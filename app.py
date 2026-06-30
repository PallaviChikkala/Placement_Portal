from flask import Flask, request, render_template, redirect, session, jsonify, flash
import pdfplumber
import docx
import mysql.connector
import os
from werkzeug.utils import secure_filename
import json
from datetime import timedelta
import pandas as pd
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# ── Email notification service (optional — works when .env is configured) ──
try:
    import email_service as _email_svc
    _EMAIL_SVC_LOADED = True
except Exception as _e:
    _EMAIL_SVC_LOADED = False
    print(f"[EmailService] Could not load email_service module: {_e}")

app = Flask(__name__)
app.secret_key = "placement_portal_secret"
app.permanent_session_lifetime = timedelta(days=30)

@app.after_request
def add_header(r):
    r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    return r


#MYSQL Connection
global_db = None
global_cursor = None

def get_connection(db_name="placement_portal"):
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="Pallavi@2007",
        database=db_name,
        connection_timeout=30,
        autocommit=False
    )

def init_global_db():
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="Pallavi@2007",
    )
    c = conn.cursor()
    
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches = json.load(f)
            for batch in batches:
                c.execute(f"CREATE DATABASE IF NOT EXISTS {batch['db']}")
    else:
        c.execute("CREATE DATABASE IF NOT EXISTS placement_portal_2025_2026")
        c.execute("CREATE DATABASE IF NOT EXISTS placement_portal_2026_2027")
        
    c.execute("CREATE DATABASE IF NOT EXISTS placement_portal2")
    conn.commit()
    conn.close()
    
    global global_db, global_cursor
    global_db = get_connection("placement_portal_2025_2026")
    global_cursor = global_db.cursor(dictionary=True)

class CursorWrapper:
    def __getattr__(self, name):
        from flask import session, has_request_context, g
        if has_request_context():
            db_name = session.get('active_year_db', 'placement_portal_2025_2026')
            if 'db' not in g:
                 g.db = get_connection(db_name)
                 g.cursor = g.db.cursor(dictionary=True)
            return getattr(g.cursor, name)
        return getattr(global_cursor, name)

class DbWrapper:
    def __getattr__(self, name):
        from flask import session, has_request_context, g
        if has_request_context():
            db_name = session.get('active_year_db', 'placement_portal_2025_2026')
            if 'db' not in g:
                 g.db = get_connection(db_name)
                 g.cursor = g.db.cursor(dictionary=True)
            return getattr(g.db, name)
        return getattr(global_db, name)

init_global_db()
cursor = CursorWrapper()
db = DbWrapper()

def switch_active_db(db_name, year_str, year_name=None):
    from flask import g, session, has_request_context
    if has_request_context():
        session["active_year_db"] = db_name
        session["active_year"] = year_str
        
        actual_name = year_name
        if not actual_name:
            try:
                import json, os
                batches_file = os.path.join(app.root_path, "database", "batches.json")
                if os.path.exists(batches_file):
                    with open(batches_file, "r") as f:
                        batches = json.load(f)
                        for b in batches:
                            if b["id"] == year_str:
                                actual_name = b["name"]
                                break
            except Exception:
                pass
                
        session["active_year_name"] = actual_name or year_str
        
        if 'db' in g:
            try:
                g.cursor.close()
            except:
                pass
            try:
                g.db.close()
            except:
                pass
            g.pop('db', None)
            g.pop('cursor', None)

@app.before_request
def check_routes_and_master_sheet():
    from flask import session, request, redirect, flash
    import os

    # Skip static file requests entirely
    if request.path.startswith("/static"):
        return None

    # 1. Enforce student Student Directory Data check on protected student pages
    if "student_id" in session:
        protected_prefixes = [
            "/student_dashboard", "/eligible_companies", "/my_applications",
            "/profile", "/change_password", "/apply_job",
            "/update_profile", "/update_profile_details",
            "/notifications", "/clear_notifications", "/upload_resume"
        ]
        is_protected = any(request.path.startswith(p) for p in protected_prefixes)
        if is_protected:
            active_year = session.get("active_year", "2025-2026")
            upload_dir = os.path.join(app.static_folder, "uploads", active_year)
            has_sheet = (
                os.path.exists(os.path.join(upload_dir, "master_sheet.xlsx")) or
                os.path.exists(os.path.join(upload_dir, "master_sheet.pdf")) or
                os.path.exists(os.path.join(upload_dir, "master_sheet.csv"))
            )
            if not has_sheet:
                session.pop("student_id", None)
                session.pop("student_name", None)
                return redirect("/student_login?error=Your+portal+is+currently+disabled.+Please+contact+faculty.")
            

    # 2. Enforce faculty login and active year database check
    is_faculty_route = (
        request.path.startswith("/faculty") or
        request.path.startswith("/api/faculty")
    )
    if is_faculty_route:
        exempt = ["/faculty_login", "/faculty_login_check", "/faculty_logout"]
        if request.path not in exempt:
            if "faculty_email" not in session:
                is_api = (request.path.startswith("/api/") or
                          request.headers.get("X-Requested-With") == "XMLHttpRequest")
                if is_api:
                    return jsonify({"error": "Unauthorized"}), 401
                return redirect("/faculty_login")

            if "active_year_db" not in session:
                year_exempt = (
                    request.path.startswith("/faculty/select_year") or
                    request.path.startswith("/faculty/set_year")
                )
                if not year_exempt:
                    is_api = (request.path.startswith("/api/") or
                              request.headers.get("X-Requested-With") == "XMLHttpRequest")
                    if is_api:
                        return jsonify({"error": "Select year required"}), 400
                    return redirect("/faculty/select_year")

@app.teardown_appcontext
def close_db(error):
    from flask import g
    if 'cursor' in g:
        g.cursor.close()
    if 'db' in g:
        g.db.close()

def ensure_connection():
    pass

# Database Tables & Mock Data Initialization
def init_database():
    global global_db, global_cursor
    
    db_names = ["placement_portal_2025_2026", "placement_portal_2026_2027"]
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches = json.load(f)
            db_names = [b["db"] for b in batches]

    for db_name in db_names:
        try:
            global_db = get_connection(db_name)
            global_cursor = global_db.cursor(dictionary=True)
            _init_database_single()
        except Exception as e:
            print(f"Failed to init {db_name}: {e}")
            
    # Restore default global connection to the first batch or fallback
    try:
        global_db = get_connection(db_names[0] if db_names else "placement_portal_2025_2026")
        global_cursor = global_db.cursor(dictionary=True)
    except Exception as e:
        print(f"Failed to restore default global DB: {e}")

def _init_database_single():
    ensure_connection()
    try:
        # ── Students table ──────────────────────────────────────────────────────
        # Students are populated exclusively via Student Directory Data uploads.
        # Do NOT insert hardcoded mock students here — every new year
        # brings a fresh set of students from the master sheet.
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                student_id    INT AUTO_INCREMENT PRIMARY KEY,
                name          VARCHAR(100),
                email         VARCHAR(150) UNIQUE,
                password      VARCHAR(100),
                branch        VARCHAR(50),
                cgpa          FLOAT DEFAULT 0,
                backlogs      INT DEFAULT 0,
                backlog_history INT DEFAULT 0,
                skills        TEXT,
                selected_tier INT DEFAULT NULL,
                batch         INT,
                roll_number   VARCHAR(50) DEFAULT NULL,
                phone_number  VARCHAR(20) DEFAULT NULL,
                aadhar        VARCHAR(20) DEFAULT NULL,
                pan           VARCHAR(20) DEFAULT NULL,
                profile_photo VARCHAR(255) DEFAULT '/static/default_avatar.png',
                must_change_password TINYINT(1) DEFAULT 1,
                profile_complete TINYINT(1) DEFAULT 0,
                tenth_score   FLOAT DEFAULT NULL,
                inter_score   FLOAT DEFAULT NULL,
                dob           DATE DEFAULT NULL,
                gender        VARCHAR(20) DEFAULT NULL,
                category      VARCHAR(20) DEFAULT NULL,
                physically_challenged VARCHAR(5) DEFAULT 'No',
                pc_percentage FLOAT DEFAULT NULL,
                internships_count INT DEFAULT 0,
                home_address  TEXT DEFAULT NULL,
                jee_rank      VARCHAR(30) DEFAULT NULL,
                academic_gap  TEXT DEFAULT NULL,
                alt_phone_number VARCHAR(20) DEFAULT NULL,
                tier_1        VARCHAR(100) DEFAULT NULL,
                tier_2        VARCHAR(100) DEFAULT NULL,
                tier_3        VARCHAR(100) DEFAULT NULL,
                career_option VARCHAR(100) DEFAULT NULL,
                personal_email VARCHAR(150) DEFAULT NULL
            )
        """)

        # Safe column additions for databases created before this schema update
        for col_sql in [
            "ALTER TABLE students ADD COLUMN backlog_history INT DEFAULT 0",
            "ALTER TABLE students ADD COLUMN profile_photo VARCHAR(255) DEFAULT '/static/default_avatar.png'",
            "ALTER TABLE students ADD COLUMN must_change_password TINYINT(1) DEFAULT 1",
            "ALTER TABLE students ADD COLUMN roll_number VARCHAR(50) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN phone_number VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN aadhar VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN pan VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN selected_tier INT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN course VARCHAR(20) DEFAULT 'B.Tech'",
            "ALTER TABLE students ADD COLUMN profile_complete TINYINT(1) DEFAULT 0",
            "ALTER TABLE students ADD COLUMN tenth_score FLOAT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN inter_score FLOAT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN dob DATE DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN gender VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN category VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN physically_challenged VARCHAR(5) DEFAULT 'No'",
            "ALTER TABLE students ADD COLUMN pc_percentage FLOAT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN internships_count INT DEFAULT 0",
            "ALTER TABLE students ADD COLUMN home_address TEXT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN jee_rank VARCHAR(30) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN academic_gap TEXT DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN alt_phone_number VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN tier_1 VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN tier_2 VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN tier_3 VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN career_option VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE students ADD COLUMN personal_email VARCHAR(150) DEFAULT NULL"
        ]:
            try:
                cursor.execute(col_sql)
            except Exception:
                pass

        # ── Faculty table ────────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS faculty (
                faculty_id INT AUTO_INCREMENT PRIMARY KEY,
                name       VARCHAR(50),
                email      VARCHAR(100) UNIQUE,
                password   VARCHAR(50)
            )
        """)

        # ── Jobs table ───────────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                job_id          VARCHAR(30) UNIQUE,
                company_name    VARCHAR(100),
                role            VARCHAR(100),
                ctc             VARCHAR(30),
                location        VARCHAR(100),
                bond            VARCHAR(50) DEFAULT 'None',
                cgpa_cutoff     DECIMAL(4,2) DEFAULT 0.0,
                active_backlogs INT DEFAULT 0,
                backlog_history INT DEFAULT 0,
                branches        TEXT,
                tier            VARCHAR(20) DEFAULT 'Tier 1',
                description     TEXT,
                req_aadhar      TINYINT(1) DEFAULT 0,
                req_pan         TINYINT(1) DEFAULT 0,
                req_other       VARCHAR(200),
                pdf_path        VARCHAR(300),
                deadline        DATETIME DEFAULT NULL,
                reminder_date   DATETIME DEFAULT NULL,
                reminder_note   TEXT DEFAULT NULL,
                reminder_sent   TINYINT(1) DEFAULT 0
            )
        """)

        for col_sql in [
            "ALTER TABLE jobs ADD COLUMN ctc VARCHAR(30)",
            "ALTER TABLE jobs ADD COLUMN location VARCHAR(100)",
            "ALTER TABLE jobs ADD COLUMN bond VARCHAR(50) DEFAULT 'None'",
            "ALTER TABLE jobs ADD COLUMN cgpa_cutoff DECIMAL(4,2) DEFAULT 0.0",
            "ALTER TABLE jobs ADD COLUMN active_backlogs INT DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN backlog_history INT DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN branches TEXT",
            "ALTER TABLE jobs ADD COLUMN req_aadhar TINYINT(1) DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN req_pan TINYINT(1) DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN req_other VARCHAR(200)",
            "ALTER TABLE jobs ADD COLUMN pdf_path VARCHAR(300)",
            "ALTER TABLE jobs ADD COLUMN deadline DATETIME DEFAULT NULL",
            "ALTER TABLE jobs ADD COLUMN reminder_date DATETIME DEFAULT NULL",
            "ALTER TABLE jobs ADD COLUMN reminder_note TEXT DEFAULT NULL",
            "ALTER TABLE jobs ADD COLUMN reminder_sent TINYINT(1) DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN deadline_dismissed TINYINT(1) DEFAULT 0",
            "ALTER TABLE jobs ADD COLUMN recruitment_finished_at DATETIME DEFAULT NULL",
            "ALTER TABLE jobs ADD COLUMN recruitment_archived TINYINT(1) DEFAULT 0"
        ]:
            try:
                cursor.execute(col_sql)
            except Exception:
                pass

        # ── Applications table ───────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS applications (
                application_id     INT AUTO_INCREMENT PRIMARY KEY,
                student_id         INT,
                job_id             VARCHAR(30),
                resume_path        TEXT,
                status             VARCHAR(50),
                applied_date       DATE,
                extra_details      TEXT,
                drive_link         TEXT DEFAULT NULL,
                status_updated_at  DATETIME DEFAULT NULL
            )
        """)

        for col_sql in [
            "ALTER TABLE applications ADD COLUMN extra_details TEXT",
            "ALTER TABLE applications MODIFY COLUMN job_id VARCHAR(30)",
            "ALTER TABLE applications ADD COLUMN drive_link TEXT DEFAULT NULL",
            "ALTER TABLE applications ADD COLUMN status_updated_at DATETIME DEFAULT NULL"
        ]:
            try:
                cursor.execute(col_sql)
            except Exception:
                pass

        # ── Recruitment rounds ───────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS recruitment_rounds (
                id         INT AUTO_INCREMENT PRIMARY KEY,
                job_id     VARCHAR(30),
                num_rounds INT DEFAULT 1,
                UNIQUE KEY unique_job (job_id)
            )
        """)

        # ── Round results ────────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS round_results (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                job_id       VARCHAR(30),
                student_id   INT,
                round_number INT,
                result       VARCHAR(20) DEFAULT 'Pending',
                drive_link   TEXT DEFAULT NULL,
                UNIQUE KEY unique_round_result (job_id, student_id, round_number)
            )
        """)

        try:
            cursor.execute("ALTER TABLE round_results ADD COLUMN drive_link TEXT DEFAULT NULL")
        except Exception:
            pass

        # ── Notifications ────────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                notif_id   INT AUTO_INCREMENT PRIMARY KEY,
                student_id INT,
                message    TEXT,
                link       TEXT,
                is_read    BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # ── Global settings ──────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS global_settings (
                setting_key   VARCHAR(50) PRIMARY KEY,
                setting_value VARCHAR(255)
            )
        """)
        cursor.execute("INSERT IGNORE INTO global_settings (setting_key, setting_value) VALUES ('recruitment_year', '2025')")
        cursor.execute("INSERT IGNORE INTO global_settings (setting_key, setting_value) VALUES ('email_notifications_enabled', 'false')")

        # ── Email logs table ─────────────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS email_logs (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                recipient_email VARCHAR(150) NOT NULL,
                student_name    VARCHAR(100),
                event_type      VARCHAR(100) NOT NULL,
                subject         VARCHAR(255),
                sent_at         TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status          ENUM('sent','failed') DEFAULT 'sent',
                error_message   TEXT DEFAULT NULL,
                INDEX idx_event (event_type),
                INDEX idx_sent_at (sent_at)
            )
        """)

        # ── Faculty seed account ─────────────────────────────────────────────────
        # Insert Dr. Shankar only if not already present.
        # Students are intentionally NOT seeded here —
        # they arrive exclusively via Student Directory Data uploads each year.
        cursor.execute("SELECT COUNT(*) as count FROM faculty")
        if cursor.fetchone()["count"] == 0:
            cursor.execute("""
               INSERT INTO faculty (faculty_id, name, email, password)
VALUES
(1, 'Placement Officer', 'tap@nitandhra.ac.in', 'placementOfficerNITandhra2015'),
(2, 'Placement Officer', 'tapc@nitandhra.ac.in', 'placementOfficerNITandhra2015');
            """)

        # ── Internship postings table ─────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS internship_postings (
                posting_id   INT AUTO_INCREMENT PRIMARY KEY,
                company_name VARCHAR(100) NOT NULL,
                role         VARCHAR(100) NOT NULL,
                details      TEXT,
                link         TEXT,
                posted_date  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # ── Student internships table ─────────────────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS student_internships (
                id                     INT AUTO_INCREMENT PRIMARY KEY,
                student_id             INT NOT NULL,
                posting_id             INT DEFAULT NULL,
                status                 VARCHAR(50) DEFAULT 'Interested',
                completion_description TEXT DEFAULT NULL,
                certificate_path       VARCHAR(255) DEFAULT NULL,
                submitted_at           DATETIME DEFAULT NULL,
                ext_company_name       VARCHAR(200) DEFAULT NULL,
                ext_role               VARCHAR(200) DEFAULT NULL,
                ext_duration           VARCHAR(100) DEFAULT NULL,
                is_external            TINYINT(1) DEFAULT 0,
                FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE
            )
        """)
        # Safe additions for older databases
        for _col_sql in [
            "ALTER TABLE student_internships ADD COLUMN ext_company_name VARCHAR(200) DEFAULT NULL",
            "ALTER TABLE student_internships ADD COLUMN ext_role VARCHAR(200) DEFAULT NULL",
            "ALTER TABLE student_internships ADD COLUMN ext_duration VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE student_internships ADD COLUMN is_external TINYINT(1) DEFAULT 0",
            "ALTER TABLE student_internships ADD COLUMN hr_name VARCHAR(200) DEFAULT NULL",
            "ALTER TABLE student_internships ADD COLUMN hr_contact VARCHAR(200) DEFAULT NULL",
        ]:
            try:
                cursor.execute(_col_sql)
                db.commit()
            except Exception:
                db.rollback()

        db.commit()
        print("Database tables initialized successfully.")
    except Exception as e:
        print("Warning: Database initialization failed. Details:", e)

init_database()

def normalize_branch(branch_name):
    if not branch_name: return ""
    b = branch_name.lower().strip()
    
    # 1. Computer Science / CSE / CS / Computer Engineering
    if "computer science" in b or "cse" in b or b == "cs" or "computer engineering" in b or "comp sci" in b:
        return "CSE"
        
    # 2. Information Technology / IT
    if "information technology" in b or b == "it" or "infotech" in b:
        return "IT"
        
    # 3. AI / ML / Data Science / CSDA
    if "artificial intelligence" in b or "machine learning" in b or "aiml" in b or "ai & ds" in b or "ai&ds" in b or "aids" in b or "data science" in b or "data analytics" in b or "csda" in b:
        return "AI"
        
    # 4. Electrical / EEE (Checked before ECE so "electrical and electronics" is mapped to EEE)
    if "electrical" in b or "eee" in b or b == "ee" or "power electronics" in b:
        return "EEE"
        
    # 5. Electronics & Communication / ECE / ACS
    if "electronics" in b or "ece" in b or b == "ec" or "communication" in b or "acs" in b:
        return "ECE"
        
    # 6. Mechanical / MECH / Thermal / Manufacturing
    if "mechanical" in b or "mech" in b or b == "me" or "manufacturing" in b or "thermal" in b:
        return "MECH"
        
    # 7. Civil / CIVIL / Geotechnical
    if "civil" in b or b == "ce" or "geotechnical" in b or "geo" in b:
        return "CIVIL"
        
    # 8. Chemical / CHEMICAL / Chem
    if "chemical" in b or "chem" in b:
        return "CHEMICAL"
        
    # 9. Biotech / BIOTECH / Bioprocess
    if "biotech" in b or "bio-technology" in b or "bioprocess" in b or b == "bt":
        return "BIOTECH"
        
    # 10. Metallurgy / METALLURGY / Materials
    if "metallurg" in b or "material" in b or "mme" in b:
        return "METALLURGY"
        
    return b.upper()

def notify_students_new_job(company_name, role):
    """
    Helper function to notify all students when a new job is posted.
    Call this function from the faculty job post route.
    """
    try:
        message = f"New Job Posted: {company_name} is hiring for {role}!"
        link = "/eligible_companies"
       
        # Insert a notification for every student in the database
        cursor.execute("""
            INSERT INTO notifications (student_id, message, link)
            SELECT student_id, %s, %s FROM students
        """, (message, link))
        db.commit()
        print(f"Successfully notified students about {company_name} - {role}")
    except Exception as e:
        print("Failed to notify students:", e)


def _is_email_notifications_enabled() -> bool:
    """Check if the global email notifications toggle is ON in the DB."""
    try:
        cursor.execute("SELECT setting_value FROM global_settings WHERE setting_key='email_notifications_enabled'")
        row = cursor.fetchone()
        return row and row['setting_value'] == 'true'
    except Exception:
        return False


def _log_email(recipient_email, student_name, event_type, subject, status, error_msg=None, db_name=None):
    """Write a row to email_logs. Safe to call from background threads — does NOT use Flask session."""
    try:
        # Resolve db_name without touching Flask session (safe in threads)
        if not db_name:
            db_name = 'placement_portal_2025_2026'
            try:
                from flask import has_request_context
                if has_request_context():
                    db_name = session.get('active_year_db', db_name)
            except Exception:
                pass
        conn2 = get_connection(db_name)
        c2 = conn2.cursor()
        c2.execute("""
            INSERT INTO email_logs (recipient_email, student_name, event_type, subject, status, error_message)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (recipient_email, student_name, event_type, subject, status, error_msg))
        conn2.commit()
        c2.close()
        conn2.close()
    except Exception as ex:
        print(f"[EmailLog] Could not write log: {ex}")



def _get_eligible_students_for_job(job_cgpa, job_backlogs, job_branches, job_tier):
    """
    Query students who meet the eligibility criteria for a given job.
    Returns list of dicts: {student_id, name, email}
    """
    try:
        # Build branch filter: branches is a comma-separated string
        branch_list = [b.strip().upper() for b in job_branches.split(',') if b.strip()] if job_branches else []

        cursor.execute("""
            SELECT student_id, name, email, cgpa, backlogs, branch, selected_tier
            FROM students
            WHERE email IS NOT NULL AND email != ''
        """)
        all_students = cursor.fetchall()

        eligible = []
        for s in all_students:
            # CGPA check
            if float(s.get('cgpa') or 0) < float(job_cgpa or 0):
                continue
            # Backlogs check
            if int(s.get('backlogs') or 0) > int(job_backlogs or 0):
                continue
            # Branch check
            if branch_list:
                norm_branch = normalize_branch(str(s.get('branch') or ''))
                if not any(b in norm_branch for b in branch_list) and not any(norm_branch in b for b in branch_list):
                    continue
            # Tier restriction (skip if student already placed at same/higher tier)
            tier_num = 1 if '1' in str(job_tier) else (2 if '2' in str(job_tier) else 3)
            student_tier = s.get('selected_tier')
            if student_tier is not None and int(student_tier) <= tier_num:
                continue

            eligible.append({
                'student_id': s['student_id'],
                'name': s['name'],
                'email': s['email']
            })

        return eligible
    except Exception as ex:
        print(f"[EmailService] Could not fetch eligible students: {ex}")
        return []


def _send_new_job_emails_async(job_id, company, role, ctc, tier, deadline, location, min_cgpa, branches, subject_override=None):
    """Fetch eligible students and send new-job emails asynchronously."""
    if not _EMAIL_SVC_LOADED or not _email_svc.is_configured():
        print("[EmailService] SMTP not configured, skipping email send.")
        return

    # Read db_name from session HERE (in the request context) before the thread starts,
    # because Flask session is not accessible inside a daemon thread.
    db_name = session.get('active_year_db', 'placement_portal_2025_2026')
    subject = subject_override or f"New Placement Drive: {company} – {role}"

    try:
        recipients = _get_eligible_students_for_job(min_cgpa, 0, branches, tier)
    except Exception as ex:
        print(f"[EmailService] Error fetching recipients: {ex}")
        return

    if not recipients:
        print("[EmailService] No eligible recipients found.")
        return

    def html_fn(rec):
        return _email_svc.build_new_job_email(
            rec, company, role, ctc, tier, deadline, location, min_cgpa, branches
        )

    def log_fn(sid, email, status, err):
        _log_email(email, None, 'new_job', subject, status, err)

    _email_svc.send_bulk_emails_async(recipients, subject, html_fn, log_fn, event_type='new_job')
    print(f"[EmailService] Triggered bulk email for {len(recipients)} eligible students for {company}")

@app.route("/")
def home():
    stats_file = os.path.join(app.root_path, "database", "yearly_statistics.json")
    past_stats = []
    if os.path.exists(stats_file):
        try:
            import json as _json
            with open(stats_file, "r") as f:
                past_stats = _json.load(f)
                # Sort by year descending
                past_stats.sort(key=lambda x: x.get('year', ''), reverse=True)
        except Exception as e:
            print("Error loading past stats:", e)
            
    # Load homepage updates
    updates_file = os.path.join(app.root_path, "database", "homepage_updates.json")
    updates = []
    if os.path.exists(updates_file):
        try:
            import json as _json
            with open(updates_file, "r") as f:
                updates = _json.load(f)
                # Sort descending by id (timestamp)
                updates.sort(key=lambda x: x.get('id', ''), reverse=True)
        except Exception as e:
            print("Error loading homepage updates:", e)
            
    return render_template("index.html", past_stats=past_stats, updates=updates)

@app.route("/resume_analyzer")
def resume_analyzer():
    return render_template("resume_analyzer.html")

@app.route("/upload",methods = ["POST"])
def upload():
    file = request.files.get("resume")
    filename = file.filename.lower() if file else ""
    role = request.form.get("role")
    custom_jd = request.form.get("custom_jd", "")
   
    if not role or role == "Select a job profile...":
        return render_template("resume_analysis_result.html", error_msg="Please select a target job role.")
   
    text = ""
   
    if filename.endswith(".pdf"):
        try:
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() or ""
        except Exception:
            return render_template("resume_analysis_result.html", error_msg="The uploaded resume file appears to be corrupted or is not a valid PDF. Please upload a valid PDF file and try again.")

    elif filename.endswith(".docx"):
        document = docx.Document(file)
        for para in document.paragraphs:
            text += para.text + "\n"
   
    else :
        return render_template("resume_analysis_result.html", error_msg="Only PDF and DOCX files are allowed for resumes.")
   
    # Skill aliases for smarter matching (handles abbreviations and alternate names)
    skill_aliases = {
        "JavaScript": ["javascript", "js", "es6", "ecmascript"],
        "TypeScript": ["typescript", "ts"],
        "HTML": ["html", "html5"],
        "CSS": ["css", "css3", "scss", "sass"],
        "React": ["react", "reactjs", "react.js"],
        "Angular": ["angular", "angularjs"],
        "Vue.js": ["vue", "vuejs", "vue.js"],
        "Node.js": ["node", "nodejs", "node.js"],
        "Spring Boot": ["spring boot", "springboot", "spring"],
        "MySQL": ["mysql", "sql"],
        "MongoDB": ["mongodb", "mongo"],
        "PostgreSQL": ["postgresql", "postgres"],
        "SQL": ["sql", "mysql", "postgresql", "sqlite"],
        "Python": ["python", "python3"],
        "Java": ["java"],
        "C++": ["c++", "cpp"],
        "Machine Learning": ["machine learning", "ml", "machine-learning"],
        "Deep Learning": ["deep learning", "dl", "deep-learning"],
        "TensorFlow": ["tensorflow", "tf"],
        "PyTorch": ["pytorch", "torch"],
        "Pandas": ["pandas"],
        "NumPy": ["numpy", "np"],
        "Scikit-learn": ["scikit-learn", "sklearn", "scikit learn"],
        "Power BI": ["power bi", "powerbi"],
        "Tableau": ["tableau"],
        "Excel": ["excel", "ms excel", "microsoft excel"],
        "Git": ["git", "github", "gitlab"],
        "Docker": ["docker"],
        "Kubernetes": ["kubernetes", "k8s"],
        "REST API": ["rest", "rest api", "restful", "api"],
        "DBMS": ["dbms", "database management"],
        "Hibernate": ["hibernate"],
        "Mathematics": ["mathematics", "math", "statistics", "linear algebra", "calculus"],
        "Statistics": ["statistics", "statistical"],
        "Natural Language Processing": ["nlp", "natural language processing"],
        "Flask": ["flask"],
        "Data Structures": ["data structures", "dsa", "algorithms"],
        "Algorithms": ["algorithms", "dsa", "data structures"],
        "Spacy": ["spacy"],
        "NLTK": ["nltk"],
        "Regex": ["regex", "regular expression"],
        "Text Parsing": ["text parsing", "parsing"],
        "Bootstrap": ["bootstrap"],
        "Tailwind CSS": ["tailwind", "tailwindcss"],
    }

    def skill_in_text(skill, text_lower):
        """Check if a skill appears in text using aliases."""
        aliases = skill_aliases.get(skill, [skill.lower()])
        return any(alias in text_lower for alias in aliases)

    role_skills = {
        "Frontend Developer": {
            "required": ["HTML", "CSS", "JavaScript"],
            "optional": ["React", "Angular", "Vue.js", "Bootstrap", "Tailwind CSS", "Git", "TypeScript"]
        },
        "Backend Developer": {
            "required": ["Java", "MySQL", "DBMS"],
            "optional": ["Spring Boot", "REST API", "Hibernate", "Git", "Docker", "Python"]
        },
        "Full Stack Developer": {
            "required": ["HTML", "CSS", "JavaScript", "Java", "MySQL"],
            "optional": ["React", "Node.js", "Spring Boot", "MongoDB", "Git", "REST API"]
        },
        "Data Analyst": {
            "required": ["Python", "SQL", "Excel"],
            "optional": ["Power BI", "Tableau", "Pandas", "NumPy", "Statistics"]
        },
        "AI/ML Engineer": {
            "required": ["Python", "Machine Learning", "Mathematics"],
            "optional": ["Pandas", "NumPy", "Scikit-learn", "TensorFlow", "PyTorch", "Deep Learning"]
        },
        "Resume Analyzer": {
            "required": ["Python", "Natural Language Processing", "Regex", "Text Parsing"],
            "optional": ["Machine Learning", "Spacy", "NLTK", "Flask"]
        }
    }
   
    text_lower = text.lower()

    if role == "Custom":
        custom_text = custom_jd
        all_possible_skills = list(skill_aliases.keys())
        required_skills = [s for s in all_possible_skills if skill_in_text(s, custom_jd.lower())]
        optional_skills = []
        if not required_skills:
            required_skills = ["Problem Solving", "Communication"]
    else:
        selected_role = role_skills.get(role, {"required": ["Java", "Python", "SQL"], "optional": []})
        required_skills = selected_role["required"]
        optional_skills = selected_role["optional"]

    found_required = []
    missing_required = []
    for skill in required_skills:
        if skill_in_text(skill, text_lower):
            found_required.append(skill)
        else:
            missing_required.append(skill)

    missing_optional = []
    found_optional = []
    for skill in optional_skills:
        if skill_in_text(skill, text_lower):
            found_optional.append(skill)
        else:
            missing_optional.append(skill)

    # --- Keyword Scoring (50 pts max) ---
    if len(required_skills) > 0:
        required_score = (len(found_required) / len(required_skills)) * 40
    else:
        required_score = 40

    if len(optional_skills) > 0:
        optional_score = (len(found_optional) / len(optional_skills)) * 10
    else:
        optional_score = 10 if found_optional else 0

    keyword_score = required_score + optional_score

    # --- AI Semantic Scoring using TF-IDF (40 pts max) ---
    def clean_text(t):
        return re.sub(r'[^a-zA-Z0-9\s]', '', t).lower()
        
    cleaned_resume = clean_text(text)
    if role == "Custom":
        target_doc = clean_text(custom_jd)
    else:
        target_doc = clean_text(" ".join(required_skills + optional_skills))

    ai_score = 0
    if cleaned_resume.strip() and target_doc.strip():
        try:
            vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1,2))
            tfidf_matrix = vectorizer.fit_transform([cleaned_resume, target_doc])
            similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
            ai_score = min(int(similarity * 100 * 2.5), 40)
        except Exception:
            pass

    score = int(keyword_score + ai_score)

    # --- Profile Bonus Scoring (up to 20 pts) ---
    genome_score = 0

    if "project" in text_lower or "projects" in text_lower or "github.com" in text_lower:
        genome_score += 6
    if "b.tech" in text_lower or "bachelor" in text_lower or "education" in text_lower or "university" in text_lower or "college" in text_lower:
        genome_score += 4
    if "certificate" in text_lower or "certification" in text_lower or "certified" in text_lower:
        genome_score += 4
    if "internship" in text_lower or "intern" in text_lower or "experience" in text_lower:
        genome_score += 4
    if "achievement" in text_lower or "award" in text_lower or "hackathon" in text_lower or "competition" in text_lower:
        genome_score += 2
   
    final_score = min(score + genome_score, 100)

    # --- Rich AI Suggestion based on analysis ---
    if final_score >= 85:
        suggestion = "Excellent match! Your resume is strongly aligned. Ensure formatting is clean, quantify achievements (e.g. 'Improved performance by 30%'), and tailor your summary to this exact role."
    elif final_score >= 70:
        top_missing = missing_required[:2] + missing_optional[:2]
        suggestion = f"Good match. Boost your score by adding these skills: {', '.join(top_missing)}. Quantify your project outcomes and add a concise professional summary."
    elif final_score >= 55:
        suggestion = f"Moderate match. Focus on adding the missing required skills: {', '.join(missing_required[:4])}. Consider building small projects to demonstrate these and link your GitHub."
    else:
        suggestion = f"Low match. Your resume needs significant tailoring. Priority skills to add: {', '.join(required_skills[:4])}. Study these, build projects, and highlight relevant coursework."

    # --- Update student profile if logged in ---
    if "student_id" in session:
        student_id = session["student_id"]
        session["resume_score"] = final_score
       
        found_skills = list(set(found_required + found_optional))
        if found_skills:
            cursor.execute("SELECT skills FROM students WHERE student_id = %s", (student_id,))
            current_skills_row = cursor.fetchone()
            existing_skills = [s.strip() for s in current_skills_row["skills"].split(",")] if current_skills_row and current_skills_row["skills"] else []
            combined_skills = list(set(existing_skills + found_skills))
            skills_str = ", ".join(combined_skills)
            cursor.execute("UPDATE students SET skills = %s WHERE student_id = %s", (skills_str, student_id))
            db.commit()

    found_skills_list = found_required + found_optional
    missing_skills_list = missing_required + missing_optional

    return render_template("resume_analysis_result.html", 
                           score=final_score, 
                           matched_skills=found_skills_list, 
                           missing_skills=missing_skills_list, 
                           suggestion=suggestion)

@app.route("/student_login")
def student_login_page():
    # Clear any residual temporary session state
    session.pop("resume_score", None)
    # Always show the login form - never auto-redirect (user must explicitly log in)
    error = request.args.get('error')
    success = request.args.get('success')
    return render_template("student/login.html", error=error, success=success)

@app.route("/student_login_check", methods = ["POST"])
def student_login_check():
    email = request.form["email"]
    password = request.form["password"]
    
    # 1. Try 2025-2026
    switch_active_db("placement_portal_2025_2026", "2025-2026")
    ensure_connection()
    query = "SELECT * FROM students WHERE email = %s AND password = %s"
    cursor.execute(query, (email,password))
    student = cursor.fetchone()

    if not student:
        # 2. Try 2026-2027
        switch_active_db("placement_portal_2026_2027", "2026-2027")
        ensure_connection()
        cursor.execute(query, (email,password))
        student = cursor.fetchone()

    if student:
        # Check if Student Directory Data exists
        upload_dir = os.path.join(app.static_folder, "uploads", session["active_year"])
        has_xlsx = os.path.exists(os.path.join(upload_dir, "master_sheet.xlsx"))
        has_pdf = os.path.exists(os.path.join(upload_dir, "master_sheet.pdf"))
        if not (has_xlsx or has_pdf):
            session.pop("student_id", None)
            session.pop("student_name", None)
            return render_template("student/login.html", error="Your portal is currently disabled (No Master Sheet uploaded by faculty).")

        session["student_id"] = student["student_id"]
        session["student_name"] = student["name"]
        session["must_change_password"] = student.get("must_change_password", 1)
        session.permanent = True
        return redirect("/student_dashboard")
    else :
        session.pop("student_id", None)
        session.pop("student_name", None)
        return render_template("student/login.html", error="Wrong password or invalid credentials.")

# ---- FORGOT PASSWORD FLOW ----
import random
import time

@app.route("/student_forgot_password", methods=["GET", "POST"])
def student_forgot_password():
    if request.method == "GET":
        error = request.args.get('error')
        return render_template("student/forgot_password_email.html", error=error)
    
    email = request.form.get("email", "").strip()
    
    found_db = None
    found_year_str = None
    
    dbs_to_check = [("placement_portal_2025_2026", "2025-2026"), ("placement_portal_2026_2027", "2026-2027")]
    for db_name, year_str in dbs_to_check:
        switch_active_db(db_name, year_str)
        ensure_connection()
        cursor.execute("SELECT * FROM students WHERE email = %s", (email,))
        student = cursor.fetchone()
        if student:
            found_db = db_name
            found_year_str = year_str
            break
            
    if not found_db:
        return render_template("student/forgot_password_email.html", error="Email not found in our records.")
        
    session["reset_email"] = email
    session["reset_db_name"] = found_db
    session["reset_year_str"] = found_year_str
    
    return redirect("/student_forgot_password/verify_identity")

@app.route("/student_forgot_password/verify_identity", methods=["GET", "POST"])
def student_forgot_identity():
    if "reset_email" not in session:
        return redirect("/student_forgot_password")
        
    if request.method == "GET":
        error = request.args.get('error')
        return render_template("student/forgot_password_identity.html", error=error)
        
    aadhar_input = request.form.get("aadhar", "").strip()
    pan_input = request.form.get("pan", "").strip().upper()
    
    switch_active_db(session["reset_db_name"], session["reset_year_str"])
    ensure_connection()
    cursor.execute("SELECT aadhar, pan FROM students WHERE email = %s", (session["reset_email"],))
    student = cursor.fetchone()
    
    if not student or not student.get("aadhar") or not student.get("pan"):
        return render_template("student/forgot_password_identity.html", error="Aadhar or PAN not registered for this account.")
        
    db_aadhar = student["aadhar"].strip()
    db_pan = student["pan"].strip().upper()
    
    if len(db_aadhar) >= 5 and db_aadhar[-5:] == aadhar_input and db_pan == pan_input:
        otp = str(random.randint(100000, 999999))
        session["reset_otp"] = otp
        session["reset_otp_expiry"] = time.time() + 180
        
        html_body = f"""
        <div style="font-family: Arial, sans-serif; padding: 20px;">
            <h2>Password Reset OTP</h2>
            <p>Your One-Time Password for resetting your placement portal password is:</p>
            <h1 style="color: #d97706; font-size: 36px; letter-spacing: 5px;">{otp}</h1>
            <p>This OTP is valid for exactly 3 minutes. Do not share it with anyone.</p>
        </div>
        """
        from email_service import send_email
        send_email(session["reset_email"], "Password Reset OTP", html_body)
        
        return redirect("/student_forgot_password/otp")
    else:
        return render_template("student/forgot_password_identity.html", error="Verification failed. Details do not match.")

@app.route("/student_forgot_password/otp", methods=["GET", "POST"])
def student_forgot_otp():
    if "reset_email" not in session or "reset_otp" not in session:
        return redirect("/student_forgot_password")
        
    if request.method == "GET":
        error = request.args.get('error')
        time_remaining = max(0, int(session.get("reset_otp_expiry", 0) - time.time()))
        return render_template("student/forgot_password_otp.html", error=error, time_remaining=time_remaining)
        
    otp_input = request.form.get("otp", "").strip()
    
    if time.time() > session.get("reset_otp_expiry", 0):
        return render_template("student/forgot_password_otp.html", error="OTP has expired. Please click Resend OTP.", time_remaining=0)
        
    if otp_input == session["reset_otp"]:
        session["reset_otp_verified"] = True
        return redirect("/student_forgot_password/reset")
    else:
        time_remaining = max(0, int(session.get("reset_otp_expiry", 0) - time.time()))
        return render_template("student/forgot_password_otp.html", error="Invalid OTP. Try again.", time_remaining=time_remaining)

@app.route("/student_forgot_password/resend_otp", methods=["POST"])
def student_resend_otp():
    if "reset_email" not in session:
        return redirect("/student_forgot_password")
        
    otp = str(random.randint(100000, 999999))
    session["reset_otp"] = otp
    session["reset_otp_expiry"] = time.time() + 180
    
    html_body = f"""
    <div style="font-family: Arial, sans-serif; padding: 20px;">
        <h2>Password Reset OTP (Resent)</h2>
        <p>Your new One-Time Password is:</p>
        <h1 style="color: #d97706; font-size: 36px; letter-spacing: 5px;">{otp}</h1>
        <p>This OTP is valid for exactly 3 minutes.</p>
    </div>
    """
    from email_service import send_email
    send_email(session["reset_email"], "Password Reset OTP", html_body)
    
    return redirect("/student_forgot_password/otp")

@app.route("/student_forgot_password/reset", methods=["GET", "POST"])
def student_forgot_reset():
    if not session.get("reset_otp_verified"):
        return redirect("/student_forgot_password")
        
    if request.method == "GET":
        return render_template("student/forgot_password_reset.html")
        
    password = request.form.get("password")
    confirm = request.form.get("confirm_password")
    
    if password != confirm:
        return render_template("student/forgot_password_reset.html", error="Passwords do not match.")
        
    if len(password) < 6:
        return render_template("student/forgot_password_reset.html", error="Password must be at least 6 characters.")
        
    switch_active_db(session["reset_db_name"], session["reset_year_str"])
    ensure_connection()
    cursor.execute("UPDATE students SET password = %s, must_change_password = 0 WHERE email = %s", 
                   (password, session["reset_email"]))
    db.commit()
    
    session.pop("reset_email", None)
    session.pop("reset_db_name", None)
    session.pop("reset_year_str", None)
    session.pop("reset_otp", None)
    session.pop("reset_otp_expiry", None)
    session.pop("reset_otp_verified", None)
    
    return redirect("/student_login?success=Password+reset+successfully")
# ---- END FORGOT PASSWORD FLOW ----

@app.route("/google_login_check", methods=["POST"])
def google_login_check():
    import base64
    credential = request.form.get("credential")
    if not credential:
        return render_template("student/login.html", error="Google Sign-In failed.")
       
    try:
        parts = credential.split(".")
        if len(parts) != 3:
            raise ValueError("Invalid JWT format")
           
        payload = parts[1]
        payload += "=" * ((4 - len(payload) % 4) % 4)
        decoded_payload = base64.urlsafe_b64decode(payload).decode('utf-8')
        user_info = json.loads(decoded_payload)
       
        email = user_info.get("email")
        if not email:
            raise ValueError("Email not found in Google token")
           
        # 1. Try 2025-2026
        switch_active_db("placement_portal_2025_2026", "2025-2026")
        ensure_connection()
        cursor.execute("SELECT * FROM students WHERE email = %s", (email,))
        student = cursor.fetchone()
       
        if not student:
            # 2. Try 2026-2027
            switch_active_db("placement_portal_2026_2027", "2026-2027")
            ensure_connection()
            cursor.execute("SELECT * FROM students WHERE email = %s", (email,))
            student = cursor.fetchone()

        if student:
            # Check if Student Directory Data exists
            upload_dir = os.path.join(app.static_folder, "uploads", session["active_year"])
            has_xlsx = os.path.exists(os.path.join(upload_dir, "master_sheet.xlsx"))
            has_pdf = os.path.exists(os.path.join(upload_dir, "master_sheet.pdf"))
            if not (has_xlsx or has_pdf):
                session.pop("student_id", None)
                session.pop("student_name", None)
                return render_template("student/login.html", error="Your portal is currently disabled (No Master Sheet uploaded by faculty).")

            session["student_id"] = student["student_id"]
            session["student_name"] = student["name"]
            session.permanent = True
            return redirect("/student_dashboard")
        else:
            session.pop("student_id", None)
            session.pop("student_name", None)
            return render_template("student/login.html", error=f"Email {email} is not registered. Please contact faculty.")
           
    except Exception as e:
        print("Google Auth Error:", e)
        session.pop("student_id", None)
        session.pop("student_name", None)
        return render_template("student/login.html", error="Google Sign-In verification failed.")
   
@app.route("/change_password", methods=["POST"])
def change_password():
    if "student_id" not in session:
        return redirect("/student_login")

    old_password = request.form.get("old_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if new_password != confirm_password:
        flash("New password and confirm password do not match", "danger")
        return redirect("/student_profile")

    cursor.execute(
        "SELECT * FROM students WHERE student_id=%s AND password=%s",
        (session["student_id"], old_password)
    )
    student = cursor.fetchone()

    if student:
        cursor.execute(
    "UPDATE students SET password=%s, must_change_password=0 WHERE student_id=%s",
    (new_password, session["student_id"])
        )
        db.commit()
        session["must_change_password"] = 0
        flash("Password changed successfully", "success")
    else:
        flash("Old password is incorrect", "danger")

    return redirect("/student_profile")

@app.route("/api/faculty_login", methods=["POST"])
def api_faculty_login():
    data = request.get_json()
    email = data.get("email")
    password = data.get("password")
   
    cursor.execute("SELECT * FROM faculty WHERE email = %s AND password = %s", (email, password))
    faculty = cursor.fetchone()
   
    if faculty:
        session["faculty_id"] = faculty["faculty_id"]
        session["faculty_name"] = faculty["name"]
        return jsonify({"success": True, "name": faculty["name"]})
    else:
        return jsonify({"success": False, "message": "Invalid email or password"})

@app.route("/clear_notifications", methods=["POST"])
def clear_notifications():
    if "student_id" in session:
        try:
            student_id = session["student_id"]
            cursor.execute("UPDATE notifications SET is_read = 1 WHERE student_id = %s", (student_id,))
            db.commit()
            return jsonify({"success": True})
        except Exception as e:
            print("Error clearing notifications:", e)
            return jsonify({"success": False, "error": str(e)})
    return jsonify({"success": False, "error": "Not logged in"})

@app.route("/api/upcoming_deadlines")
def api_upcoming_deadlines():
    """Returns upcoming job deadlines for the calendar widget."""
    ensure_connection()
    if "student_id" not in session:
        return jsonify({"deadlines": []})
    try:
        from datetime import datetime, timedelta
        now = datetime.now()
        future = now + timedelta(days=30)
        cursor.execute("""
            SELECT company_name, deadline FROM jobs
            WHERE deadline IS NOT NULL
              AND deadline >= %s
              AND deadline <= %s
            ORDER BY deadline ASC
            LIMIT 5
        """, (now.strftime("%Y-%m-%d"), future.strftime("%Y-%m-%d")))
        rows = cursor.fetchall()
        deadlines = []
        for r in rows:
            try:
                d = r["deadline"]
                if hasattr(d, "strftime"):
                    date_str = d.strftime("%b %d")
                else:
                    date_str = str(d)[:10]
                deadlines.append({"company": r["company_name"], "date": date_str})
            except Exception:
                pass
        return jsonify({"deadlines": deadlines})
    except Exception as e:
        return jsonify({"deadlines": [], "error": str(e)})

@app.route("/student_dashboard")
def student_dashboard():
    ensure_connection()

    if "student_id" not in session:
        return redirect("/student_login")

    cursor.execute("SELECT * FROM students WHERE student_id = %s", (session["student_id"],))
    student = cursor.fetchone()

    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")

    try:
        cursor.execute("SELECT * FROM jobs WHERE deadline IS NULL OR deadline >= NOW() ORDER BY id DESC")
    except Exception:
        cursor.execute("SELECT * FROM jobs WHERE deadline IS NULL OR deadline >= NOW() ORDER BY job_id DESC")
    all_jobs = cursor.fetchall()

    cursor.execute("SELECT job_id, status FROM applications WHERE student_id = %s", (session["student_id"],))
    apps = cursor.fetchall()
    app_status_map = {a["job_id"]: a["status"] for a in apps}

    upcoming_drives = []

    for job in all_jobs:
        job_tier_str = str(job.get("tier", "Tier 3")).lower()
        job_tier_num = 1 if "1" in job_tier_str else (2 if "2" in job_tier_str else 3)

        eligible_branches = [normalize_branch(b) for b in job.get("branches", "").split(",")] if job.get("branches") else []
        student_branch = normalize_branch(student["branch"]) if student["branch"] else ""

        cgpa_ok = float(student.get("cgpa") or 0) >= float(job.get("cgpa_cutoff") or 0)
        
        s_backlogs = int(student.get("backlogs") or 0)
        s_history = int(student.get("backlog_history") or 0)
        j_backlogs = int(job.get("active_backlogs") or 0)
        j_history = int(job.get("backlog_history") or 0)
        
        backlogs_ok = s_backlogs <= j_backlogs
        backlog_history_ok = s_history <= j_history
        
        branch_ok = student_branch in eligible_branches or not eligible_branches

        student_selected_tier = student.get("selected_tier")
        
        tier_1_val = student.get("tier_1")
        tier_2_val = student.get("tier_2")
        tier_3_val = student.get("tier_3")
        offers_count = sum(1 for t in [tier_1_val, tier_2_val, tier_3_val] if t and str(t).strip())
        
        tier_ok = True
        
        if offers_count >= 2:
            tier_ok = False
        elif student_selected_tier is not None and student_selected_tier > 0:
            if student_selected_tier == 1:
                tier_ok = False
            elif student_selected_tier == 2 and job_tier_num in [2, 3]:
                tier_ok = False
            elif student_selected_tier == 3 and job_tier_num == 3:
                tier_ok = False

        pwd_only = bool(job.get('pwd_only', False))
        is_pwd_student = (str(student.get('physically_challenged', '')).strip().lower() in ['yes', 'y', 'true', '1'])
        if pwd_only and not is_pwd_student:
            continue

        is_eligible = cgpa_ok and backlogs_ok and backlog_history_ok and branch_ok and tier_ok

        job_item = dict(job)
        job_item["is_eligible"] = is_eligible
        job_item["package_lpa"] = job.get("ctc") or ""
        job_item["deadline"] = str(job.get("deadline")) if job.get("deadline") else "Ongoing"
        job_item["application_status"] = app_status_map.get(job["job_id"])

        upcoming_drives.append(job_item)

    eligible_count = sum(1 for job in upcoming_drives if job["is_eligible"])

    cursor.execute("SELECT COUNT(*) AS count FROM applications WHERE student_id = %s", (session["student_id"],))
    applied_count = cursor.fetchone()["count"]

    cursor.execute("SELECT COUNT(*) AS count FROM applications WHERE student_id = %s AND status = 'Interview'", (session["student_id"],))
    interview_count = cursor.fetchone()["count"]

    cursor.execute("SELECT COUNT(*) AS count FROM applications WHERE student_id = %s AND status = 'Selected'", (session["student_id"],))
    selected_count = cursor.fetchone()["count"]

    resume_score = session.get("resume_score", 0)

    # Recent applications for the dashboard card
    cursor.execute("""
        SELECT a.applied_date, a.status, a.resume_path,
               j.company_name, j.role, j.ctc as package_lpa, j.tier, j.deadline
        FROM applications a
        JOIN jobs j ON a.job_id = j.job_id
        WHERE a.student_id = %s
        ORDER BY a.applied_date DESC
        LIMIT 5
    """, (session["student_id"],))
    recent_applications = cursor.fetchall()
    # Replace 'Not Selected' display for 'Rejected'
    for app in recent_applications:
        if app.get('status') == 'Rejected':
            app['status'] = 'Not Selected'

    cursor.execute("""
        SELECT * FROM notifications
        WHERE student_id = %s AND is_read = 0
        ORDER BY created_at DESC
        """, (session["student_id"],))

    notifications = cursor.fetchall()

    return render_template(
        "student/dashboard.html",
        name=student["name"],
        student=student,
        upcoming_drives=upcoming_drives,
        eligible_count=eligible_count,
        applied_count=applied_count,
        interview_count=interview_count,
        selected_count=selected_count,
        resume_score=resume_score,
        notifications=notifications,
        recent_applications=recent_applications
    )
@app.route("/student_profile")
def student_profile():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")

    student_id = session["student_id"]

    query = "SELECT * FROM students WHERE student_id = %s"
    cursor.execute(query, (student_id,))
    student = cursor.fetchone()

    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")

    return render_template(
    "student/profile.html",
    student=student,
    must_change_password=session.get("must_change_password", 0)
)

@app.route("/update_profile", methods=["POST"])
def update_profile():
    if "student_id" not in session:
        return redirect("/student_login")
       
    student_id = session["student_id"]
   
    if "profile_photo" in request.files:
        photo = request.files["profile_photo"]
        if photo.filename != "":
            filename = secure_filename(photo.filename)
            active_year = session.get("active_year", "2025-2026")
            upload_folder = os.path.join(app.root_path, "static", "uploads", active_year)
            os.makedirs(upload_folder, exist_ok=True)
            filepath = os.path.join(upload_folder, f"student_{student_id}_{filename}")
            photo.save(filepath)
           
            db_path = f"/static/uploads/{active_year}/student_{student_id}_{filename}"
            cursor.execute("UPDATE students SET profile_photo = %s WHERE student_id = %s", (db_path, student_id))
            db.commit()
           
    return redirect("/student_profile")

@app.route("/update_profile_details", methods=["POST"])
def update_profile_details():
    if "student_id" not in session:
        return redirect("/student_login")
       
    student_id = session["student_id"]
    roll_number = request.form.get("roll_number", "").strip()
    phone_number = request.form.get("phone_number", "").strip()
    aadhar = request.form.get("aadhar", "").strip()
    pan = request.form.get("pan", "").strip()
   
    try:
        from flask import flash
        # Allow updating phone number only
        cursor.execute("UPDATE students SET phone_number = %s WHERE student_id = %s", (phone_number, student_id))
        db.commit()
        flash("Profile updated successfully.", "success")
    except Exception as e:
        flash("An error occurred while updating profile.", "error")
        pass
       
    return redirect("/student_profile")

@app.route("/update_skills", methods=["POST"])
def update_skills():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
       
    student_id = session["student_id"]
    skills = request.form.get("skills", "").strip()
   
    try:
        # Clean up skills string (remove extra spaces around commas)
        if skills:
            skills_list = [s.strip() for s in skills.split(',') if s.strip()]
            skills = ", ".join(skills_list)
           
        cursor.execute("UPDATE students SET skills = %s WHERE student_id = %s", (skills, student_id))
        db.commit()
        from flask import flash
        flash("Skills updated successfully!", "success")
    except Exception as e:
        print("Error updating skills:", e)
        from flask import flash
        flash("Error updating skills.", "danger")
        
    return redirect("/student_profile")

@app.route("/remove_skill", methods=["POST"])
def remove_skill():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
        
    student_id = session["student_id"]
    skill_to_remove = request.form.get("skill", "").strip()
    
    if skill_to_remove:
        try:
            cursor.execute("SELECT skills FROM students WHERE student_id = %s", (student_id,))
            student = cursor.fetchone()
            if student and student.get("skills"):
                current_skills = [s.strip() for s in student["skills"].split(',') if s.strip()]
                if skill_to_remove in current_skills:
                    current_skills.remove(skill_to_remove)
                    new_skills = ", ".join(current_skills)
                    cursor.execute("UPDATE students SET skills = %s WHERE student_id = %s", (new_skills, student_id))
                    db.commit()
                    from flask import flash
                    flash(f"Skill '{skill_to_remove}' removed successfully!", "success")
        except Exception as e:
            print("Error removing skill:", e)
            from flask import flash
            flash("Error removing skill.", "danger")
            
    return redirect("/student_profile")


@app.route("/eligible_companies")
def eligible_companies():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
   
    student_id = session["student_id"]
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    
    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")
   
    try:
        cursor.execute("SELECT * FROM jobs ORDER BY id DESC")
    except Exception:
        cursor.execute("SELECT * FROM jobs ORDER BY job_id DESC")
    all_jobs = cursor.fetchall()
   
    from datetime import datetime, timedelta
    # Check eligibility for each job
    active_jobs_list = []
    past_jobs_list = []
    for job in all_jobs:
        pwd_only = bool(job.get('pwd_only', False))
        is_pwd_student = (str(student.get('physically_challenged', '')).strip().lower() in ['yes', 'y', 'true', '1'])
        if pwd_only and not is_pwd_student:
            continue

        job_tier_str = str(job.get('tier', 'Tier 3')).lower()
        job_tier_num = 1 if '1' in job_tier_str else (2 if '2' in job_tier_str else 3)

        eligible_branches = [normalize_branch(b) for b in job.get('branches', '').split(',')] if job.get('branches') else []
        student_branch = normalize_branch(student['branch']) if student['branch'] else ""
       
        cgpa_ok = float(student.get('cgpa') or 0) >= float(job.get('cgpa_cutoff') or 0)
        
        s_backlogs = int(student.get("backlogs") or 0)
        s_history = int(student.get("backlog_history") or 0)
        j_backlogs = int(job.get("active_backlogs") or 0)
        j_history = int(job.get("backlog_history") or 0)
        
        backlogs_ok = s_backlogs <= j_backlogs
        backlog_history_ok = s_history <= j_history
        
        branch_ok = (student_branch in eligible_branches) or (not eligible_branches)
       
        # Tier check
        student_selected_tier = student.get('selected_tier')
        tier_1_val = student.get('tier_1')
        tier_2_val = student.get('tier_2')
        tier_3_val = student.get('tier_3')
        offers_count = sum(1 for t in [tier_1_val, tier_2_val, tier_3_val] if t and str(t).strip() and str(t).strip().lower() != 'nan')

        tier_ok = True
        tier_reason = None
        
        if offers_count >= 2:
            tier_ok = False
            tier_reason = "Tier Policy restriction: Selected in 2 companies, cannot apply for more."
        elif student_selected_tier is not None and student_selected_tier > 0:
            if student_selected_tier == 1:
                tier_ok = False
                tier_reason = f"Tier Policy restriction: Selected in Tier {student_selected_tier}, cannot apply for Tier {job_tier_num}"
            elif student_selected_tier == 2 and job_tier_num in [2, 3]:
                tier_ok = False
                tier_reason = f"Tier Policy restriction: Selected in Tier {student_selected_tier}, cannot apply for Tier {job_tier_num}"
            elif student_selected_tier == 3 and job_tier_num == 3:
                tier_ok = False
                tier_reason = f"Tier Policy restriction: Selected in Tier {student_selected_tier}, cannot apply for Tier {job_tier_num}"

        reasons = []
        if not cgpa_ok:
            reasons.append(f"CGPA below requirement ({student['cgpa']} < {job.get('cgpa_cutoff')})")
        if not backlogs_ok:
            reasons.append(f"Backlogs exceed maximum allowed ({student['backlogs']} > {job.get('active_backlogs')})")
        if not backlog_history_ok:
            reasons.append("Backlog History is not allowed for this job")

        if not branch_ok:
            reasons.append(f"Branch not eligible (Your branch: {student['branch'].upper()}, Eligible: {job.get('branches')})")
        if not tier_ok:
            if tier_reason:
                reasons.append(tier_reason)
            else:
                reasons.append(f"Tier Policy restriction: Selected in Tier {student_selected_tier}, cannot apply for Tier {job_tier_num}")
           
        is_eligible = cgpa_ok and backlogs_ok and backlog_history_ok and branch_ok and tier_ok
       
        # Check if already applied
        cursor.execute("SELECT * FROM applications WHERE student_id = %s AND job_id = %s", (student_id, job['job_id']))
        application = cursor.fetchone()
        applied = True if application else False
        status = application['status'] if application else None
        status_updated_at = application['status_updated_at'] if application else None
       
        job_item = dict(job)
        job_item['is_eligible'] = is_eligible
        job_item['reasons'] = reasons
        job_item['applied'] = applied
        job_item['application_status'] = status
       
        job_item['package_lpa'] = job.get('ctc') or ''
        job_item['min_cgpa'] = job.get('cgpa_cutoff') or 0
        job_item['max_backlogs'] = job.get('active_backlogs') or 0
        job_item['eligible_branches'] = job.get('branches') or ''
        job_item['deadline'] = str(job.get('deadline')) if job.get('deadline') else 'Ongoing'
        job_item['job_description'] = job.get('description') or ''
        job_item['required_skills'] = job.get('skills') or job.get('required_skills') or ''

        # Fetch number of rounds if set by faculty
        cursor.execute("SELECT num_rounds FROM recruitment_rounds WHERE job_id = %s", (job['job_id'],))
        r_round = cursor.fetchone()
        job_item['num_rounds'] = r_round['num_rounds'] if r_round else None

        is_past = False
        deadline_date = None
        if job.get('deadline'):
            try:
                deadline_date = datetime.strptime(str(job['deadline']), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    deadline_date = datetime.strptime(str(job['deadline']), "%Y-%m-%dT%H:%M")
                except ValueError:
                    pass
        
        # Rule 1: If job deadline has passed → move immediately
        deadline_expired = deadline_date and datetime.now() > deadline_date
        job_item['is_expired'] = deadline_expired
        if deadline_expired:
            is_past = True
            
        # Rule 2: If student is Selected or Not Selected → wait 3 days after status was set
        if status in ['Selected', 'Not Selected', 'Rejected'] and not is_past:
            if status_updated_at:
                if datetime.now() > status_updated_at + timedelta(days=3):
                    is_past = True
            else:
                # No timestamp recorded yet (legacy row) → move immediately
                is_past = True

        if is_past:
            past_jobs_list.append(job_item)
        else:
            active_jobs_list.append(job_item)
        
    return render_template("student/eligible_companies.html", active_jobs=active_jobs_list, past_jobs=past_jobs_list, student=student)

@app.route("/apply_job", methods=["POST"])
def apply_job():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
   
    student_id = session["student_id"]
    job_id = request.form.get("job_id")
    drive_link = request.form.get("drive_link")
   
    # Fetch student and job details to verify tier restrictions in backend
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    cursor.execute("SELECT * FROM jobs WHERE job_id = %s", (job_id,))
    job = cursor.fetchone()
   
    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")

    if not job:
        return "Invalid request."
       
    # Check if deadline has passed
    from datetime import datetime
    if job.get('deadline'):
        deadline = job['deadline']
        if isinstance(deadline, str):
            try:
                deadline = datetime.strptime(deadline, "%Y-%m-%d %H:%M:%S")
            except Exception:
                try:
                    deadline = datetime.strptime(deadline, "%Y-%m-%dT%H:%M")
                except Exception:
                    pass
        if isinstance(deadline, datetime) and datetime.now() > deadline:
            return """
            <script>
                alert("Application deadline has passed for this job!");
                window.location.href = "/eligible_companies";
            </script>
            """
       
    # Tier calculation
    job_tier_str = str(job.get('tier', 'Tier 3')).lower()
    job_tier_num = 1 if '1' in job_tier_str else (2 if '2' in job_tier_str else 3)
       
    tier_1_val = student.get('tier_1')
    tier_2_val = student.get('tier_2')
    tier_3_val = student.get('tier_3')
    offers_count = sum(1 for t in [tier_1_val, tier_2_val, tier_3_val] if t and str(t).strip() and str(t).strip().lower() != 'nan')

    student_selected_tier = student.get('selected_tier')
    tier_ok = True
    tier_reason = ""
    
    if offers_count >= 2:
        tier_ok = False
        tier_reason = "You are already selected in 2 companies. You cannot apply for more."
    elif student_selected_tier is not None and student_selected_tier > 0:
        if student_selected_tier == 1:
            tier_ok = False
            tier_reason = f"You are already selected in a Tier {student_selected_tier} job. You cannot apply for Tier {job_tier_num} roles."
        elif student_selected_tier == 2 and job_tier_num in [2, 3]:
            tier_ok = False
            tier_reason = f"You are already selected in a Tier {student_selected_tier} job. You cannot apply for Tier {job_tier_num} roles."
        elif student_selected_tier == 3 and job_tier_num == 3:
            tier_ok = False
            tier_reason = f"You are already selected in a Tier {student_selected_tier} job. You cannot apply for Tier {job_tier_num} roles."
           
    if not tier_ok:
        return """
        <script>
            alert("Policy Violation: """ + tier_reason + """");
            window.location.href = "/eligible_companies";
        </script>
        """
   
    # PWD check
    pwd_only = bool(job.get('pwd_only', False))
    is_pwd_student = (str(student.get('physically_challenged', '')).strip().lower() in ['yes', 'y', 'true', '1'])
    if pwd_only and not is_pwd_student:
        return """
        <script>
            alert("Policy Violation: This job opening is restricted to PWD students only.");
            window.location.href = "/eligible_companies";
        </script>
        """

    # Check if already applied
    cursor.execute("SELECT * FROM applications WHERE student_id = %s AND job_id = %s", (student_id, job_id))
    existing = cursor.fetchone()
    if existing:
        return """
        <script>
            alert("Already applied for this job!");
            window.location.href = "/eligible_companies";
        </script>
        """
   
    # Generate new application_id safely
    cursor.execute("SELECT COALESCE(MAX(application_id), 0) + 1 as next_id FROM applications")
    result = cursor.fetchone()
    next_id = result['next_id'] if result else 1
   
    # Extract dynamic extra details if present
    extra_data = {}
    aadhar = request.form.get("aadhar_number")
    pan = request.form.get("pan_number")
    other_info = request.form.get("other_info")
   
    if aadhar: extra_data["aadhar_number"] = aadhar
    if pan: extra_data["pan_number"] = pan
    if other_info: extra_data["other_info"] = other_info
   
    import json
    extra_details_json = json.dumps(extra_data) if extra_data else None
   
    query = """
        INSERT INTO applications (application_id, student_id, job_id, resume_path, status, applied_date, extra_details)
        VALUES (%s, %s, %s, %s, %s, CURDATE(), %s)
    """
    cursor.execute(query, (next_id, student_id, job_id, drive_link, "Pending", extra_details_json))
    db.commit()
   
    return """
    <script>
        alert("Application submitted successfully!");
        window.location.href = "/my_applications";
    </script>
    """

@app.route("/my_applications")
def my_applications():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
        
    student_id = session["student_id"]
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    
    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")
    
    query = """
        SELECT a.applied_date, a.status, a.resume_path, a.job_id, a.drive_link,
               j.company_name, j.role, j.ctc as package_lpa, j.tier, j.deadline,
               (SELECT MAX(round_number) FROM round_results rr WHERE rr.job_id = a.job_id AND rr.student_id = a.student_id AND rr.result = 'Selected') as max_cleared_round,
               (SELECT MAX(round_number) FROM round_results rr WHERE rr.job_id = a.job_id AND rr.student_id = a.student_id) as max_attempted_round,
               (SELECT num_rounds FROM recruitment_rounds WHERE job_id = a.job_id) as total_rounds,
               (SELECT result FROM round_results rr WHERE rr.job_id = a.job_id AND rr.student_id = a.student_id ORDER BY round_number DESC LIMIT 1) as latest_round_result,
               (SELECT drive_link FROM round_results rr WHERE rr.job_id = a.job_id AND rr.student_id = a.student_id AND drive_link IS NOT NULL ORDER BY round_number DESC LIMIT 1) as latest_drive_link,
               (SELECT round_number FROM round_results rr WHERE rr.job_id = a.job_id AND rr.student_id = a.student_id AND drive_link IS NOT NULL ORDER BY round_number DESC LIMIT 1) as latest_drive_round
        FROM applications a
        JOIN jobs j ON a.job_id = j.job_id
        WHERE a.student_id = %s
        ORDER BY a.applied_date DESC
    """
    cursor.execute(query, (student_id,))
    apps = cursor.fetchall()
    for app in apps:
        if app.get('status') == 'Rejected':
            app['status'] = 'Not Selected'
            
        app['pipeline_status'] = 'Applied'
        if app.get('status') == 'Selected':
            app['pipeline_status'] = 'Selected'
        elif app.get('status') == 'Rounds Cleared':
            app['pipeline_status'] = 'All Rounds Cleared'
        elif app.get('status') == 'Not Selected':
            app['pipeline_status'] = 'Not Selected'
        elif app.get('max_cleared_round'):
            next_round = app['max_cleared_round'] + 1
            if app.get('total_rounds') and next_round > app['total_rounds']:
                app['pipeline_status'] = f"Cleared Round {app['max_cleared_round']}"
            else:
                app['pipeline_status'] = f"Qualified for Round {next_round}"
        elif app.get('status') == 'Shortlisted' or app.get('status') == 'Interview':
            app['pipeline_status'] = 'Interviewing'
        elif app.get('status') == 'Pending':
            app['pipeline_status'] = 'Applied'

    cursor.execute("""
        SELECT job_id, round_number, drive_link 
        FROM round_results 
        WHERE student_id = %s AND drive_link IS NOT NULL
        ORDER BY round_number ASC
    """, (student_id,))
    links_data = cursor.fetchall()
    
    round_links_map = {}
    import re
    def is_valid_url(url):
        regex = re.compile(
            r'^(?:http|ftp)s?://' 
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'
            r'localhost|'
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}|'
            r'\[?[A-F0-9]*:[A-F0-9:]+\]?)'
            r'(?::\d+)?'
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        return re.match(regex, url) is not None

    for row in links_data:
        jid = row['job_id']
        if jid not in round_links_map:
            round_links_map[jid] = []
        raw_val = str(row['drive_link']).strip()
        is_u = is_valid_url(raw_val)
        round_links_map[jid].append({'round_number': row['round_number'], 'is_url': is_u, 'url': raw_val if is_u else '', 'raw_text': raw_val})

    for app in apps:
        app['round_links'] = round_links_map.get(app['job_id'], [])

    return render_template("student/my_applications.html", applications=apps, student=student)

@app.route("/student_selected_companies")
def student_selected_companies():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
        
    student_id = session["student_id"]
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    
    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")
    
    # Fetch jobs where they are formally 'Selected'
    query = """
        SELECT a.applied_date, j.job_id, j.company_name, j.role, j.tier, j.ctc
        FROM applications a
        JOIN jobs j ON a.job_id = j.job_id
        WHERE a.student_id = %s AND a.status = 'Selected'
        ORDER BY j.company_name ASC
    """
    cursor.execute(query, (student_id,))
    selected_jobs = cursor.fetchall()

    return render_template("student/selected_companies.html", student=student, selected_jobs=selected_jobs)

@app.route("/ongoing_rounds")
def ongoing_rounds():
    return redirect("/my_applications")

@app.route("/student/internships")
def student_internships():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
        
    student_id = session["student_id"]
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    if not student:
        session.pop("student_id", None)
        session.pop("student_name", None)
        return redirect("/student_login?error=Session+expired.+Please+log+in+again.")
        
    cursor.execute("""
        SELECT * FROM internship_postings 
        WHERE details IS NULL OR details NOT LIKE '[EXTERNAL]%' 
        ORDER BY posting_id DESC
    """)
    postings = cursor.fetchall()
    
    cursor.execute("""
        SELECT si.*, 
               IFNULL(ip.company_name, si.ext_company_name) AS company_name,
               IFNULL(ip.role, si.ext_role) AS role,
               IFNULL(ip.details, IF(si.is_external=1, '[EXTERNAL] External (self-arranged)', NULL)) AS details
        FROM student_internships si
        LEFT JOIN internship_postings ip ON si.posting_id = ip.posting_id
        WHERE si.student_id = %s AND si.status = 'Completed'
    """, (student_id,))
    completions = cursor.fetchall()
    completed_posting_ids = {c["posting_id"] for c in completions}
    
    for post in postings:
        post["completed"] = post["posting_id"] in completed_posting_ids
        
    return render_template("student/internships.html", postings=postings, completions=completions, student=student)

@app.route("/student/internships/submit", methods=["POST"])
def student_internships_submit():
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")
        
    student_id  = session["student_id"]
    posting_id  = request.form.get("posting_id", type=int)
    description = request.form.get("description", "").strip()
    hr_name     = request.form.get("hr_name", "").strip()
    hr_contact  = request.form.get("hr_contact", "").strip()
    
    if not posting_id or not description:
        flash("Invalid submission details.", "error")
        return redirect("/student/internships")

    cursor.execute("SELECT company_name, role FROM internship_postings WHERE posting_id = %s", (posting_id,))
    posting = cursor.fetchone()
    if not posting:
        flash("Internship posting not found.", "error")
        return redirect("/student/internships")

    # Certificate upload is OPTIONAL
    db_path = None
    file = request.files.get("certificate")
    if file and file.filename and file.filename.strip() != "":
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        active_year   = session.get("active_year", "2025-2026")
        upload_folder = os.path.join(app.root_path, "static", "uploads", active_year, "certificates")
        os.makedirs(upload_folder, exist_ok=True)
        import time
        timestamp     = int(time.time())
        dest_filename = f"student_{student_id}_{posting_id}_{timestamp}_{filename}"
        filepath      = os.path.join(upload_folder, dest_filename)
        file.save(filepath)
        db_path = f"/static/uploads/{active_year}/certificates/{dest_filename}"

    cursor.execute("SELECT id FROM student_internships WHERE student_id = %s AND posting_id = %s", (student_id, posting_id))
    existing = cursor.fetchone()
    
    if existing:
        if db_path:
            cursor.execute("""
                UPDATE student_internships 
                SET status = 'Completed', completion_description = %s, certificate_path = %s, submitted_at = NOW(), hr_name = %s, hr_contact = %s
                WHERE id = %s
            """, (description, db_path, hr_name, hr_contact, existing["id"]))
        else:
            cursor.execute("""
                UPDATE student_internships 
                SET status = 'Completed', completion_description = %s, submitted_at = NOW(), hr_name = %s, hr_contact = %s
                WHERE id = %s
            """, (description, hr_name, hr_contact, existing["id"]))
    else:
        cursor.execute("""
            INSERT INTO student_internships (student_id, posting_id, status, completion_description, certificate_path, submitted_at, hr_name, hr_contact)
            VALUES (%s, %s, 'Completed', %s, %s, NOW(), %s, %s)
        """, (student_id, posting_id, description, db_path, hr_name, hr_contact))
        
    cursor.execute("SELECT COUNT(*) as count FROM student_internships WHERE student_id = %s AND status = 'Completed'", (student_id,))
    completed_count = cursor.fetchone()["count"]
    cursor.execute("UPDATE students SET internships_count = %s WHERE student_id = %s", (completed_count, student_id))
    db.commit()
    flash("Internship completion details submitted successfully!", "success")
    return redirect("/student/internships")


@app.route("/student/internships/add_external", methods=["POST"])
def student_add_external_internship():
    """
    Allows students to add an internship they arranged themselves (outside college)
    without needing a faculty-posted listing. Certificate upload is optional.
    """
    ensure_connection()
    if "student_id" not in session:
        return redirect("/student_login")

    student_id   = session["student_id"]
    company_name = request.form.get("ext_company", "").strip()
    role         = request.form.get("ext_role", "").strip()
    description  = request.form.get("ext_description", "").strip()
    duration     = request.form.get("ext_duration", "").strip()
    hr_name      = request.form.get("ext_hr_name", "").strip()
    hr_contact   = request.form.get("ext_hr_contact", "").strip()

    if not company_name or not role or not description:
        flash("Company name, role, and description are required.", "error")
        return redirect("/student/internships")

    # Certificate upload — optional
    db_path = None
    file = request.files.get("ext_certificate")
    if file and file.filename and file.filename.strip() != "":
        from werkzeug.utils import secure_filename
        filename      = secure_filename(file.filename)
        active_year   = session.get("active_year", "2025-2026")
        upload_folder = os.path.join(app.root_path, "static", "uploads", active_year, "certificates")
        os.makedirs(upload_folder, exist_ok=True)
        import time
        timestamp     = int(time.time())
        dest_filename = f"ext_{student_id}_{timestamp}_{filename}"
        filepath      = os.path.join(upload_folder, dest_filename)
        file.save(filepath)
        db_path = f"/static/uploads/{active_year}/certificates/{dest_filename}"

    cursor.execute("""
        INSERT INTO student_internships (
            student_id, posting_id, status, completion_description, 
            certificate_path, submitted_at, ext_company_name, ext_role, 
            ext_duration, is_external, hr_name, hr_contact
        )
        VALUES (%s, NULL, 'Completed', %s, %s, NOW(), %s, %s, %s, 1, %s, %s)
    """, (student_id, description, db_path, company_name, role, duration, hr_name, hr_contact))

    cursor.execute("SELECT COUNT(*) as count FROM student_internships WHERE student_id = %s AND status = 'Completed'", (student_id,))
    completed_count = cursor.fetchone()["count"]
    cursor.execute("UPDATE students SET internships_count = %s WHERE student_id = %s", (completed_count, student_id))
    db.commit()

    flash(f"External internship at {company_name} added successfully!", "success")
    return redirect("/student/internships")


@app.route("/faculty/internships", methods=["GET", "POST"])
def faculty_internships():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    if request.method == "POST":
        action = request.form.get("action")
        if action == "post_internship":
            company_name = request.form.get("company_name", "").strip()
            role = request.form.get("role", "").strip()
            details = request.form.get("details", "").strip()
            link = request.form.get("link", "").strip()
            
            if not company_name or not role or not details or not link:
                flash("All fields are required to post an internship.", "error")
                return redirect("/faculty/internships")
                
            cursor.execute("""
                INSERT INTO internship_postings (company_name, role, details, link)
                VALUES (%s, %s, %s, %s)
            """, (company_name, role, details, link))
            
            cursor.execute("SELECT student_id FROM students")
            students = cursor.fetchall()
            for s in students:
                cursor.execute("""
                    INSERT INTO notifications (student_id, message, link, is_read)
                    VALUES (%s, %s, %s, 0)
                """, (s["student_id"], f"New Internship Opportunity: {company_name} - {role}", "/student/internships"))
                
            db.commit()
            flash("Internship opportunity posted and students notified successfully!", "success")
            
        elif action == "delete_posting":
            posting_id = request.form.get("posting_id", type=int)
            if posting_id:
                cursor.execute("DELETE FROM internship_postings WHERE posting_id = %s", (posting_id,))
                db.commit()
                flash("Internship posting deleted successfully.", "success")
                
        return redirect("/faculty/internships")
        
    cursor.execute("""
        SELECT * FROM internship_postings
        WHERE details IS NULL OR details NOT LIKE '[EXTERNAL]%'
        ORDER BY posting_id DESC
    """)
    postings = cursor.fetchall()

    
    cursor.execute("""
        SELECT si.*, s.name as student_name, s.roll_number, s.branch, s.cgpa,
               IFNULL(ip.company_name, si.ext_company_name) AS company_name,
               IFNULL(ip.role, si.ext_role) AS role,
               IFNULL(ip.details, IF(si.is_external=1, '[EXTERNAL] External (self-arranged)', NULL)) AS posting_details
        FROM student_internships si
        JOIN students s ON si.student_id = s.student_id
        LEFT JOIN internship_postings ip ON si.posting_id = ip.posting_id
        WHERE si.status = 'Completed'
        ORDER BY si.submitted_at DESC
    """)
    completions = cursor.fetchall()
    
    active_year_name = get_active_batch_name()
    return render_template("faculty/internships.html", postings=postings, completions=completions, active_year=active_year_name)

@app.route("/faculty/internships/export")
def faculty_internships_export():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    cursor.execute("""
        SELECT s.roll_number, s.name as student_name, s.branch, s.cgpa, 
               IFNULL(ip.company_name, si.ext_company_name) AS company_name, 
               IFNULL(ip.role, si.ext_role) AS role, 
               si.completion_description, si.certificate_path, si.submitted_at
        FROM student_internships si
        JOIN students s ON si.student_id = s.student_id
        LEFT JOIN internship_postings ip ON si.posting_id = ip.posting_id
        WHERE si.status = 'Completed'
        ORDER BY s.roll_number ASC
    """)
    completions = cursor.fetchall()
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Internship Certificates"
    
    title_font = Font(name="Calibri", size=14, bold=True, color="78350F")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=11)
    amber_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Internship Completion Report — Batch: {get_active_batch_name()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    headers = ["Roll Number", "Student Name", "Branch", "CGPA", "Company", "Role", "Completion Details", "Certificate Link", "Submitted Date"]
    ws.append([])
    ws.row_dimensions[2].height = 10
    
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    row_num = 4
    for comp in completions:
        sub_date = comp["submitted_at"].strftime('%Y-%m-%d %H:%M') if comp["submitted_at"] else "N/A"
        cert_url = f"{request.host_url.rstrip('/')}{comp['certificate_path']}" if comp["certificate_path"] else "N/A"
        
        row_data = [
            comp["roll_number"] or "N/A",
            comp["student_name"],
            comp["branch"].upper(),
            comp["cgpa"],
            comp["company_name"],
            comp["role"],
            comp["completion_description"],
            cert_url,
            sub_date
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 3, 4, 8, 9]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if col_idx == 8 and cert_url != "N/A":
                cell.hyperlink = cert_url
                cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        row_num += 1
        
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"internships_{session.get('active_year', 'batch').replace('-', '_')}.xlsx"
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

@app.route("/faculty/update_manual_stats", methods=["POST"])
def faculty_update_manual_stats():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    placement_rate = request.form.get("placement_rate", "").strip()
    avg_lpa = request.form.get("avg_lpa", "").strip()
    active_year_id = session.get("active_year", "2025-2026")
    
    manual_stats_file = os.path.join(app.root_path, "database", "manual_stats.json")
    stats = {}
    if os.path.exists(manual_stats_file):
        try:
            with open(manual_stats_file, "r") as f:
                stats = json.load(f)
        except Exception:
            pass
            
    stats[active_year_id] = {
        "placement_rate": placement_rate,
        "avg_lpa": avg_lpa
    }
    
    os.makedirs(os.path.dirname(manual_stats_file), exist_ok=True)
    with open(manual_stats_file, "w") as f:
        json.dump(stats, f)
        
    flash("Manual stats updated successfully.", "success")
    return redirect("/faculty/master_sheet")

@app.route("/faculty/manage_homepage", methods=["GET", "POST"])
def faculty_manage_homepage():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    updates_file = os.path.join(app.root_path, "database", "homepage_updates.json")
    updates = []
    if os.path.exists(updates_file):
        try:
            import json as _json
            with open(updates_file, "r") as f:
                updates = _json.load(f)
        except Exception as e:
            print("Error loading homepage updates:", e)
            
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_update":
            title = request.form.get("title", "").strip()
            type_val = request.form.get("type", "").strip()
            description = request.form.get("description", "").strip()
            link = request.form.get("link", "").strip()
            
            if not title or not type_val or not description:
                flash("Title, type, and description are required.", "error")
                return redirect("/faculty/manage_homepage")
                
            image_path = None
            if "photo" in request.files:
                photo = request.files["photo"]
                if photo.filename != "":
                    from werkzeug.utils import secure_filename
                    filename = secure_filename(photo.filename)
                    upload_folder = os.path.join(app.root_path, "static", "uploads", "homepage")
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    import time
                    timestamp = int(time.time())
                    dest_filename = f"{timestamp}_{filename}"
                    filepath = os.path.join(upload_folder, dest_filename)
                    photo.save(filepath)
                    image_path = f"/static/uploads/homepage/{dest_filename}"
                    
            import time
            from datetime import datetime
            new_item = {
                "id": str(int(time.time() * 1000)),
                "title": title,
                "type": type_val,
                "description": description,
                "link": link if link else None,
                "image_path": image_path,
                "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            updates.append(new_item)
            
            try:
                import json as _json
                with open(updates_file, "w") as f:
                    _json.dump(updates, f, indent=4)
                flash("Homepage item added successfully!", "success")
            except Exception as e:
                flash(f"Error saving update: {str(e)}", "error")
                
        elif action == "delete_update":
            update_id = request.form.get("update_id")
            if update_id:
                item_to_remove = None
                for item in updates:
                    if item.get("id") == update_id:
                        item_to_remove = item
                        break
                        
                if item_to_remove:
                    updates.remove(item_to_remove)
                    if item_to_remove.get("image_path"):
                        try:
                            photo_path = os.path.join(app.root_path, item_to_remove["image_path"].lstrip("/"))
                            if os.path.exists(photo_path):
                                os.remove(photo_path)
                        except Exception as e:
                            print("Error deleting image file:", e)
                            
                    try:
                        import json as _json
                        with open(updates_file, "w") as f:
                            _json.dump(updates, f, indent=4)
                        flash("Homepage item deleted successfully.", "success")
                    except Exception as e:
                        flash(f"Error deleting update: {str(e)}", "error")
                        
        return redirect("/faculty/manage_homepage")
        
    updates.sort(key=lambda x: x.get("id", ""), reverse=True)
    return render_template("faculty/manage_homepage.html", updates=updates)

@app.route("/homepage_updates/<string:update_id>")
def homepage_update_detail(update_id):
    ensure_connection()
    updates_file = os.path.join(app.root_path, "database", "homepage_updates.json")
    updates = []
    if os.path.exists(updates_file):
        try:
            import json as _json
            with open(updates_file, "r") as f:
                updates = _json.load(f)
        except Exception as e:
            print("Error loading homepage updates:", e)
            
    selected_update = None
    for u in updates:
        if u.get("id") == update_id:
            selected_update = u
            break
            
    if not selected_update:
        return "Update not found", 404
        
    return render_template("homepage_update_detail.html", item=selected_update)

@app.route("/student_logout")
def student_logout():
    session.pop("student_id", None)
    session.pop("student_name", None)
    session.pop("must_change_password", None)
    session.pop("resume_score", None)
    return redirect("/student_login")

def get_active_batch_name():
    from flask import session
    import json, os
    year_str = session.get("active_year", "2025-2026")
    try:
        batches_file = os.path.join(app.root_path, "database", "batches.json")
        if os.path.exists(batches_file):
            with open(batches_file, "r") as f:
                batches = json.load(f)
                for b in batches:
                    if b["id"] == year_str:
                        return b["name"]
    except Exception:
        pass
    return session.get("active_year_name", year_str)

@app.context_processor
def inject_global_settings():
    ensure_connection()
    active_year_name = get_active_batch_name()
    try:
        cursor.execute("SELECT setting_key, setting_value FROM global_settings")
        rows = cursor.fetchall()
        settings = {}
        for r in rows:
            settings[r["setting_key"]] = r["setting_value"]
        if 'recruitment_title' not in settings:
            settings['recruitment_title'] = 'Recruitment Season Live'
        return dict(global_settings=settings, active_year=active_year_name)
    except Exception:
        return dict(global_settings={'recruitment_title': 'Recruitment Season Live'}, active_year=active_year_name)

@app.route("/faculty/update_settings", methods=["POST"])
def faculty_update_settings():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    title = request.form.get("recruitment_title", "").strip()
    if title:
        cursor.execute("""
            INSERT INTO global_settings (setting_key, setting_value) 
            VALUES ('recruitment_title', %s) 
            ON DUPLICATE KEY UPDATE setting_value = %s
        """, (title, title))
        db.commit()
        flash("Settings updated successfully!", "success")
    return redirect("/faculty_dashboard")


@app.route("/faculty/email_manager")
def faculty_email_manager():
    """Render the dedicated email management page."""
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    return render_template("faculty/email_manager.html")


@app.route("/faculty/email_settings", methods=["POST"])
def faculty_email_settings():
    """Toggle the global email notifications ON/OFF."""
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Unauthorized"}), 401

    data    = request.get_json() or {}
    enabled = data.get("enabled", False)
    value   = "true" if enabled else "false"

    try:
        cursor.execute("""
            INSERT INTO global_settings (setting_key, setting_value)
            VALUES ('email_notifications_enabled', %s)
            ON DUPLICATE KEY UPDATE setting_value = %s
        """, (value, value))
        db.commit()

        smtp_ok = _EMAIL_SVC_LOADED and _email_svc.is_configured()

        return jsonify({
            "success":   True,
            "enabled":   enabled,
            "smtp_ready": smtp_ok,
            "message":  ("Email notifications enabled." if enabled else "Email notifications disabled.")
                        + ("" if smtp_ok else " ⚠️ SMTP not configured in .env — emails will not be sent.")
        })
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/test_smtp_email", methods=["POST"])
def faculty_test_smtp_email():
    """
    Send a SYNCHRONOUS test email so faculty can verify SMTP credentials work.
    Returns the real SMTP result immediately (not async).
    """
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Unauthorized"}), 401

    if not _EMAIL_SVC_LOADED:
        return jsonify({"success": False, "error": "email_service.py failed to load. Check server logs."})

    # Reload credentials fresh from disk each time (in case .env was updated)
    try:
        from dotenv import load_dotenv
        load_dotenv(override=True)
        import importlib
        import email_service as _esvc_mod
        importlib.reload(_esvc_mod)
    except Exception:
        pass

    if not _email_svc.is_configured():
        return jsonify({
            "success": False,
            "error": "SMTP not configured. Fill in SMTP_EMAIL and SMTP_PASSWORD in .env, then restart the server."
        })

    data       = request.get_json() or {}
    to_email   = data.get("to_email", "").strip()
    if not to_email:
        # Default: send to the logged-in faculty's email
        to_email = session.get("faculty_email")
        if not to_email:
            to_email = _email_svc.SMTP_EMAIL

    subject  = "✅ NIT AP Portal — SMTP Test Email"
    html_body = _email_svc._base_template(f"""
        <p class="greeting">Test Email 🎉</p>
        <p>This is a test email sent from the <strong>NIT AP Placement Portal</strong> to verify that SMTP credentials are working correctly.</p>
        <div class="info-box">
          <div class="info-row"><span class="info-label">Sent To</span><span class="info-value">{to_email}</span></div>
          <div class="info-row"><span class="info-label">SMTP Host</span><span class="info-value">{_email_svc.SMTP_HOST}:{_email_svc.SMTP_PORT}</span></div>
          <div class="info-row"><span class="info-label">From Address</span><span class="info-value">{_email_svc.SMTP_EMAIL}</span></div>
        </div>
        <p>✅ If you can read this, email delivery is working correctly!</p>
    """, "SMTP Test")

    result = _email_svc.send_email(to_email, subject, html_body)
    return jsonify(result)


@app.route("/api/faculty/email_logs")
def api_faculty_email_logs():
    """Return recent email log records for the admin dashboard."""
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    ensure_connection()
    try:
        cursor.execute("""
            SELECT id, recipient_email, student_name, event_type, subject,
                   sent_at, status, error_message
            FROM email_logs
            ORDER BY sent_at DESC
            LIMIT 50
        """)
        rows = cursor.fetchall()
        # Convert datetime objects to strings
        for r in rows:
            if r.get("sent_at"):
                r["sent_at"] = str(r["sent_at"])
        return jsonify({"logs": rows})
    except Exception as e:
        return jsonify({"logs": [], "error": str(e)})


@app.route("/api/faculty/jobs_list")
def api_faculty_jobs_list():
    """Return a lightweight list of active jobs for the selective email modal dropdown."""
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    ensure_connection()
    try:
        cursor.execute("SELECT job_id, company_name, role FROM jobs ORDER BY company_name, role")
        rows = cursor.fetchall()
        return jsonify({"jobs": rows})
    except Exception as e:
        return jsonify({"jobs": [], "error": str(e)})


@app.route("/api/faculty/email_status")

def api_faculty_email_status():
    """Return current email notification toggle state + SMTP readiness."""
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    ensure_connection()
    try:
        cursor.execute("SELECT setting_value FROM global_settings WHERE setting_key='email_notifications_enabled'")
        row = cursor.fetchone()
        enabled   = row and row["setting_value"] == "true"
        smtp_ok   = _EMAIL_SVC_LOADED and _email_svc.is_configured()

        # Count eligible recipients (all students with email)
        cursor.execute("SELECT COUNT(*) AS c FROM students WHERE email IS NOT NULL AND email != ''")
        total_r = cursor.fetchone()
        total_email_students = total_r["c"] if total_r else 0

        # Stats from email_logs
        cursor.execute("SELECT COUNT(*) AS c FROM email_logs WHERE status='sent'")
        r2 = cursor.fetchone()
        total_sent = r2["c"] if r2 else 0

        cursor.execute("SELECT COUNT(*) AS c FROM email_logs WHERE status='failed'")
        r3 = cursor.fetchone()
        total_failed = r3["c"] if r3 else 0

        return jsonify({
            "enabled":             enabled,
            "smtp_ready":          smtp_ok,
            "smtp_email":          _email_svc.SMTP_EMAIL if _EMAIL_SVC_LOADED else "",
            "total_email_students": total_email_students,
            "total_sent":          total_sent,
            "total_failed":        total_failed
        })
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/faculty/send_announcement_email", methods=["POST"])
def faculty_send_announcement_email():
    """Send an announcement email to all students with registered emails."""
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Unauthorized"}), 401

    if not _is_email_notifications_enabled():
        return jsonify({"success": False, "error": "Email notifications are disabled. Enable them first."})

    if not _EMAIL_SVC_LOADED or not _email_svc.is_configured():
        return jsonify({"success": False, "error": "SMTP is not configured. Please fill in .env credentials."})

    data    = request.get_json(silent=True) or {}
    title   = request.form.get("title", "").strip() or data.get("title", "").strip()
    message = request.form.get("message", "").strip() or data.get("message", "").strip()

    attachments = []
    for file in request.files.getlist("attachments"):
        if file and file.filename:
            attachments.append((file.filename, file.read()))

    if not title or not message:
        return jsonify({"success": False, "error": "Title and message are required."})

    try:
        cursor.execute("SELECT student_id, name, email FROM students WHERE email IS NOT NULL AND email != ''")
        recipients = cursor.fetchall()
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not fetch students: {e}"})

    if not recipients:
        return jsonify({"success": False, "error": "No students with registered emails found."})

    subject_ann = f"Placement Announcement: {title}"

    def html_fn_ann(rec):
        return _email_svc.build_announcement_email(rec, title, message)

    def log_fn_ann(sid, email, status, err):
        _log_email(email, None, "announcement", subject_ann, status, err)

    _email_svc.send_bulk_emails_async(list(recipients), subject_ann, html_fn_ann, log_fn_ann, event_type="announcement", attachments=attachments)

    return jsonify({
        "success":    True,
        "recipients": len(recipients),
        "message":    f"Announcement is being sent to {len(recipients)} students asynchronously."
    })


@app.route("/api/faculty/students_with_email")
def api_faculty_students_with_email():
    """
    Return students who have a registered email — used by the Selective Email modal.
    Optional filters: ?job_id=X  or  ?status=Shortlisted
    """
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    ensure_connection()

    job_id     = request.args.get("job_id")
    status_fil = request.args.get("status")   # e.g. 'Shortlisted', 'Selected', etc.

    try:
        if job_id:
            # Return students who applied to a specific job (optionally filtered by status)
            query = """
                SELECT s.student_id, s.name, s.email, s.roll_number AS roll_no, s.branch,
                       a.status AS application_status
                FROM applications a
                JOIN students s ON a.student_id = s.student_id
                WHERE a.job_id = %s
                  AND s.email IS NOT NULL AND s.email != ''
            """
            params = [job_id]
            if status_fil:
                query += " AND a.status = %s"
                params.append(status_fil)
            query += " ORDER BY s.name"
            cursor.execute(query, params)
        elif status_fil:
            # Filter by status across ALL jobs
            query = """
                SELECT DISTINCT s.student_id, s.name, s.email, s.roll_number AS roll_no, s.branch,
                       a.status AS application_status
                FROM applications a
                JOIN students s ON a.student_id = s.student_id
                WHERE s.email IS NOT NULL AND s.email != ''
                  AND a.status = %s
                ORDER BY s.name
            """
            cursor.execute(query, [status_fil])
        else:
            # Return all students with emails
            cursor.execute("""
                SELECT student_id, name, email, roll_number AS roll_no, branch,
                       '' AS application_status
                FROM students
                WHERE email IS NOT NULL AND email != ''
                ORDER BY name
                LIMIT 500
            """)

        rows = cursor.fetchall()
        return jsonify({"students": rows, "total": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e), "students": []})


@app.route("/faculty/send_selective_email", methods=["POST"])
def faculty_send_selective_email():
    """
    Send a custom email to a manually selected list of students.
    Body: { student_ids: [1,2,3], subject: '...', title: '...', message: '...' }
    """
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Unauthorized"}), 401

    if not _is_email_notifications_enabled():
        return jsonify({"success": False, "error": "Email notifications are disabled."})

    if not _EMAIL_SVC_LOADED or not _email_svc.is_configured():
        return jsonify({"success": False, "error": "SMTP not configured. Fill in .env credentials."})

    data = request.get_json(silent=True) or {}
    
    student_ids_str = request.form.get("student_ids")
    if student_ids_str:
        import json
        try:
            student_ids = json.loads(student_ids_str)
        except:
            student_ids = []
    else:
        student_ids = data.get("student_ids", [])
        
    subject     = request.form.get("subject", "").strip() or data.get("subject", "").strip()
    title       = request.form.get("title", "").strip() or data.get("title", "").strip()
    message_txt = request.form.get("message", "").strip() or data.get("message", "").strip()

    attachments = []
    for file in request.files.getlist("attachments"):
        if file and file.filename:
            attachments.append((file.filename, file.read()))

    if not student_ids:
        return jsonify({"success": False, "error": "No students selected."})
    if not subject or not message_txt:
        return jsonify({"success": False, "error": "Subject and message are required."})

    try:
        # Fetch details for only the selected student IDs
        fmt = ','.join(['%s'] * len(student_ids))
        cursor.execute(f"""
            SELECT student_id, name, email FROM students
            WHERE student_id IN ({fmt})
              AND email IS NOT NULL AND email != ''
        """, student_ids)
        recipients = cursor.fetchall()
    except Exception as e:
        return jsonify({"success": False, "error": f"DB error: {e}"})

    if not recipients:
        return jsonify({"success": False, "error": "None of the selected students have registered emails."})

    # Capture DB name now (in request context) for use inside the thread
    db_name_cap = session.get('active_year_db', 'placement_portal_2025_2026')

    def html_fn_sel(rec):
        return _email_svc.build_announcement_email(rec, title or subject, message_txt)

    def log_fn_sel(sid, email, status, err):
        _log_email(email, None, "selective", subject, status, err, db_name=db_name_cap)

    _email_svc.send_bulk_emails_async(
        list(recipients), subject, html_fn_sel, log_fn_sel, event_type="selective", attachments=attachments
    )

    return jsonify({
        "success":    True,
        "recipients": len(recipients),
        "message":    f"Email is being sent to {len(recipients)} selected student(s)."
    })


def get_dashboard_stats(active_year_id=None):
    ensure_connection()
    try:
        cursor.execute("SELECT COUNT(*) AS c FROM students")
        total_students = cursor.fetchone()["c"]
    except Exception:
        total_students = 0
    try:
        cursor.execute("SELECT COUNT(*) AS c FROM jobs")
        active_jobs = cursor.fetchone()["c"]
    except Exception:
        active_jobs = 0
    
    placement_rate = 0.0
    avg_lpa = 0.0
    
    manual_stats_file = os.path.join(app.root_path, "database", "manual_stats.json")
    if active_year_id and os.path.exists(manual_stats_file):
        try:
            with open(manual_stats_file, "r") as f:
                manual_stats = json.load(f)
            if active_year_id in manual_stats:
                if "placement_rate" in manual_stats[active_year_id] and manual_stats[active_year_id]["placement_rate"] != "":
                    placement_rate = float(manual_stats[active_year_id]["placement_rate"])
                if "avg_lpa" in manual_stats[active_year_id] and manual_stats[active_year_id]["avg_lpa"] != "":
                    avg_lpa = float(manual_stats[active_year_id]["avg_lpa"])
        except Exception:
            pass

    return total_students, active_jobs, placement_rate, avg_lpa

def get_login_batches_stats():
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    batches_stats = []
    
    if os.path.exists(batches_file):
        try:
            with open(batches_file, "r") as f:
                batches = json.load(f)
                for batch in batches[:2]:
                    try:
                        switch_active_db(batch["db"], batch["id"])
                        ts, aj, pr, alpa = get_dashboard_stats(active_year_id=batch["id"])
                    except Exception:
                        ts, aj, pr, alpa = 0, 0, 0, 0.0
                        
                    batches_stats.append({
                        "name": batch["name"],
                        "total_students": ts,
                        "active_jobs": aj,
                        "placement_rate": pr,
                        "avg_lpa": alpa
                    })
        except Exception:
            pass
    return batches_stats

@app.route("/faculty_login")
def faculty_login_page():
    # Always show the login form - never auto-redirect
    return render_template("faculty/login.html", batches_stats=get_login_batches_stats())

@app.route("/faculty_login_check", methods=["POST"])
def faculty_login_check():
    ensure_connection()
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "").strip()

    # 1. Try 2025-2026
    switch_active_db("placement_portal_2025_2026", "2025-2026")
    ensure_connection()
    cursor.execute("SELECT * FROM faculty WHERE email = %s AND password = %s", (email, password))
    faculty = cursor.fetchone()

    if not faculty:
        # 2. Try 2026-2027
        switch_active_db("placement_portal_2026_2027", "2026-2027")
        ensure_connection()
        cursor.execute("SELECT * FROM faculty WHERE email = %s AND password = %s", (email, password))
        faculty = cursor.fetchone()

    if faculty:
        session["faculty_email"] = email
        session["faculty_name"] = faculty["name"]
        session.permanent = True
        return redirect("/faculty/select_year")

    return render_template("faculty/login.html", error="Invalid email or password.",
                           batches_stats=get_login_batches_stats())


def faculty_required():
    """Returns None if faculty is logged in, else a redirect response."""
    if "faculty_email" not in session:
        return redirect("/faculty_login")
    if "active_year_db" not in session and request.endpoint not in ["faculty_select_year", "faculty_set_year"]:
        return redirect("/faculty/select_year")
    return None


@app.route("/faculty_dashboard")
def faculty_dashboard():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    cursor.execute("SELECT COUNT(*) AS count FROM students")
    total_students = cursor.fetchone()["count"]
    cursor.execute("SELECT COUNT(*) AS count FROM jobs")
    active_jobs = cursor.fetchone()["count"]

    # Stats are now manually set via Master Sheet
    avg_lpa = "0.0"
    placement_rate = "0.0"

    # Get active batch name instead of just ID
    active_year_id = session.get("active_year", "2025-2026")
    active_year_name = active_year_id
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches_list = json.load(f)
            for b in batches_list:
                if b["id"] == active_year_id:
                    active_year_name = b["name"]
                    break

    upload_dir = os.path.join(app.static_folder, "uploads", active_year_id)
    master_sheet_status = "Empty"
    if os.path.exists(os.path.join(upload_dir, "master_sheet.pdf")) or \
       os.path.exists(os.path.join(upload_dir, "master_sheet.xlsx")) or \
       os.path.exists(os.path.join(upload_dir, "master_sheet.csv")):
        master_sheet_status = "Uploaded"

    # Fetch due reminders for the dashboard alert
    try:
        cursor.execute("SELECT id, job_id, company_name, role, reminder_date, reminder_note FROM jobs WHERE reminder_date <= NOW() AND (reminder_sent IS NULL OR reminder_sent = 0) ORDER BY reminder_date DESC")
        due_reminders = cursor.fetchall()
    except Exception:
        due_reminders = []
        
    try:
        cursor.execute("SELECT id, job_id, company_name, role, deadline FROM jobs WHERE deadline IS NOT NULL AND (deadline_dismissed IS NULL OR deadline_dismissed = 0)")
        all_jobs_for_deadline = cursor.fetchall()
        deadline_passed_jobs = []
        from datetime import datetime, timedelta
        now = datetime.now()
        for jb in all_jobs_for_deadline:
            dl = jb['deadline']
            if isinstance(dl, str):
                try: dl = datetime.strptime(dl, "%Y-%m-%d %H:%M:%S")
                except:
                    try: dl = datetime.strptime(dl, "%Y-%m-%dT%H:%M")
                    except: continue
            if isinstance(dl, datetime):
                if dl < now and dl >= now - timedelta(days=3):
                    deadline_passed_jobs.append(jb)
    except Exception:
        deadline_passed_jobs = []

    manual_stats_file = os.path.join(app.root_path, "database", "manual_stats.json")
    if os.path.exists(manual_stats_file):
        try:
            with open(manual_stats_file, "r") as f:
                manual_stats = json.load(f)
            if active_year_id in manual_stats:
                if "placement_rate" in manual_stats[active_year_id] and manual_stats[active_year_id]["placement_rate"] != "":
                    placement_rate = float(manual_stats[active_year_id]["placement_rate"])
                if "avg_lpa" in manual_stats[active_year_id] and manual_stats[active_year_id]["avg_lpa"] != "":
                    avg_lpa = float(manual_stats[active_year_id]["avg_lpa"])
        except Exception as e:
            print("Error loading manual stats:", e)

    return render_template(
        "faculty/dashboard.html",
        name=session["faculty_name"],
        total_students=total_students,
        active_jobs=active_jobs,
        placement_rate=placement_rate,
        avg_lpa=avg_lpa,
        master_sheet_status=master_sheet_status,
        active_year=active_year_name,
        due_reminders=due_reminders,
        deadline_passed_jobs=deadline_passed_jobs
    )

@app.route("/faculty/jobs/dismiss_deadline", methods=["POST"])
def faculty_dismiss_deadline():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    job_id = data.get("job_id")
    if not job_id:
        return jsonify({"success": False, "error": "Missing job ID"})
    try:
        cursor.execute("UPDATE jobs SET deadline_dismissed = 1 WHERE id = %s", (job_id,))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})

@app.route("/faculty/jobs/dismiss_reminder", methods=["POST"])
def faculty_dismiss_reminder():
    """Mark a reminder as dismissed (sent) so it hides from dashboard."""
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    job_id = data.get("job_id")
    if not job_id:
        return jsonify({"success": False, "error": "Missing job ID"})
    try:
        cursor.execute("UPDATE jobs SET reminder_sent = 1 WHERE id = %s", (job_id,))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})



@app.route("/faculty_logout")
def faculty_logout():
    session.pop("faculty_email", None)
    session.pop("faculty_name", None)
    session.pop("active_year", None)
    session.pop("active_year_db", None)
    return redirect("/faculty_login")


@app.route("/api/faculty/stats")
def api_faculty_stats():
    """Live stats endpoint for auto-refresh polling on faculty dashboard."""
    ensure_connection()
    redir = faculty_required()
    if redir:
        return jsonify({"error": "unauthorized"}), 403
    try:
        active_year_id = session.get("active_year", "2025-2026")
        ts, aj, pr, alpa = get_dashboard_stats(active_year_id=active_year_id)
        
        return jsonify({
            "total_students": ts,
            "active_jobs": aj,
            "placement_rate": pr,
            "avg_lpa": alpa
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/student/stats")
def api_student_stats():
    """Live stats endpoint for auto-refresh polling on student dashboard."""
    ensure_connection()
    if "student_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    try:
        student_id = session["student_id"]
        cursor.execute("SELECT COUNT(*) AS c FROM applications WHERE student_id = %s", (student_id,))
        applied_count = cursor.fetchone()["c"] or 0

        cursor.execute("SELECT COUNT(*) AS c FROM applications WHERE student_id = %s AND status = 'Selected'", (student_id,))
        selected_count = cursor.fetchone()["c"] or 0

        cursor.execute("SELECT COUNT(*) AS c FROM jobs")
        active_jobs = cursor.fetchone()["c"] or 0

        cursor.execute("SELECT COUNT(*) AS c FROM notifications WHERE student_id = %s AND is_read = 0", (student_id,))
        unread_notifications = cursor.fetchone()["c"] or 0

        return jsonify({
            "applied_count": applied_count,
            "selected_count": selected_count,
            "active_jobs": active_jobs,
            "unread_notifications": unread_notifications
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/faculty/select_year")
def faculty_select_year():
    ensure_connection()
    if "faculty_email" not in session:
        return redirect("/faculty_login")
        
    import json, os
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    batches = []
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches = json.load(f)
            
    return render_template("faculty/select_year.html", batches=batches)


@app.route("/faculty/set_year/<year>")
def faculty_set_year(year):
    ensure_connection()
    if "faculty_email" not in session:
        return redirect("/faculty_login")
    
    import json, os
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    batches = []
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches = json.load(f)
            
    valid_years = [b["id"] for b in batches]
    
    if year not in valid_years:
        flash("Invalid year batch selection.", "error")
        return redirect("/faculty/select_year")
        
    db_name = next(b["db"] for b in batches if b["id"] == year)
    batch_name = next((b["name"] for b in batches if b["id"] == year), year)
    switch_active_db(db_name, year, batch_name)
    
    flash(f"Switched to {batch_name} Batch successfully.", "success")
    return redirect("/faculty_dashboard")

@app.route("/faculty/manage_batches", methods=["POST"])
def faculty_manage_batches():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    action = request.form.get("action")
    import json, os
    batches_file = os.path.join(app.root_path, "database", "batches.json")
    batches = []
    if os.path.exists(batches_file):
        with open(batches_file, "r") as f:
            batches = json.load(f)
            
    if action == "edit":
        edit_id = request.form.get("batch_id")
        edit_name = request.form.get("batch_name")
        edit_desc = request.form.get("batch_desc")
        
        for b in batches:
            if b["id"] == edit_id:
                b["name"] = edit_name
                if edit_desc:
                    b["desc"] = edit_desc
                else:
                    b["desc"] = f"Manage recruitment drives, student details, eligibility criteria, and outcomes for {edit_name}."
                break
                
        with open(batches_file, "w") as f:
            json.dump(batches, f, indent=2)

        # Update session if renaming the currently active batch
        if session.get("active_year") == edit_id:
            session["active_year_name"] = edit_name
            
        flash(f"Batch '{edit_name}' updated successfully.", "success")
        
    return redirect("/faculty/select_year")


@app.route("/faculty/reset_batch", methods=["POST"])

@app.route("/faculty/edit_yearly_stats", methods=["GET", "POST"])
def faculty_edit_yearly_stats():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    stats_file = os.path.join(app.root_path, "database", "yearly_statistics.json")
    past_stats = []
    
    if os.path.exists(stats_file):
        import json as _json
        try:
            with open(stats_file, "r") as f:
                past_stats = _json.load(f)
        except Exception as e:
            print("Error loading stats:", e)
            
    if request.method == "POST":
        try:
            new_stats = []
            for stat in past_stats:
                year = stat.get("year")
                
                # Fetch new values from form
                tot_stu = request.form.get(f"tot_stu_{year}")
                pl_stu = request.form.get(f"pl_stu_{year}")
                pr_rate = request.form.get(f"pr_rate_{year}")
                avg_lpa = request.form.get(f"avg_lpa_{year}")
                
                if tot_stu is not None: stat["total_students"] = int(tot_stu)
                if pl_stu is not None: stat["placed_students"] = int(pl_stu)
                if pr_rate is not None: stat["placement_rate"] = float(pr_rate)
                if avg_lpa is not None: stat["average_lpa"] = float(avg_lpa)
                
                new_stats.append(stat)
                
            import json as _json
            with open(stats_file, "w") as f:
                _json.dump(new_stats, f, indent=4)
                
            flash("Yearly statistics updated successfully.", "success")
            return redirect("/faculty_dashboard")
        except Exception as e:
            flash(f"Error updating stats: {str(e)}", "error")
            
    return render_template("faculty/edit_yearly_stats.html", past_stats=past_stats, name=session.get("faculty_name"))


def faculty_reset_batch():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
    
    confirm_checkbox = request.form.get("confirm_save")
    if not confirm_checkbox:
        flash("You must confirm that you have saved student data before resetting.", "error")
        return redirect("/faculty/master_sheet")
    
    # Verify faculty email and password
    confirm_email = request.form.get("confirm_email", "").strip().lower()
    confirm_password = request.form.get("confirm_password", "").strip()
    
    if not confirm_email or not confirm_password:
        flash("You must enter your faculty email and password to reset.", "error")
        return redirect("/faculty/master_sheet")
    
    cursor.execute("SELECT * FROM faculty WHERE LOWER(email) = %s AND password = %s", (confirm_email, confirm_password))
    faculty = cursor.fetchone()
    if not faculty:
        flash("Incorrect email or password. Reset denied.", "error")
        return redirect("/faculty/master_sheet")
        
    active_year = session.get("active_year", "2025-2026")
    active_year_name = get_active_batch_name()
    
    # --- ARCHIVE YEARLY STATISTICS ---
    try:
        cursor.execute("SELECT COUNT(*) AS c FROM students")
        total_students_val = cursor.fetchone()["c"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS c FROM students
            WHERE LOWER(TRIM(COALESCE(career_option,''))) IN ('job', 'psu')
               OR career_option IS NULL
        """)
        tot_interested = cursor.fetchone()["c"] or 0

        cursor.execute("""
            SELECT COUNT(DISTINCT s.student_id) AS c
            FROM students s
            JOIN applications a ON a.student_id = s.student_id
            WHERE LOWER(TRIM(COALESCE(s.career_option,''))) NOT IN ('job', 'psu')
              AND s.career_option IS NOT NULL
        """)
        not_int_app = cursor.fetchone()["c"] or 0

        denom = tot_interested + not_int_app
        
        cursor.execute("""
            SELECT COUNT(DISTINCT student_id) AS c 
            FROM students 
            WHERE (selected_tier IS NOT NULL AND selected_tier > 0)
               OR (tier_1 IS NOT NULL AND TRIM(tier_1) != '' AND LOWER(TRIM(tier_1)) != 'nan')
               OR (tier_2 IS NOT NULL AND TRIM(tier_2) != '' AND LOWER(TRIM(tier_2)) != 'nan')
               OR (tier_3 IS NOT NULL AND TRIM(tier_3) != '' AND LOWER(TRIM(tier_3)) != 'nan')
        """)
        tot_placed = cursor.fetchone()["c"] or 0
        
        pr_rate = round((tot_placed / denom * 100), 1) if denom > 0 else 0.0

        cursor.execute("SELECT ctc FROM jobs WHERE ctc IS NOT NULL AND ctc != ''")
        ctc_rows = cursor.fetchall()
        tot_lpa, cnt_lpa = 0, 0
        import re as _re
        for r in ctc_rows:
            m = _re.search(r'([\d.]+)', str(r['ctc']))
            if m:
                tot_lpa += float(m.group(1))
                cnt_lpa += 1
        avg_lpa_val = round(tot_lpa / cnt_lpa, 1) if cnt_lpa > 0 else 0.0

        stats_file = os.path.join(app.root_path, "database", "yearly_statistics.json")
        past_stats = []
        if os.path.exists(stats_file):
            with open(stats_file, "r") as sf:
                try:
                    import json as _json
                    past_stats = _json.load(sf)
                except:
                    past_stats = []
        
        # update or append
        found = False
        for stat in past_stats:
            if stat.get("year") == active_year_name:
                stat["total_students"] = total_students_val
                stat["placed_students"] = tot_placed
                stat["placement_rate"] = pr_rate
                stat["average_lpa"] = avg_lpa_val
                found = True
                break
        
        if not found:
            past_stats.append({
                "id": active_year,
                "year": active_year_name,
                "total_students": total_students_val,
                "placed_students": tot_placed,
                "placement_rate": pr_rate,
                "average_lpa": avg_lpa_val
            })
            
        import json as _json
        with open(stats_file, "w") as sf:
            _json.dump(past_stats, sf, indent=4)
            
    except Exception as e:
        print("Error archiving stats:", e)
    # ---------------------------------

    # 1. Delete all uploaded files for this year (profile photos, job PDFs, master sheets)
    import shutil
    upload_dir = os.path.join(app.static_folder, "uploads", active_year)
    if os.path.exists(upload_dir):
        try:
            shutil.rmtree(upload_dir)
        except Exception as e:
            print(f"Error removing upload dir {upload_dir}: {e}")
            
    # 2. Clear all tables in the active year database (except faculty)
    try:
        cursor.execute("SELECT setting_value FROM global_settings WHERE setting_key = 'recruitment_title'")
        res = cursor.fetchone()
        current_title = res["setting_value"] if res else "Recruitment Season Live"

        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        cursor.execute("DELETE FROM round_results")
        cursor.execute("DELETE FROM recruitment_rounds")
        cursor.execute("DELETE FROM applications")
        cursor.execute("DELETE FROM notifications")
        cursor.execute("DELETE FROM students")
        cursor.execute("DELETE FROM jobs")
        cursor.execute("DELETE FROM global_settings")
        cursor.execute("INSERT INTO global_settings (setting_key, setting_value) VALUES ('recruitment_title', %s)", (current_title,))
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        db.commit()
        
        flash(f"Database for batch '{active_year_name}' has been reset. All student data and files were removed.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error resetting database: {str(e)}", "error")
        
    return redirect("/faculty/master_sheet")


# ─── FACULTY: JOBS ───────────────────────────────────────────────────────────

@app.route("/faculty/jobs")
def faculty_jobs():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    try:
        cursor.execute("SELECT * FROM jobs ORDER BY id DESC")
    except Exception:
        # Fallback if 'id' column doesn't exist in older schema
        cursor.execute("SELECT * FROM jobs ORDER BY job_id DESC")
    jobs = cursor.fetchall()
   
    for j in jobs:
        cursor.execute("SELECT COUNT(*) as c FROM applications WHERE job_id=%s", (j["job_id"],))
        j["applicant_count"] = cursor.fetchone()["c"]
   
    # Get custom columns dynamically
    cursor.execute("SHOW COLUMNS FROM jobs")
    all_columns = cursor.fetchall()
    custom_cols = [{"db_name": c["Field"], "label": c["Field"].replace("custom_", "").replace("_", " ").title()} for c in all_columns if c["Field"].startswith("custom_")]

    jobs_json = json.dumps([dict(j) for j in jobs], default=str)
    return render_template("faculty/jobs.html", jobs=jobs, jobs_json=jobs_json, custom_cols=custom_cols)


@app.route("/faculty/jobs/add", methods=["POST"])
def faculty_job_add():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    job_id    = request.form.get("job_id", "").strip()
    company   = request.form.get("company_name", "").strip()
    role      = request.form.get("role", "").strip()
    ctc       = request.form.get("ctc", "").strip()
    location  = request.form.get("location", "").strip()
    bond      = request.form.get("bond", "None").strip()
    cgpa      = float(request.form.get("cgpa_cutoff", 0))
    act_bl    = int(request.form.get("active_backlogs", 0))
    bl_hist   = int(request.form.get("backlog_history", 0))
    branches_list = request.form.getlist("branches")

    custom_branch = request.form.get("custom_branch", "").strip()
    if custom_branch:
        branches_list.append(custom_branch)

    branches = ", ".join(branches_list)

    tier      = request.form.get("tier", "Tier 1")
    desc      = request.form.get("description", "").strip()
    req_aadhar = 1 if request.form.get("req_aadhar") else 0
    req_pan    = 1 if request.form.get("req_pan") else 0
    req_other  = request.form.get("req_other", "").strip()
    pwd_only   = 1 if request.form.get("pwd_only") else 0
    deadline   = request.form.get("deadline", "").strip() or None

    custom_fields = [
        k for k in request.form.keys()
        if k.startswith("custom_") and k != "custom_branch"
    ]
    custom_cols_str = ", ".join(custom_fields)
    custom_placeholders = ", ".join(["%s"] * len(custom_fields))
    custom_values = [request.form.get(k, "").strip() for k in custom_fields]

    pdf_path = None
    pdf_file = request.files.get("pdf_file")
    if pdf_file and pdf_file.filename:
        active_year = session.get("active_year", "2025-2026")
        upload_dir = os.path.join(app.static_folder, "uploads", active_year, "job_pdfs")
        os.makedirs(upload_dir, exist_ok=True)
        fname = secure_filename(f"{job_id}_{pdf_file.filename}")
        pdf_file.save(os.path.join(upload_dir, fname))
        pdf_path = f"/static/uploads/{active_year}/job_pdfs/{fname}"

    try:
        col_sql = f", {custom_cols_str}" if custom_cols_str else ""
        val_sql = f", {custom_placeholders}" if custom_placeholders else ""
        cursor.execute(f"""
            INSERT INTO jobs (job_id, company_name, role, ctc, location, bond,
                cgpa_cutoff, active_backlogs, backlog_history, branches, tier,
                description, req_aadhar, req_pan, req_other, pwd_only, pdf_path, deadline{col_sql})
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s{val_sql})
        """, [job_id, company, role, ctc, location, bond,
               cgpa, act_bl, bl_hist, branches, tier,
               desc, req_aadhar, req_pan, req_other, pwd_only, pdf_path, deadline] + custom_values)
        db.commit()
        # ── In-app notification for all students ────────────────────────────
        notify_students_new_job(company, role)

        # ── Email notification (if enabled and checkbox checked) ─────────────
        send_email_flag = request.form.get("send_email_notification") == "1"
        if send_email_flag and _is_email_notifications_enabled():
            _send_new_job_emails_async(
                job_id=job_id, company=company, role=role, ctc=ctc,
                tier=tier, deadline=deadline or 'As announced',
                location=location, min_cgpa=cgpa, branches=branches
            )

        from flask import flash
        flash(f"Job opening for {company} added successfully!", "success")
    except Exception as e:
        db.rollback()
        from flask import flash
        flash(f"Error adding job: {str(e)}", "error")

    return redirect("/faculty/jobs")


@app.route("/faculty/jobs/edit", methods=["POST"])
def faculty_job_edit():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    db_job_id = request.form.get("job_id_edit", "").strip() or request.form.get("job_id", "").strip()
    company   = request.form.get("company_name", "").strip()
    role      = request.form.get("role", "").strip()
    ctc       = request.form.get("ctc", "").strip()
    location  = request.form.get("location", "").strip()
    bond      = request.form.get("bond", "None").strip()
    cgpa      = float(request.form.get("cgpa_cutoff", 0))
    act_bl    = int(request.form.get("active_backlogs", 0))
    bl_hist   = int(request.form.get("backlog_history", 0))
    branches_list = request.form.getlist("branches")

    custom_branch = request.form.get("custom_branch", "").strip()
    if custom_branch:
        branches_list.append(custom_branch)

    branches = ", ".join(branches_list)
   
    tier      = request.form.get("tier", "Tier 1")
    desc      = request.form.get("description", "").strip()
    req_aadhar = 1 if request.form.get("req_aadhar") else 0
    req_pan    = 1 if request.form.get("req_pan") else 0
    req_other  = request.form.get("req_other", "").strip()
    pwd_only   = 1 if request.form.get("pwd_only") else 0
    deadline   = request.form.get("deadline", "").strip() or None

    custom_fields = [
    k for k in request.form.keys()
    if k.startswith("custom_") and k != "custom_branch"
]
    custom_set_sql = "".join([f", {k}=%s" for k in custom_fields])
    custom_values = [request.form.get(k, "").strip() for k in custom_fields]

    # Check if a new PDF was uploaded
    pdf_file = request.files.get("pdf_file")
    pdf_update_sql = ""
    pdf_args = []
    if pdf_file and pdf_file.filename:
        active_year = session.get("active_year", "2025-2026")
        upload_dir = os.path.join(app.static_folder, "uploads", active_year, "job_pdfs")
        os.makedirs(upload_dir, exist_ok=True)
        fname = secure_filename(f"job_{db_job_id}_{pdf_file.filename}")
        pdf_file.save(os.path.join(upload_dir, fname))
        pdf_update_sql = ", pdf_path=%s"
        pdf_args = [f"/static/uploads/{active_year}/job_pdfs/{fname}"]

    try:
        cursor.execute(f"""
            UPDATE jobs SET company_name=%s, role=%s, ctc=%s, location=%s, bond=%s,
                cgpa_cutoff=%s, active_backlogs=%s, backlog_history=%s, branches=%s,
                tier=%s, description=%s, req_aadhar=%s, req_pan=%s, req_other=%s, deadline=%s
                {pdf_update_sql}
                {custom_set_sql}
            WHERE job_id = %s
        """, [company, role, ctc, location, bond,
               cgpa, act_bl, bl_hist, branches, tier,
               desc, req_aadhar, req_pan, req_other, deadline] + pdf_args + custom_values + [db_job_id])
       
        db.commit()

        # ── Email notification for job update ────────────────────────────────
        send_email_flag = request.form.get("send_email_notification") == "1"
        if send_email_flag and _is_email_notifications_enabled() and _EMAIL_SVC_LOADED and _email_svc.is_configured():
            try:
                # Notify students who applied OR are eligible
                cursor.execute("""
                    SELECT s.student_id, s.name, s.email
                    FROM applications a
                    JOIN students s ON a.student_id = s.student_id
                    WHERE a.job_id = %s AND s.email IS NOT NULL AND s.email != ''
                """, (db_job_id,))
                applied_students = cursor.fetchall()

                subject = f"Job Update: {company} – {role}"
                deadline_val = deadline or 'As announced'

                def html_fn_edit(rec):
                    return _email_svc.build_job_update_email(rec, company, role, ctc, deadline_val)

                def log_fn_edit(sid, email, status, err):
                    _log_email(email, None, 'job_update', subject, status, err)

                if applied_students:
                    _email_svc.send_bulk_emails_async(list(applied_students), subject, html_fn_edit, log_fn_edit, event_type='job_update')
            except Exception as email_ex:
                print(f"[EmailService] Error in job update email: {email_ex}")

        from flask import flash
        flash(f"Job for {company} updated successfully!", "success")
    except Exception as e:
        db.rollback()
        from flask import flash
        flash(f"Error updating job: {str(e)}", "error")

    return redirect("/faculty/jobs")


@app.route("/faculty/jobs/set_reminder", methods=["POST"])
def faculty_job_set_reminder():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    job_db_id = data.get("job_db_id")
    r_date = data.get("reminder_date")
    r_note = data.get("reminder_note")
    
    if not job_db_id:
        return jsonify({"success": False, "error": "Missing job ID"})
        
    try:
        if not r_date:
            r_date = None
            
        cursor.execute("""
            UPDATE jobs 
            SET reminder_date = %s, reminder_note = %s, reminder_sent = 0 
            WHERE id = %s
        """, (r_date, r_note, job_db_id))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/jobs/add_column", methods=["POST"])
def faculty_job_add_column():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir
   
    name = request.form.get("name", "").strip()
    col_type = request.form.get("type", "").strip()
   
    if not name:
        from flask import flash
        flash("Column name cannot be empty.", "error")
        return redirect("/faculty/jobs")
       
    import re
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', name.lower())
    col_name = f"custom_{sanitized}"
   
    sql_type = "VARCHAR(255)"
    if col_type == "number":
        sql_type = "DECIMAL(10,2)"
    elif col_type == "boolean":
        sql_type = "TINYINT(1)"
       
    try:
        cursor.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {sql_type}")
        db.commit()
        from flask import flash
        flash(f"Custom column '{name}' added successfully!", "success")
    except Exception as e:
        db.rollback()
        from flask import flash
        flash(f"Error adding column: {str(e)}", "error")
       
    return redirect("/faculty/jobs")

@app.route("/faculty/jobs/delete_all", methods=["POST"])
def faculty_jobs_delete_all():
    ensure_connection()
    redir = faculty_required()
    if redir:
        return redir

    try:
        cursor.execute("DELETE FROM applications")
        cursor.execute("DELETE FROM jobs")
        db.commit()
        flash("All job openings deleted successfully!", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error deleting job openings: {str(e)}", "error")

    return redirect("/faculty/jobs")


@app.route("/faculty/jobs/delete/<string:job_db_id>", methods=["POST"])
def faculty_job_delete(job_db_id):
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Not logged in"})
    try:
        # Fetch job info before deleting (for cancellation emails)
        cursor.execute("SELECT company_name, role, cgpa_cutoff, branches, tier FROM jobs WHERE job_id=%s", (job_db_id,))
        job_info = cursor.fetchone()

        # Fetch students who had applied
        cursor.execute("""
            SELECT s.student_id, s.name, s.email
            FROM applications a JOIN students s ON a.student_id=s.student_id
            WHERE a.job_id=%s AND s.email IS NOT NULL AND s.email!=''
        """, (job_db_id,))
        applied_students = cursor.fetchall()

        cursor.execute("DELETE FROM jobs WHERE job_id=%s", (job_db_id,))
        db.commit()

        # ── Send cancellation emails asynchronously ──────────────────────────
        if job_info and applied_students and _is_email_notifications_enabled() and _EMAIL_SVC_LOADED and _email_svc.is_configured():
            company_c = job_info['company_name']
            role_c    = job_info['role']
            subject_c = f"Drive Cancelled: {company_c} – {role_c}"

            def html_fn_cancel(rec):
                return _email_svc.build_job_cancelled_email(rec, company_c, role_c)

            def log_fn_cancel(sid, email, status, err):
                _log_email(email, None, 'job_cancelled', subject_c, status, err)

            _email_svc.send_bulk_emails_async(list(applied_students), subject_c, html_fn_cancel, log_fn_cancel, event_type='job_cancelled')

        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/jobs/delete_column/<string:col_name>", methods=["POST"])
def faculty_job_delete_column(col_name):
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Not logged in"})
    try:
        import re
        if not re.match(r'^custom_[a-zA-Z0-9_]+$', col_name):
            raise Exception("Invalid column name")
        cursor.execute(f"ALTER TABLE jobs DROP COLUMN {col_name}")
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/jobs/delete_column_by_name", methods=["POST"])
def faculty_job_delete_column_by_name():
    ensure_connection()
    redir = faculty_required()
    if redir:
        return redir
        
    name = request.form.get("name")
    if not name:
        from flask import flash
        flash("Column name is required.", "error")
        return redirect("/faculty/jobs")
        
    import re
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', name.lower())
    db_col_name = f"custom_{sanitized}"
    
    try:
        cursor.execute("SHOW COLUMNS FROM jobs LIKE %s", (db_col_name,))
        col = cursor.fetchone()
        if not col:
            from flask import flash
            flash(f"Criteria column '{name}' does not exist.", "error")
            return redirect("/faculty/jobs")
            
        cursor.execute(f"ALTER TABLE jobs DROP COLUMN {db_col_name}")
        db.commit()
        from flask import flash
        flash(f"Criteria column '{name}' removed successfully!", "success")
    except Exception as e:
        db.rollback()
        from flask import flash
        flash(f"Error removing column: {str(e)}", "error")
        
    return redirect("/faculty/jobs")


@app.route("/faculty/jobs/applicants/<string:job_db_id>")
def faculty_job_applicants(job_db_id):
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"applicants": []})

    cursor.execute("""
        SELECT a.*, s.name, s.branch, s.student_id
        FROM applications a
        JOIN students s ON a.student_id = s.student_id
        WHERE a.job_id = %s
    """, (job_db_id,))
    apps = cursor.fetchall()
    return jsonify({"applicants": [dict(a) for a in apps]})


@app.route("/faculty/job_pdf/<string:job_db_id>")
def faculty_job_pdf(job_db_id):
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    cursor.execute("SELECT pdf_path FROM jobs WHERE job_id=%s", (job_db_id,))
    job = cursor.fetchone()
    if not job or not job["pdf_path"]:
        from flask import flash
        flash("No PDF attached to this job.", "error")
        return redirect("/faculty/jobs")
   
    # Strip the leading '/static/' to get the true relative path in the static folder
    pdf_path = job["pdf_path"]
    if pdf_path.startswith("/static/"):
        pdf_path = pdf_path[8:]
    from flask import send_from_directory
    return send_from_directory(app.static_folder, pdf_path)


@app.route("/faculty/applications/update_status", methods=["POST"])
def faculty_applications_update_status():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Faculty authorization required"})

    data = request.get_json() or {}
    app_id = data.get("application_id")
    status = data.get("status")
    # Normalize 'Not Selected' to 'Not Selected' (was 'Rejected')
    if status == 'Rejected':
        status = 'Not Selected'

    if not app_id or not status:
        return jsonify({"success": False, "error": "Missing application_id or status"})

    try:
        # Fetch current details before updating
        cursor.execute("SELECT student_id, job_id, status FROM applications WHERE application_id = %s", (app_id,))
        app_record = cursor.fetchone()
        if not app_record:
            return jsonify({"success": False, "error": "Application not found"})

        student_id = app_record["student_id"]
        job_db_id = app_record["job_id"]
        old_status = app_record["status"]

        drive_link = data.get("drive_link")
        
        # Update applications table
        if drive_link:
            cursor.execute("UPDATE applications SET status = %s, drive_link = %s, status_updated_at = NOW() WHERE application_id = %s", (status, drive_link, app_id))
        else:
            cursor.execute("UPDATE applications SET status = %s, status_updated_at = NOW() WHERE application_id = %s", (status, app_id))

        # Check the job company and role/tier
        cursor.execute("SELECT company_name, role, tier FROM jobs WHERE job_id = %s", (job_db_id,))
        job_details = cursor.fetchone()
        company_name = job_details["company_name"] if job_details else "Company"
        role_name = job_details["role"] if job_details else "Role"
        job_tier_str = job_details["tier"] if job_details else "Tier 3"
        job_tier_num = 1 if '1' in job_tier_str else (2 if '2' in job_tier_str else 3)

        
        # Notify student if status changed or drive link provided
        if old_status != status or drive_link:
            message = f"Your application status for {company_name} - {role_name} has been updated to {status}."
            if status == "Selected":
                message = f"Congratulations! You have been Selected by {company_name} for the {role_name} role (Tier {job_tier_str})!"
            
            cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                           (student_id, message, "/my_applications"))

            # ── Send status-update email to this student ──────────────────────
            if _is_email_notifications_enabled() and _EMAIL_SVC_LOADED and _email_svc.is_configured():
                try:
                    cursor.execute("SELECT name, email FROM students WHERE student_id=%s", (student_id,))
                    stu = cursor.fetchone()
                    if stu and stu.get('email'):
                        subject_su = f"Application Update: {company_name} – {status}"
                        rec_su = {'student_id': student_id, 'name': stu['name'], 'email': stu['email']}

                        def html_fn_status(r):
                            return _email_svc.build_status_update_email(r, company_name, role_name, status, drive_link)

                        def log_fn_status(sid, email, st, err):
                            _log_email(email, stu['name'], 'status_update', subject_su, st, err)

                        _email_svc.send_bulk_emails_async([rec_su], subject_su, html_fn_status, log_fn_status, event_type='status_update')
                except Exception as email_ex:
                    print(f"[EmailService] Error sending status email: {email_ex}")

        # Re-evaluate the student selected_tier:
        # Find the highest tier level among all 'Selected' applications for this student
        cursor.execute("""
            SELECT j.tier, j.company_name
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            WHERE a.student_id = %s AND a.status = 'Selected'
        """, (student_id,))
        selected_apps = cursor.fetchall()
       
        if selected_apps:
            highest_tier_num = 3
            tier_1_val, tier_2_val, tier_3_val = None, None, None
            for sa in selected_apps:
                t_str = sa["tier"] or "Tier 3"
                c_name = sa["company_name"]
                t_num = 1 if '1' in t_str else (2 if '2' in t_str else 3)
                if t_num < highest_tier_num:
                    highest_tier_num = t_num
                    
                if t_num == 1:
                    tier_1_val = c_name
                elif t_num == 2:
                    tier_2_val = c_name
                elif t_num == 3:
                    tier_3_val = c_name
            cursor.execute("UPDATE students SET selected_tier = %s, tier_1 = %s, tier_2 = %s, tier_3 = %s WHERE student_id = %s", (highest_tier_num, tier_1_val, tier_2_val, tier_3_val, student_id))
        else:
            # Clear selected tier
            cursor.execute("UPDATE students SET selected_tier = NULL, tier_1 = NULL, tier_2 = NULL, tier_3 = NULL WHERE student_id = %s", (student_id,))

        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/upload_selected_students_excel", methods=["POST"])
def faculty_upload_selected_students_excel():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No file selected"})
        
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
            
        required_cols = ['Roll Number', 'Job ID', 'Company Name', 'CTC']
        for col in required_cols:
            if col not in df.columns:
                return jsonify({"success": False, "error": f"Missing required column: {col}"})
                
        updated_count = 0
        for index, row in df.iterrows():
            roll_number = str(row['Roll Number']).strip()
            job_ids_raw = str(row['Job ID']).strip()
            
            if not roll_number or roll_number.lower() in ('nan', 'none', '') or \
               not job_ids_raw or job_ids_raw.lower() in ('nan', 'none', ''):
                continue
                
            cursor.execute("SELECT student_id FROM students WHERE roll_number = %s", (roll_number,))
            student = cursor.fetchone()
            if not student:
                continue
            student_id = student['student_id']

            job_ids = [x.strip() for x in job_ids_raw.split(',')]
            companies = [x.strip() for x in str(row['Company Name']).split(',')] if pd.notna(row['Company Name']) else []
            ctcs = [x.strip() for x in str(row['CTC']).split(',')] if pd.notna(row['CTC']) else []
            
            # pad companies and ctcs to match job_ids length
            while len(companies) < len(job_ids): companies.append("Unknown")
            while len(ctcs) < len(job_ids): ctcs.append("0")
            
            has_tier_cols = 'Tier 1' in df.columns or 'Tier 2' in df.columns or 'Tier 3' in df.columns
            t1_str = str(row.get('Tier 1', '')).strip().lower() if 'Tier 1' in df.columns else ''
            t2_str = str(row.get('Tier 2', '')).strip().lower() if 'Tier 2' in df.columns else ''
            t3_str = str(row.get('Tier 3', '')).strip().lower() if 'Tier 3' in df.columns else ''
            
            for job_id, company_name, ctc_val in zip(job_ids, companies, ctcs):
                is_selected_in_excel = False
                tier = None
                c_lower = company_name.lower()
                
                if has_tier_cols:
                    if (c_lower and c_lower in t1_str) or t1_str in ('yes', '1', 'true', '✓', 'tick', 'y'):
                        tier = 'Tier 1'
                        is_selected_in_excel = True
                    elif (c_lower and c_lower in t2_str) or t2_str in ('yes', '1', 'true', '✓', 'tick', 'y'):
                        tier = 'Tier 2'
                        is_selected_in_excel = True
                    elif (c_lower and c_lower in t3_str) or t3_str in ('yes', '1', 'true', '✓', 'tick', 'y'):
                        tier = 'Tier 3'
                        is_selected_in_excel = True

                if has_tier_cols and not is_selected_in_excel:
                    cursor.execute("UPDATE applications SET status = 'Not Selected' WHERE student_id = %s AND job_id = %s", (student_id, job_id))
                    db.commit()
                    
                    cursor.execute("""
                        SELECT j.tier
                        FROM applications a
                        JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.student_id = %s AND a.status = 'Selected'
                    """, (student_id,))
                    selected_apps = cursor.fetchall()
                    
                    highest_tier_num = 3
                    if selected_apps:
                        for sa in selected_apps:
                            t_str = sa["tier"] or "Tier 3"
                            t_num = 1 if '1' in t_str else (2 if '2' in t_str else 3)
                            if t_num < highest_tier_num:
                                highest_tier_num = t_num
                    cursor.execute("UPDATE students SET selected_tier = %s WHERE student_id = %s", (highest_tier_num, student_id))
                    db.commit()
                    updated_count += 1
                    continue
                    
                if not has_tier_cols:
                    ctc_num = 0.0
                    try:
                        import re
                        match = re.search(r'(\d+(\.\d+)?)', str(ctc_val).lower())
                        if match: ctc_num = float(match.group(1))
                    except: pass
                    
                    if ctc_num < 7.0: tier = 'Tier 3'
                    elif 7.0 <= ctc_num <= 17.0: tier = 'Tier 2'
                    else: tier = 'Tier 1'
                    
                ctc_str = str(ctc_val)
                if ctc_str.replace('.', '', 1).isdigit(): ctc_str += " LPA"
                
                cursor.execute("SELECT id FROM jobs WHERE job_id = %s", (job_id,))
                if not cursor.fetchone():
                    cursor.execute("INSERT INTO jobs (job_id, company_name, ctc, tier, role) VALUES (%s, %s, %s, %s, %s)",
                                   (job_id, company_name, ctc_str, tier, "Selected Role"))
                else:
                    cursor.execute("UPDATE jobs SET tier = %s WHERE job_id = %s", (tier, job_id))
                db.commit()
                    
                cursor.execute("SELECT application_id, status FROM applications WHERE student_id = %s AND job_id = %s", (student_id, job_id))
                app = cursor.fetchone()
                was_selected = app and app['status'] == 'Selected'
                
                if app:
                    cursor.execute("UPDATE applications SET status = 'Selected' WHERE application_id = %s", (app['application_id'],))
                else:
                    cursor.execute("SELECT MAX(application_id) as max_id FROM applications")
                    max_row = cursor.fetchone()
                    next_id = (max_row['max_id'] or 0) + 1
                    from datetime import datetime
                    today = datetime.today().strftime('%Y-%m-%d')
                    cursor.execute("INSERT INTO applications (application_id, student_id, job_id, status, applied_date) VALUES (%s, %s, %s, %s, %s)",
                                   (next_id, student_id, job_id, 'Selected', today))
                db.commit()
                
                cursor.execute("""
                    SELECT j.tier
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.student_id = %s AND a.status = 'Selected'
                """, (student_id,))
                selected_apps = cursor.fetchall()
               
                if selected_apps:
                    highest_tier_num = 3
                    for sa in selected_apps:
                        t_str = sa["tier"] or "Tier 3"
                        t_num = 1 if '1' in t_str else (2 if '2' in t_str else 3)
                        if t_num < highest_tier_num:
                            highest_tier_num = t_num
                    cursor.execute("UPDATE students SET selected_tier = %s WHERE student_id = %s", (highest_tier_num, student_id))
                
                if not was_selected:
                    tier_label = tier.replace('Tier ', 'Tier ')
                    notif_msg = f"🏆 Congratulations! You have been <strong>Selected</strong> by <strong>{company_name}</strong> under <strong>{tier_label}</strong> (CTC: {ctc_val} LPA). Your tier eligibility has been updated."
                    cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                                   (student_id, notif_msg, "/my_applications"))
                db.commit()
                updated_count += 1
                
        return jsonify({"success": True, "message": f"Successfully updated {updated_count} records."})
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route("/faculty/download_selected_students_template")
def faculty_download_selected_students_template():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    course_filter = request.args.get("course", "all").lower()
    branch_filter = request.args.get("branch", "all").lower()

    try:
        import io
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        query = """
            SELECT s.roll_number, s.name, s.course, s.branch,
                   j.job_id, j.company_name, j.ctc, j.tier
            FROM students s
            JOIN applications a ON s.student_id = a.student_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE a.status = 'Selected'
            ORDER BY s.name ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        # Build data rows
        from collections import defaultdict
        student_groups = defaultdict(list)
        for r in rows:
            c = (r['course'] or 'b.tech').lower()
            b = (r['branch'] or '').strip()
            c_match = (course_filter == 'all') or (c == course_filter)
            b_match = True if branch_filter == 'all' else (normalize_branch(b) == normalize_branch(branch_filter))
            if not (c_match and b_match):
                continue
            student_groups[r['roll_number']].append(r)
            
        data_rows = []
        for roll_no, apps in student_groups.items():
            first = apps[0]
            job_ids = [str(a['job_id']) for a in apps]
            ctcs = [str(a['ctc']) for a in apps if a.get('ctc')]
            
            t1, t2, t3 = '', '', ''
            for a in apps:
                tier = str(a['tier'] or '')
                c_name = str(a['company_name'] or '')
                if '1' in tier:
                    t1 = c_name if not t1 else t1 + ", " + c_name
                elif '2' in tier:
                    t2 = c_name if not t2 else t2 + ", " + c_name
                else:
                    t3 = c_name if not t3 else t3 + ", " + c_name
                    
            data_rows.append([
                len(data_rows) + 1,
                first['roll_number'],
                first['name'],
                first['course'],
                first['branch'],
                ", ".join(job_ids),
                ", ".join([str(a['company_name']) for a in apps]),
                ", ".join(ctcs),
                t1, t2, t3
            ])

        # Add 5 blank rows at the end for admin to fill new entries
        for _ in range(5):
            data_rows.append([''] * 11)

        # Build workbook with styling
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Selected Students'

        headers = ['S.No', 'Roll Number', 'Name', 'Course', 'Branch',
                   'Job ID', 'Company Name', 'CTC', 'Tier 1', 'Tier 2', 'Tier 3']

        # Header style
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='B8540A')   # dark amber
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_widths = [6, 14, 22, 10, 28, 10, 22, 8, 8, 8, 8]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 20

        # Data rows
        data_font = Font(size=10)
        data_center = Alignment(horizontal='center', vertical='center')
        alt_fill = PatternFill(fill_type='solid', fgColor='FFF8EE')

        for row_idx, row_data in enumerate(data_rows, start=2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_center
                cell.border = border
                cell.fill = fill

        ws.freeze_panes = 'A2'  # freeze header row

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="Selected_Students.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback; traceback.print_exc()
        return f"Error creating Excel: {str(e)}", 500

@app.route("/faculty/selected_students")
def faculty_selected_students():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    # Query all students who are selected for any job
    query = """
        SELECT s.student_id, s.name, s.email, s.branch, s.course, s.roll_number, s.phone_number,
               a.status, j.company_name, j.tier, j.job_id, j.ctc
        FROM students s
        JOIN applications a ON s.student_id = a.student_id
        JOIN jobs j ON a.job_id = j.job_id
        WHERE a.status = 'Selected'
        ORDER BY s.student_id ASC
    """
    cursor.execute(query)
    placed = cursor.fetchall()

    return render_template("faculty/selected_students.html", placed_students=placed)


@app.route("/faculty/applied_students")
def faculty_applied_students():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    try:
        cursor.execute("SELECT * FROM jobs ORDER BY id DESC")
    except Exception:
        cursor.execute("SELECT * FROM jobs ORDER BY job_id DESC")
    jobs = cursor.fetchall()

    # Format deadlines and query applicants count
    jobs_list = []
    for job in jobs:
        j = dict(job)
        # Check if deadline passed
        from datetime import datetime, timedelta
        is_passed = False
        hide_job = False
        if j.get('deadline'):
            deadline = j['deadline']
            if isinstance(deadline, str):
                try:
                    deadline = datetime.strptime(deadline, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    try:
                        deadline = datetime.strptime(deadline, "%Y-%m-%dT%H:%M")
                    except Exception:
                        pass
            if isinstance(deadline, datetime):
                if datetime.now() > deadline:
                    is_passed = True
                if datetime.now() > deadline + timedelta(days=3):
                    hide_job = True
       
        if hide_job:
            continue
            
        j['is_deadline_passed'] = is_passed
        j['deadline_str'] = str(j.get('deadline')) if j.get('deadline') else 'Ongoing'

        # Query applicants details
        cursor.execute("""
            SELECT s.student_id, s.name, s.branch, s.roll_number, s.phone_number, s.email, s.aadhar, s.pan, a.resume_path, a.applied_date, a.extra_details
            FROM applications a
            JOIN students s ON a.student_id = s.student_id
            WHERE a.job_id = %s
            ORDER BY a.applied_date DESC
        """, (j['job_id'],))
        applicants = cursor.fetchall()
        for applicant in applicants:
            if applicant.get('extra_details'):
                try:
                    applicant['extra_dict'] = json.loads(applicant['extra_details'])
                except:
                    applicant['extra_dict'] = {}
            else:
                applicant['extra_dict'] = {}
               
        j['applicants'] = applicants
        jobs_list.append(j)

    return render_template("faculty/applied_students.html", jobs=jobs_list)


@app.route("/faculty/jobs/archive_recruitment", methods=["POST"])
def faculty_archive_recruitment():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    job_id = data.get("job_id")
    action = data.get("action") # 'yes' or 'no'
    
    if not job_id:
        return jsonify({"success": False, "error": "Missing job ID"})
        
    archived_val = 1 if action == 'yes' else -1
    
    try:
        cursor.execute("UPDATE jobs SET recruitment_archived = %s WHERE id = %s", (archived_val, job_id))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})

@app.route("/faculty/job_results")
def faculty_job_results():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    try:
        cursor.execute("SELECT * FROM jobs ORDER BY id DESC")
    except Exception:
        cursor.execute("SELECT * FROM jobs ORDER BY job_id DESC")
    jobs = cursor.fetchall()
    
    for j in jobs:
        cursor.execute("SELECT COUNT(*) as c FROM applications WHERE job_id=%s", (j["job_id"],))
        j["applicant_count"] = cursor.fetchone()["c"]

    return render_template("faculty/job_results.html", jobs=jobs)


@app.route("/faculty/job_analysis")
def faculty_job_analysis():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    try:
        cursor.execute("""
            SELECT
                j.job_id,
                j.company_name,
                j.role,
                j.tier,
                j.ctc,
                j.deadline,
                COUNT(a.student_id) AS total_applied,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(a.status,''))) = 'selected' THEN 1 ELSE 0 END) AS total_selected,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(a.status,''))) IN ('rejected','not selected') THEN 1 ELSE 0 END) AS total_rejected,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(a.status,''))) IN ('applied','pending','interview','under review') THEN 1 ELSE 0 END) AS total_pending
            FROM jobs j
            LEFT JOIN applications a ON j.job_id = a.job_id
            GROUP BY j.job_id, j.company_name, j.role, j.tier, j.ctc, j.deadline
            ORDER BY j.company_name, j.role
        """)
        jobs_analysis = cursor.fetchall()

        # Compute totals
        grand_applied = sum((r["total_applied"] or 0) for r in jobs_analysis)
        grand_selected = sum((r["total_selected"] or 0) for r in jobs_analysis)

        # Compute per-job selection rate
        analysis_list = []
        for row in jobs_analysis:
            applied = row["total_applied"] or 0
            selected = row["total_selected"] or 0
            rejected = row["total_rejected"] or 0
            pending = row["total_pending"] or 0
            rate = round((selected / applied * 100), 1) if applied > 0 else 0.0
            analysis_list.append({
                "job_id": row["job_id"],
                "company_name": row["company_name"],
                "role": row["role"],
                "tier": row["tier"] or "—",
                "ctc": row["ctc"] or "—",
                "deadline": str(row["deadline"]) if row["deadline"] else "Ongoing",
                "total_applied": applied,
                "total_selected": selected,
                "total_rejected": rejected,
                "total_pending": pending,
                "selection_rate": rate,
            })

    except Exception as e:
        print("Job analysis error:", e)
        analysis_list = []
        grand_applied = 0
        grand_selected = 0

    grand_rate = round((grand_selected / grand_applied * 100), 1) if grand_applied > 0 else 0.0

    return render_template(
        "faculty/job_analysis.html",
        analysis_list=analysis_list,
        grand_applied=grand_applied,
        grand_selected=grand_selected,
        grand_rate=grand_rate,
        total_jobs=len(analysis_list),
    )


# ─── FACULTY: RECRUITMENT PROCESS TRACKER ─────────────────────────────────────

@app.route("/faculty/recruitment_process")
def faculty_recruitment_process():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    try:
        cursor.execute("SELECT * FROM jobs ORDER BY id DESC")
    except Exception:
        cursor.execute("SELECT * FROM jobs ORDER BY job_id DESC")
    jobs = cursor.fetchall()

    jobs_data = []
    for j in jobs:
        jdict = dict(j)
        # Get number of rounds configured for this job
        cursor.execute("SELECT num_rounds FROM recruitment_rounds WHERE job_id=%s", (j["job_id"],))
        rr = cursor.fetchone()
        jdict["num_rounds"] = rr["num_rounds"] if rr else 1

        # Get all applicants for this job
        cursor.execute("""
            SELECT s.student_id, s.name, s.branch, s.roll_number, s.email, s.phone_number
            FROM applications a
            JOIN students s ON a.student_id = s.student_id
            WHERE a.job_id = %s
            ORDER BY s.name ASC
        """, (j["job_id"],))
        students = cursor.fetchall()

        # Get round results for all students in this job (include drive_link)
        cursor.execute("""
            SELECT student_id, round_number, result, drive_link
            FROM round_results
            WHERE job_id=%s
        """, (j["job_id"],))
        rr_rows = cursor.fetchall()
        # Build dicts: {student_id: {round_number: result}} and {student_id: {round_number: drive_link}}
        round_map = {}
        link_map  = {}
        for rrow in rr_rows:
            sid = rrow["student_id"]
            rnd = rrow["round_number"]
            round_map.setdefault(sid, {})[rnd] = rrow["result"]
            link_map.setdefault(sid,  {})[rnd] = rrow["drive_link"] or ""

        # Check if recruitment is completed
        has_final_round = False
        final_complete = True
        
        # Filter out eliminated students to keep the tracker clean
        students_list = []
        for s in students:
            sid = s["student_id"]
            s_dict = dict(s)
            s_dict["rounds"] = {}
            s_dict["links"]  = {}
            
            is_eliminated_before_final = False
            for rnd in range(1, jdict["num_rounds"]):
                if round_map.get(sid, {}).get(rnd) == "Not Selected":
                    is_eliminated_before_final = True
                    break
                    
            if not is_eliminated_before_final:
                has_final_round = True
                f_res = round_map.get(sid, {}).get(jdict["num_rounds"], "Pending")
                if f_res == "Pending":
                    final_complete = False
            
            is_eliminated = False
            for rnd in range(1, jdict["num_rounds"] + 1):
                s_dict["rounds"][rnd] = round_map.get(sid, {}).get(rnd, "Pending")
                s_dict["links"][rnd]  = link_map.get(sid,  {}).get(rnd, "")
                if s_dict["rounds"][rnd] == "Not Selected":
                    is_eliminated = True
                    
            if not is_eliminated:
                students_list.append(s_dict)
                
        # If final round complete, update finished_at
        if has_final_round and final_complete and not jdict.get("recruitment_finished_at"):
            try:
                cursor.execute("UPDATE jobs SET recruitment_finished_at = NOW() WHERE id = %s", (jdict["id"],))
                db.commit()
                from datetime import datetime
                jdict["recruitment_finished_at"] = datetime.now()
            except Exception:
                pass
                
        # Check archive logic
        prompt_archive = False
        archived_val = jdict.get("recruitment_archived", 0)
        
        if archived_val == 1:
            continue # hide from list
            
        if archived_val == 0 and jdict.get("recruitment_finished_at"):
            from datetime import datetime, timedelta
            f_time = jdict["recruitment_finished_at"]
            if isinstance(f_time, str):
                try: f_time = datetime.strptime(f_time, "%Y-%m-%d %H:%M:%S")
                except:
                    try: f_time = datetime.strptime(f_time, "%Y-%m-%dT%H:%M")
                    except: pass
            if isinstance(f_time, datetime) and datetime.now() > f_time + timedelta(days=3):
                prompt_archive = True
                
        jdict["prompt_archive"] = prompt_archive

        jdict["students"] = students_list
        jdict["applicant_count"] = len(list(students))
        jobs_data.append(jdict)

    return render_template("faculty/recruitment_process.html", jobs=jobs_data)


@app.route("/faculty/recruitment_process/set_rounds", methods=["POST"])
def set_recruitment_rounds():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Not logged in"})
    data = request.get_json() or {}
    job_id = data.get("job_id")
    num_rounds = int(data.get("num_rounds", 1))
    try:
        cursor.execute("""
            INSERT INTO recruitment_rounds (job_id, num_rounds)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE num_rounds = %s
        """, (job_id, num_rounds, num_rounds))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/recruitment_process/set_result", methods=["POST"])
def set_round_result():
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Not logged in"})
    data = request.get_json() or {}
    job_id = data.get("job_id")
    student_id = data.get("student_id")
    round_number = int(data.get("round_number", 1))
    result = data.get("result", "Pending")  # 'Selected', 'Not Selected', 'Pending'
    drive_link = data.get("drive_link")
    try:
        # Check previous result
        cursor.execute("SELECT result, drive_link FROM round_results WHERE job_id=%s AND student_id=%s AND round_number=%s", (job_id, student_id, round_number))
        prev = cursor.fetchone()
        prev_result = prev["result"] if prev else None
        prev_link = prev["drive_link"] if prev else None
        
        # If nothing changed, we can skip inserting duplicate notifications
        changed = (prev_result != result) or (str(prev_link or "").strip() != str(drive_link or "").strip())

        cursor.execute("""
            INSERT INTO round_results (job_id, student_id, round_number, result, drive_link)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE result = VALUES(result), drive_link = VALUES(drive_link)
        """, (job_id, student_id, round_number, result, drive_link))

        if changed:
            # Fetch job details for notification message
            cursor.execute("SELECT company_name, role FROM jobs WHERE job_id = %s", (job_id,))
            job_det = cursor.fetchone()
            company_name = job_det["company_name"] if job_det else f"Job {job_id}"
            role_name = job_det["role"] if job_det else "Role"
    
            # Create notification for the student
            if result == "Selected":
                msg = f"🎉 You have been <strong>Qualified for Round {round_number}</strong> of <strong>{company_name} - {role_name}</strong>!"
                # Update application status to reflect progress
                cursor.execute("""
                    UPDATE applications SET status = 'Shortlisted'
                    WHERE student_id = %s AND job_id = %s AND status NOT IN ('Selected', 'Not Selected')
                """, (student_id, job_id))
            elif result == "Not Selected":
                msg = f"Your application for <strong>{company_name} - {role_name}</strong> has been updated: <strong>Not Selected</strong> in Round {round_number}."
                # Update application status
                cursor.execute("""
                    UPDATE applications SET status = 'Not Selected'
                    WHERE student_id = %s AND job_id = %s
                """, (student_id, job_id))
            else:
                msg = f"Your Round {round_number} status for <strong>{company_name} - {role_name}</strong> is being reviewed."
    
            cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                           (student_id, msg, "/my_applications"))

        # Check if all rounds are cleared for auto-finalise
        cursor.execute("SELECT num_rounds FROM recruitment_rounds WHERE job_id=%s", (job_id,))
        rr = cursor.fetchone()
        num_rounds = rr["num_rounds"] if rr else 1

        cursor.execute("SELECT round_number, result FROM round_results WHERE job_id=%s AND student_id=%s", (job_id, student_id))
        student_rounds = cursor.fetchall()
        round_map = {r["round_number"]: r["result"] for r in student_rounds}

        all_selected = all(round_map.get(r, "Pending") == "Selected" for r in range(1, num_rounds + 1))
        
        cursor.execute("SELECT status FROM applications WHERE student_id=%s AND job_id=%s", (student_id, job_id))
        cur_app = cursor.fetchone()
        current_status = cur_app["status"] if cur_app else "Pending"

        if all_selected and len(round_map) == num_rounds and current_status not in ('Selected', 'Rounds Cleared', 'Not Selected'):
            cursor.execute("UPDATE applications SET status='Rounds Cleared', status_updated_at=NOW() WHERE student_id=%s AND job_id=%s", (student_id, job_id))
            final_msg = f"🎉 Well done! You have cleared all <strong>{num_rounds} rounds</strong> for <strong>{company_name} – {role_name}</strong>. Your selection will be confirmed once the final results are published by the placement office."
            cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                           (student_id, final_msg, "/my_applications"))

        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/recruitment_process/export_excel/<string:job_id>")
def export_recruitment_excel(job_id):
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    round_num = request.args.get('round', type=int)

    cursor.execute("SELECT company_name, role FROM jobs WHERE job_id = %s", (job_id,))
    job = cursor.fetchone()
    if not job:
        return "Job not found", 404

    company_name = job["company_name"]
    role_name = job["role"]

    cursor.execute("SELECT num_rounds FROM recruitment_rounds WHERE job_id=%s", (job_id,))
    rr = cursor.fetchone()
    num_rounds = rr["num_rounds"] if rr else 1

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    amber_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    blue_fill  = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    red_fill   = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    white_font   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font   = Font(name="Calibri", size=14, bold=True, color="78350F")
    regular_font = Font(name="Calibri", size=11)
    instr_font   = Font(name="Calibri", size=10, italic=True, color="6B7280")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin_border  = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),  bottom=Side(style='thin', color='E5E7EB')
    )

    # Fetch all applicants
    cursor.execute("""
        SELECT s.student_id, s.name, s.branch, s.roll_number, s.email, s.phone_number
        FROM applications a
        JOIN students s ON a.student_id = s.student_id
        WHERE a.job_id = %s
        ORDER BY s.name ASC
    """, (job_id,))
    all_students = cursor.fetchall()

    # Fetch existing round results (including drive_link)
    cursor.execute("SELECT student_id, round_number, result, drive_link FROM round_results WHERE job_id=%s", (job_id,))
    rr_rows = cursor.fetchall()
    round_map = {}
    link_map  = {}
    for rrow in rr_rows:
        sid = rrow["student_id"]
        rnd = rrow["round_number"]
        round_map.setdefault(sid, {})[rnd] = rrow["result"]
        link_map.setdefault(sid,  {})[rnd] = rrow["drive_link"] or ""

    # ── PER-ROUND EXPORT ─────────────────────────────────────────────────────────
    if round_num:
        # Filter eligible students: selected in all prior rounds (or all for round 1)
        eligible = []
        for s in all_students:
            sid = s["student_id"]
            ok = True
            for prev in range(1, round_num):
                if round_map.get(sid, {}).get(prev, "Pending") == "Not Selected":
                    ok = False
                    break
            if ok:
                eligible.append(s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Round {round_num}"

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws.cell(1, 1).value = f"Round {round_num} — {company_name} ({role_name})"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = left_align
        ws.row_dimensions[1].height = 30

        # Instructions
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
        ws.cell(2, 1).value = (
            f"INSTRUCTIONS: In column 'Round {round_num}', enter {round_num} if student is selected, "
            f"0 if not selected. Paste the Google Drive / interview link in 'Drive Link' column for selected students."
        )
        ws.cell(2, 1).font = instr_font
        ws.cell(2, 1).alignment = left_align
        ws.row_dimensions[2].height = 22

        # Headers (row 3)
        hdrs = [
            "S.No", "Student ID", "Roll No", "Student Name", "Branch", "Email", "Phone",
            f"Round {round_num}  (enter {round_num}=selected / 0=not selected)",
            "Drive Link  (paste here for selected students)"
        ]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(3, ci, h)
            c.font  = white_font
            c.fill  = amber_fill if ci <= 7 else blue_fill
            c.alignment = center_align
            c.border = thin_border
        ws.row_dimensions[3].height = 36

        # Data rows
        for ri, s in enumerate(eligible, 1):
            row_num = ri + 3
            sid = s["student_id"]
            existing_res  = round_map.get(sid, {}).get(round_num, "")
            existing_link = link_map.get(sid, {}).get(round_num, "")
            if existing_res == "Selected":
                rval = round_num
            elif existing_res == "Not Selected":
                rval = 0
            else:
                rval = ""
            vals = [ri, sid, s["roll_number"] or "—", s["name"], s["branch"],
                    s["email"], s["phone_number"] or "—", rval, existing_link]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row_num, ci, v)
                c.font  = regular_font
                c.alignment = left_align if ci in (4, 6, 9) else center_align
                c.border = thin_border

        # Column widths
        for ci, w in enumerate([6, 12, 15, 28, 16, 32, 14, 44, 52], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        from flask import send_file
        clean = "".join([ch for ch in company_name if ch.isalnum() or ch in (' ', '_')]).strip()
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name=f"Round{round_num}_{clean}_{job_id}.xlsx")

    # ── FULL SUMMARY EXPORT (no round param) ─────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recruitment Process"

    # num_cols = 6 base + (Round + Drive Link) * num_rounds
    num_cols = 6 + num_rounds * 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.cell(1, 1).value = f"Recruitment Process — {company_name} ({role_name})"
    ws.cell(1, 1).font  = title_font
    ws.cell(1, 1).alignment = left_align
    ws.row_dimensions[1].height = 30

    base_headers = ["S.No", "Roll Number", "Student Name", "Branch", "Email", "Phone"]
    round_headers = []
    for i in range(1, num_rounds + 1):
        round_headers.append(f"Round {i}")
        round_headers.append(f"Drive Link {i}")
    all_headers = base_headers + round_headers

    for ci, h in enumerate(all_headers, 1):
        c = ws.cell(3, ci, h)
        c.font  = white_font
        c.fill  = amber_fill
        c.alignment = center_align
        c.border = thin_border
    ws.row_dimensions[3].height = 24

    for ri, s in enumerate(all_students, 1):
        row_num = ri + 3
        sid = s["student_id"]
        vals = [ri, s["roll_number"] or "—", s["name"], s["branch"], s["email"], s["phone_number"] or "—"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row_num, ci, v)
            c.font  = regular_font
            c.alignment = left_align if ci in (3, 5) else center_align
            c.border = thin_border
        for rnd in range(1, num_rounds + 1):
            res  = round_map.get(sid, {}).get(rnd, "Pending")
            link = link_map.get(sid, {}).get(rnd, "")
            ci_res  = 6 + (rnd - 1) * 2 + 1
            ci_link = ci_res + 1
            c = ws.cell(row_num, ci_res, res)
            if res == "Selected":
                c.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                c.fill  = green_fill
            elif res == "Not Selected":
                c.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                c.fill  = red_fill
            else:
                c.font  = regular_font
                c.fill  = PatternFill()
            c.alignment = center_align
            c.border = thin_border
            cl = ws.cell(row_num, ci_link, link or "")
            cl.font  = regular_font
            cl.alignment = left_align
            cl.border = thin_border

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col if c.row >= 3), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    from flask import send_file
    clean = "".join([ch for ch in company_name if ch.isalnum() or ch in (' ', '_')]).strip()
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"RecruitmentProcess_{clean}_{job_id}.xlsx")


@app.route("/faculty/recruitment_process/upload_excel/<string:job_id>/<int:round_num>", methods=["POST"])
def upload_recruitment_excel(job_id, round_num):
    """Parse an uploaded Round-N Excel and update round_results + auto-finalise application status."""
    ensure_connection()
    redir = faculty_required()
    if redir: return jsonify({"success": False, "error": "Not logged in"})

    if 'excel_file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"})

    file = request.files['excel_file']
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({"success": False, "error": "Please upload a .xlsx file"})

    import openpyxl, io as _io

    try:
        content = file.read()
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        ws = wb.active

        # Locate header row (row 3) and map column names → 1-based column index
        headers = {}
        for cell in ws[3]:
            if cell.value is not None:
                key = str(cell.value).strip().lower()
                headers[key] = cell.column

        # Resolve Student ID, Round N, Drive Link columns
        student_id_col = drive_link_col = round_col = None
        for key, col in headers.items():
            if 'student id' in key:
                student_id_col = col
            if f'round {round_num}' in key:
                round_col = col
            if 'drive link' in key:
                drive_link_col = col

        if not student_id_col:
            return jsonify({"success": False, "error": "Could not find 'Student ID' column in the Excel file."})
        if not round_col:
            return jsonify({"success": False, "error": f"Could not find 'Round {round_num}' column in the Excel file."})

        updated = 0
        errors  = []

        cursor.execute("SELECT company_name, role FROM jobs WHERE job_id = %s", (job_id,))
        job_det = cursor.fetchone()
        cname = job_det["company_name"] if job_det else f"Job {job_id}"
        rname = job_det["role"] if job_det else "Role"
        
        cursor.execute("SELECT num_rounds FROM recruitment_rounds WHERE job_id=%s", (job_id,))
        rr = cursor.fetchone()
        num_rounds = rr["num_rounds"] if rr else 1

        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_sid = row[student_id_col - 1]
            if raw_sid is None:
                continue
            try:
                student_id   = int(float(str(raw_sid)))
                round_value  = row[round_col - 1]
                drive_link   = row[drive_link_col - 1] if drive_link_col else None

                if round_value is None or str(round_value).strip() == "":
                    continue  # faculty left blank — skip

                rv = float(str(round_value).strip())
                if rv == 0:
                    result = "Not Selected"
                    drive_link = None          # clear link for eliminated students
                elif rv == round_num:
                    result = "Selected"
                else:
                    continue  # unexpected value — skip

                dl = str(drive_link).strip() if drive_link else None
                if dl in (None, "", "None"):
                    dl = None

                cursor.execute("SELECT result, drive_link FROM round_results WHERE job_id=%s AND student_id=%s AND round_number=%s", (job_id, student_id, round_num))
                prev = cursor.fetchone()
                prev_result = prev["result"] if prev else None
                prev_link = prev["drive_link"] if prev else None
                
                changed = (prev_result != result) or (str(prev_link or "").strip() != str(dl or "").strip())

                cursor.execute("""
                    INSERT INTO round_results (job_id, student_id, round_number, result, drive_link)
                    VALUES (%s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE result = VALUES(result), drive_link = VALUES(drive_link)
                """, (job_id, student_id, round_num, result, dl))
                updated += 1
                
                if changed:
                    # Update application status & Create Notification for Intermediate Rounds
                    if result == "Selected":
                        msg = f"🎉 You have been <strong>Qualified for Round {round_num}</strong> of <strong>{cname} - {rname}</strong>!"
                        cursor.execute("""
                            UPDATE applications SET status = 'Shortlisted'
                            WHERE student_id = %s AND job_id = %s AND status NOT IN ('Selected', 'Not Selected')
                        """, (student_id, job_id))
                    elif result == "Not Selected":
                        msg = f"Your application for <strong>{cname} - {rname}</strong> has been updated: <strong>Not Selected</strong> in Round {round_num}."
                        cursor.execute("""
                            UPDATE applications SET status = 'Not Selected'
                            WHERE student_id = %s AND job_id = %s
                        """, (student_id, job_id))
                    else:
                        msg = f"Your Round {round_num} status for <strong>{cname} - {rname}</strong> is being reviewed."
                        
                    if dl:
                        dl_lower = str(dl).lower()
                        if dl_lower.startswith('http://') or dl_lower.startswith('https://') or dl_lower.startswith('www.'):
                            msg += f" <br><strong>Attachment:</strong> <a href='{dl}' target='_blank'>View File/Drive Link</a>"
                        else:
                            msg += f" <br><strong>Message:</strong> {dl}"
                        
                    if round_num != num_rounds:
                        cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                                       (student_id, msg, "/my_applications"))

            except Exception as row_err:
                errors.append(str(row_err))

        # ── AUTO-FINALISE WHEN LAST ROUND IS UPLOADED ────────────────────────────
        if round_num == num_rounds:
            cursor.execute("""
                SELECT s.student_id, a.status
                FROM applications a
                JOIN students s ON a.student_id = s.student_id
                WHERE a.job_id = %s
            """, (job_id,))
            all_students = cursor.fetchall()
            app_status_map = {s["student_id"]: s["status"] for s in all_students}

            cursor.execute("SELECT student_id, round_number, result FROM round_results WHERE job_id=%s", (job_id,))
            all_results = cursor.fetchall()
            res_map = {}
            for r in all_results:
                res_map.setdefault(r["student_id"], {})[r["round_number"]] = r["result"]

            cursor.execute("SELECT tier FROM jobs WHERE job_id=%s", (job_id,))
            job_det_t   = cursor.fetchone()
            tier_str  = job_det_t["tier"] if job_det_t else "Tier 3"
            tier_num  = 1 if '1' in tier_str else (2 if '2' in tier_str else 3)

            for s in all_students:
                sid = s["student_id"]
                current_status = app_status_map.get(sid, "Pending")
                sres = res_map.get(sid, {})
                all_sel = all(sres.get(r, "Pending") == "Selected" for r in range(1, num_rounds + 1))
                any_not = any(sres.get(r, "Pending") == "Not Selected" for r in range(1, num_rounds + 1))
                
                final_msg = None
                if all_sel and current_status not in ('Selected', 'Rounds Cleared', 'Not Selected'):
                    cursor.execute(
                        "UPDATE applications SET status='Rounds Cleared', status_updated_at=NOW() WHERE student_id=%s AND job_id=%s",
                        (sid, job_id))
                    final_msg = f"🎉 Well done! You have cleared all <strong>{num_rounds} rounds</strong> for <strong>{cname} – {rname}</strong>. Your selection will be confirmed once the final results are published by the placement office."
                elif any_not and current_status not in ('Not Selected',):
                    cursor.execute(
                        "UPDATE applications SET status='Not Selected', status_updated_at=NOW() WHERE student_id=%s AND job_id=%s",
                        (sid, job_id))
                    final_msg = f"Your application for <strong>{cname} - {rname}</strong> has been updated: <strong>Not Selected</strong>."
                
                if final_msg:
                    cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)",
                                   (sid, final_msg, "/my_applications"))

        db.commit()
        return jsonify({"success": True, "updated": updated, "errors": errors,
                        "finalised": (round_num == num_rounds)})

    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})


@app.route("/faculty/job_results/<string:job_id>")
def faculty_job_results_api(job_id):
    ensure_connection()
    if "faculty_email" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    # Get selected
    cursor.execute("""
        SELECT s.student_id, s.name, s.email, s.branch, s.roll_number, s.phone_number
        FROM students s
        JOIN applications a ON s.student_id = a.student_id
        WHERE a.job_id = %s AND a.status = 'Selected'
        ORDER BY s.name ASC
    """, (job_id,))
    selected = cursor.fetchall()

    return jsonify({"selected": selected})


@app.route("/faculty/download_applied_excel/<string:job_id>")
def faculty_download_applied_excel(job_id):
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    # Fetch job company & role
    cursor.execute("SELECT company_name, role FROM jobs WHERE job_id = %s", (job_id,))
    job = cursor.fetchone()
    if not job:
        return "Job not found", 404

    company_name = job["company_name"]
    role_name = job["role"]

    # Fetch applied students
    cursor.execute("""
        SELECT s.student_id, s.name, s.branch, s.roll_number, s.phone_number, s.email, s.aadhar, s.pan, a.resume_path, a.extra_details
        FROM applications a
        JOIN students s ON a.student_id = s.student_id
        WHERE a.job_id = %s
        ORDER BY s.student_id ASC
    """, (job_id,))
    applicants = cursor.fetchall()
    
    import json
    for app_row in applicants:
        if app_row.get('extra_details'):
            try:
                app_row['extra_dict'] = json.loads(app_row['extra_details'])
            except:
                app_row['extra_dict'] = {}
        else:
            app_row['extra_dict'] = {}

    extra_keys = []
    for app_row in applicants:
        for k in app_row['extra_dict'].keys():
            if k not in extra_keys:
                extra_keys.append(k)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applied Students"

    # Set grid lines visible
    ws.views.sheetView[0].showGridLines = True

    # Styling colors
    amber_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="78350F")
    regular_font = Font(name="Calibri", size=11, bold=False)
    bold_font = Font(name="Calibri", size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Applied Students - {company_name} ({role_name}) - Job ID: {job_id}"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "S.No",
        "Student ID",
        "Roll Number",
        "Student Name",
        "Branch",
        "Aadhar Number",
        "PAN Number",
        "Email ID",
        "Phone Number",
        "Resume Drive Link"
    ]
    
    filtered_extra_keys = [k for k in extra_keys if k not in ("aadhar_number", "pan_number")]
    
    for k in filtered_extra_keys:
        headers.append(k.replace('_', ' ').title())
   
    # In openpyxl: A=1, B=2, C=3, etc. We will write to Row 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = white_font
        cell.fill = amber_fill
        cell.alignment = center_align
        cell.border = thin_border
   
    ws.row_dimensions[3].height = 24

    # Data Rows
    for r_idx, app in enumerate(applicants, 1):
        row_num = r_idx + 3
        ws.row_dimensions[row_num].height = 20

        # S.No
        c = ws.cell(row=row_num, column=1, value=r_idx)
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # Student ID
        c = ws.cell(row=row_num, column=2, value=app["student_id"])
        c.font = bold_font
        c.alignment = center_align
        c.border = thin_border

        # Roll Number
        c = ws.cell(row=row_num, column=3, value=app["roll_number"] or "—")
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # Name
        c = ws.cell(row=row_num, column=4, value=app["name"])
        c.font = regular_font
        c.alignment = left_align
        c.border = thin_border

        # Branch
        c = ws.cell(row=row_num, column=5, value=app["branch"])
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # Aadhar
        aadhar_val = app.get("aadhar") or app["extra_dict"].get("aadhar_number") or "—"
        c = ws.cell(row=row_num, column=6, value=aadhar_val)
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # PAN
        pan_val = app.get("pan") or app["extra_dict"].get("pan_number") or "—"
        c = ws.cell(row=row_num, column=7, value=pan_val)
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # Email
        c = ws.cell(row=row_num, column=8, value=app["email"])
        c.font = regular_font
        c.alignment = left_align
        c.border = thin_border

        # Phone
        c = ws.cell(row=row_num, column=9, value=app["phone_number"] or "—")
        c.font = regular_font
        c.alignment = center_align
        c.border = thin_border

        # Resume Drive Link
        resume_val = app["resume_path"] or "—"
        c = ws.cell(row=row_num, column=10, value=resume_val)
        c.font = regular_font
        c.alignment = left_align
        c.border = thin_border
        if resume_val and str(resume_val).startswith("http"):
            c.hyperlink = resume_val
            c.font = Font(name="Calibri", size=11, color="0000FF", underline="single")
            
        # Extra Details
        for i, key in enumerate(filtered_extra_keys):
            val = app['extra_dict'].get(key, "—")
            c = ws.cell(row=row_num, column=11+i, value=str(val))
            c.font = regular_font
            c.alignment = left_align
            c.border = thin_border

    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip merged cells title length for autofit width calculation
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            # If hyperlink, cap length calculation to avoid overly wide column
            if len(val_str) > 30 and cell.column == 10:
                val_str = "https://drive.google.com/..."
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # Save to buffer and send
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    from flask import send_file
    clean_company = "".join([c for c in company_name if c.isalnum() or c in (' ', '_', '-')]).strip()
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Applicants_{clean_company}_{job_id}.xlsx"
    )

@app.route("/faculty/student_details_download")
def faculty_student_details_download():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    cursor.execute("""
        SELECT
            roll_number as `Roll No`,
            name as `Name`,
            cgpa as `CGPA`,
            branch as `Branch`,
            backlogs as `Active Backlogs`,
            backlog_history as `Backlog History`,
            batch as `Graduation Year`,
            email as `College email ID`,
            phone_number as `Phone number`,
            alt_phone_number as `Alternate phone number`,
            tenth_score as `10th score`,
            inter_score as `12th score`,
            dob as `Date of Birth`,
            gender as `Gender`,
            category as `Category`,
            physically_challenged as `Physically challenged if yes how much percentage?`,
            pc_percentage as `PC Percentage`,
            internships_count as `Number of internships completed`,
            home_address as `Address`,
            jee_rank as `JEE rank`,
            academic_gap as `Academic year gap right from 10th`
        FROM students
    """)
    students_data = cursor.fetchall()
    
    # Process physically challenged column to append percentage if applicable
    for student in students_data:
        if student['Physically challenged if yes how much percentage?'] == 'Yes' and student['PC Percentage']:
            student['Physically challenged if yes how much percentage?'] = f"Yes ({student['PC Percentage']}%)"
        del student['PC Percentage']

    import pandas as pd
    import io
    from flask import send_file
    
    df = pd.DataFrame(students_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Student Details')
    
    output.seek(0)
    
    active_year = session.get("active_year", "2025-2026")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Student_Details_{active_year}.xlsx"
    )


# ─── FACULTY: STUDENTS ───────────────────────────────────────────────────────

PREV_YEARS_STATS = {
    "CSE":          {"2023": 94, "2024": 96, "2025": 92},
    "ECE":          {"2023": 86, "2024": 88, "2025": 85},
    "Mechanical":   {"2023": 78, "2024": 82, "2025": 80},
    "Chemical":     {"2023": 72, "2024": 75, "2025": 78},
    "Bio-Technology": {"2023": 70, "2024": 74, "2025": 76},
}

@app.route("/faculty/students")
def faculty_students():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    cursor.execute("SELECT branch, COUNT(*) as total FROM students GROUP BY branch")
    branch_counts = {r["branch"]: r["total"] for r in cursor.fetchall()}

    branch_stats = []
    for br in ["CSE", "ECE", "Mechanical", "Chemical", "Bio-Technology"]:
        total = branch_counts.get(br, 0)
        # Approximate placed count via applications
        cursor.execute("""
            SELECT COUNT(DISTINCT a.student_id) as cnt
            FROM applications a
            JOIN students s ON a.student_id = s.student_id
            WHERE s.branch = %s
        """, (br,))
        placed = cursor.fetchone()["cnt"]
        rate = round((placed / total) * 100) if total else 0
        branch_stats.append({"name": br, "total": total, "placed": placed, "rate": rate})

    return render_template("faculty/students.html", branch_stats=branch_stats, branch=None)


@app.route("/faculty/students/<branch_name>")
def faculty_branch_detail(branch_name):
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    cursor.execute("SELECT * FROM students WHERE branch = %s ORDER BY name", (branch_name,))
    students = cursor.fetchall()

    placed_count = 0
    for s in students:
        cursor.execute("SELECT * FROM applications WHERE student_id=%s", (s["student_id"],))
        s["applications"] = cursor.fetchall()
        if s["applications"]:
            placed_count += 1

    total = len(students)
    current_rate = round((placed_count / total) * 100) if total else 0
    prev_stats = PREV_YEARS_STATS.get(branch_name, {})

    return render_template(
        "faculty/students.html",
        branch=branch_name,
        students=students,
        placed_count=placed_count,
        current_rate=current_rate,
        prev_stats=prev_stats,
        branch_stats=[]
    )


# ─── FACULTY: MASTER SHEET ────────────────────────────────────────────────────
@app.route("/faculty/master_sheet")
def faculty_master_sheet():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    import os
    active_year = session.get("active_year", "2025-2026")
    upload_dir = os.path.join(app.static_folder, "uploads", active_year)
    current_file = None
    file_type = None
    file_size = 0

    if os.path.exists(os.path.join(upload_dir, "master_sheet.xlsx")):
        current_file = "master_sheet.xlsx"
        file_type = "Excel Document"
        file_size = round(os.path.getsize(os.path.join(upload_dir, current_file)) / 1024)
    elif os.path.exists(os.path.join(upload_dir, "master_sheet.pdf")):
        current_file = "master_sheet.pdf"
        file_type = "PDF Document"
        file_size = round(os.path.getsize(os.path.join(upload_dir, current_file)) / 1024)
    elif os.path.exists(os.path.join(upload_dir, "master_sheet.csv")):
        current_file = "master_sheet.csv"
        file_type = "CSV Document"
        file_size = round(os.path.getsize(os.path.join(upload_dir, current_file)) / 1024)

    active_year_name = get_active_batch_name()

    manual_stats_file = os.path.join(app.root_path, "database", "manual_stats.json")
    manual_pr = ""
    manual_lpa = ""
    if os.path.exists(manual_stats_file):
        try:
            with open(manual_stats_file, "r") as f:
                stats = json.load(f)
                if active_year in stats:
                    manual_pr = stats[active_year].get("placement_rate", "")
                    manual_lpa = stats[active_year].get("avg_lpa", "")
        except Exception:
            pass

    return render_template("faculty/master_sheet.html", current_file=current_file, file_type=file_type, file_size=file_size, active_year=active_year_name, manual_pr=manual_pr, manual_lpa=manual_lpa)

@app.route("/faculty/upload_master_sheet", methods=["POST"])
def faculty_upload_master_sheet():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    file = request.files.get("master_file")

    if not file or file.filename == "":
        flash("Please choose a file first.", "error")
        return redirect("/faculty/master_sheet")

    filename = secure_filename(file.filename)
    active_year = session.get("active_year", "2025-2026")
    upload_folder = os.path.join(app.static_folder, "uploads", active_year)
    os.makedirs(upload_folder, exist_ok=True)

    ext = filename.split('.')[-1].lower()
    master_name = f"master_sheet.{ext}"

    for ext_del in ["xlsx", "csv", "pdf"]:
        old_path = os.path.join(upload_folder, f"master_sheet.{ext_del}")
        if os.path.exists(old_path):
            try: os.remove(old_path)
            except: pass

    file_path = os.path.join(upload_folder, master_name)
    file.save(file_path)

    if ext == "xlsx":
        df = pd.read_excel(file_path)

    elif ext == "csv":
        df = pd.read_csv(file_path)

    elif ext == "pdf":
        all_rows = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()

                for table in tables:
                    for row in table:
                        all_rows.append(row)

        if not all_rows:
            return "No table found in PDF. Please upload a proper table PDF."

        headers = all_rows[0]
        data = all_rows[1:]

        df = pd.DataFrame(data, columns=headers)

    else:
        return "Only Excel, CSV, or PDF files are allowed"

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace(".", "", regex=False)
        .str.replace(" ", "_")
    )

    print("COLUMNS:", df.columns.tolist())

    updated = 0
    inserted = 0
    uploaded_emails = set()

    for index, row in df.iterrows():
        # Roll number: try 'Roll Number' (→ roll_number) or 'Register Number' (→ register_number)
        raw_roll = row.get("roll_number",
           row.get("roll_no",
           row.get("register_number",
           row.get("registration_number",
           row.get("reg_no", "")))))
        if pd.isna(raw_roll):
            roll_no = ""
        else:
            roll_no = str(raw_roll).strip()
            if roll_no.endswith('.0'):
                roll_no = roll_no[:-2]

        # Name: 'Name(in block letters)' normalizes to 'name(in_block_letters)'
        name = str(row.get("name(in_block_letters)", row.get("name", row.get("student_name", "")))).strip()
        email = str(
    row.get("email_address",
    row.get("institute_email",
    row.get("college_email",
    row.get("email",
    row.get("email_id",
    row.get("mail_id",
    row.get("institute_mail_id",
    row.get("official_email", ""))))))))
).strip().lower()
        # Branch: 'Specialization' or 'Branch' column
        branch = normalize_branch(str(row.get("branch", row.get("specialization", ""))).strip())

        if email == "" or email == "nan":
            continue
            
        # Try to infer roll number from email if not provided (e.g. 421235@student.nitandhra.ac.in)
        if not roll_no and '@' in email:
            inferred_roll = email.split('@')[0]
            if inferred_roll.isdigit():
                roll_no = inferred_roll

        uploaded_emails.add(email)

        def _safe_float(val):
            if pd.isna(val):
                return None
            val_str = str(val).strip()
            if not val_str:
                return None
            try:
                v = float(val_str)
                return v if v > 0 else None
            except (TypeError, ValueError):
                import re as _re
                m = _re.search(r'\d+(?:\.\d+)?', val_str)
                if m:
                    try:
                        v = float(m.group(0))
                        return v if v > 0 else None
                    except:
                        pass
                return None

        def _safe_int_val(val):
            if pd.isna(val):
                return 0
            val_str = str(val).strip()
            if not val_str:
                return 0
            try:
                return int(float(val_str))
            except (ValueError, TypeError):
                import re as _re
                m = _re.search(r'\d+', val_str)
                if m:
                    return int(m.group(0))
                return 0

        cgpa_val = _safe_float(row.get("cgpa", 0))
        cgpa = cgpa_val if cgpa_val is not None else 0.0
        # If student gives percentage for CGPA (e.g. 85.5), convert it to out of 10.
        if cgpa > 10.0 and cgpa <= 100.0:
            cgpa = round(cgpa / 10.0, 2)

        active_backlogs = _safe_int_val(row.get("active_backlogs", row.get("backlog_count", 0)))
        raw_backlog_hist = row.get("backlog_history", 0)
        if str(raw_backlog_hist).strip().lower() in ["yes", "y", "true"]:
            backlog_history = 1
        elif str(raw_backlog_hist).strip().lower() in ["no", "n", "false"]:
            backlog_history = 0
        else:
            backlog_history = _safe_int_val(raw_backlog_hist)

        # Try to get batch, fallback to parsing active_year (e.g. "2025-2026" -> 2026) or default 0
        raw_batch = row.get("graduation_batch", row.get("batch", None))
        if raw_batch and str(raw_batch).strip().lower() != "nan":
            batch = _safe_int_val(raw_batch)
        else:
            from flask import has_request_context
            active_year = "2025-2026"
            if has_request_context():
                active_year = session.get("active_year", "2025-2026")
            try:
                batch = int(active_year.split("-")[-1])
            except:
                batch = 0

        # '10th Score' normalizes to '10th_score'; '12th Score' → '12th_score'
        tenth_score_ms = _safe_float(row.get("10th_score", row.get("10th_%", row.get("tenth_score", None))))
        inter_score_ms = _safe_float(row.get("12th_score", row.get("12th_%", row.get("inter_score", None))))
        
        # If student gives CGPA for 10th/12th (e.g. 9.8), convert to percentage
        if tenth_score_ms is not None and tenth_score_ms <= 10.0 and tenth_score_ms > 0:
            tenth_score_ms = round(tenth_score_ms * 10.0, 2)
        if inter_score_ms is not None and inter_score_ms <= 10.0 and inter_score_ms > 0:
            inter_score_ms = round(inter_score_ms * 10.0, 2)
        # 'Phone Number' normalizes to 'phone_number'
        raw_phone = str(row.get("phone_number", row.get("phone_no", row.get("phone", "")))).strip()
        phone_ms = raw_phone if raw_phone and raw_phone != "nan" else None
        
        tier_1_ms = str(row.get("tier_1", row.get("tier 1", row.get("tier1", row.get("tier-1", ""))))).strip()
        tier_2_ms = str(row.get("tier_2", row.get("tier 2", row.get("tier2", row.get("tier-2", ""))))).strip()
        tier_3_ms = str(row.get("tier_3", row.get("tier 3", row.get("tier3", row.get("tier-3", ""))))).strip()
        tier_1_ms = tier_1_ms if tier_1_ms and tier_1_ms.lower() != "nan" else None
        tier_2_ms = tier_2_ms if tier_2_ms and tier_2_ms.lower() != "nan" else None
        tier_3_ms = tier_3_ms if tier_3_ms and tier_3_ms.lower() != "nan" else None

        selected_tier_ms = None
        if tier_1_ms: selected_tier_ms = 1
        elif tier_2_ms: selected_tier_ms = 2
        elif tier_3_ms: selected_tier_ms = 3

        def _get_str(keys):
            for k in keys:
                val = str(row.get(k, "")).strip()
                if val and val.lower() != "nan":
                    return val
            return None

        # Personal info fields — keys are the NORMALIZED column names from the Student Directory Data
        gender_ms = _get_str(["gender", "sex"])
        category_ms = _get_str(["category", "caste"])

        # 'PWD' column → normalized to 'pwd'
        physically_challenged_ms = _get_str(["pwd", "physically_challenged", "ph"])
        if physically_challenged_ms and physically_challenged_ms.lower() in ['yes', 'y', 'true']:
            physically_challenged_ms = 'Yes'
        else:
            physically_challenged_ms = 'No'

        # 'If Yes Reason with Percentage' → normalized to 'if_yes_reason_with_percentage'
        pwd_reason_raw = _get_str(["if_yes_reason_with_percentage", "pwd_percentage", "pc_percentage"])
        pc_percentage_ms = None
        if pwd_reason_raw:
            # Try to extract a number from the reason text (e.g. "40%" → 40.0)
            import re as _re
            nums = _re.findall(r'[\d]+(?:\.\d+)?', str(pwd_reason_raw))
            if nums:
                pc_percentage_ms = float(nums[0])

        internships_count_ms = int(_safe_float(row.get("internships_count", row.get("internships", 0))) or 0)

        # 'Address' column → normalized to 'address'
        home_address_ms = _get_str(["address", "home_address"])

        # 'JEE Rank' → normalized to 'jee_rank'
        jee_rank_ms = _get_str(["jee_rank", "jee_mains_rank", "jeerank", "jee", "jee_main_rank"])

        # 'Any Academic Gap' → normalized to 'any_academic_gap'
        academic_gap_ms = _get_str(["academic_year_gap", "any_academic_gap", "academic_gap", "gap_years"])

        # 'Alternate Phone Number' → normalized to 'alternate_phone_number'
        alt_phone_number_ms = _get_str(["alternate_phone_number", "alternate_phone_no", "alt_phone_number", "alternate_phone"])

        # 'Aadhar Number' → normalized to 'aadhar_number'
        aadhar_ms = _get_str(["aadhar_number", "aadhar_no", "aadhar"])

        # 'PAN Number' → normalized to 'pan_number'
        pan_ms = _get_str(["pan_number", "pan_no", "pan"])

        # 'Career Option(HE, Job, Others)' → normalized to 'career_option(he,_job,_others)'
        career_option_ms = _get_str(["career_option(he,_job,_others)", "career_option", "career"])

        # 'Personal Email' → normalized to 'personal_email'
        personal_email_ms = _get_str(["personal_email", "email_id_(_g-mail_or_other_)", "personal_email_id", "personal_mail"])
        
        # Handle date specifically to avoid invalid datetime DB errors
        # 'Date of Birth' → normalized to 'date_of_birth'
        dob_raw = _get_str(["date_of_birth", "dob"])
        dob_ms = None
        if dob_raw:
            try:
                # If it's passed as a pandas timestamp/datetime string
                parsed = pd.to_datetime(dob_raw)
                dob_ms = parsed.strftime('%Y-%m-%d')
            except Exception:
                dob_ms = None

        cursor.execute("SELECT * FROM students WHERE LOWER(email) = %s", (email,))
        existing_student = cursor.fetchone()

        if existing_student:
            cursor.execute("""
                UPDATE students
                SET roll_number = %s,
                    name = %s,
                    cgpa = %s,
                    branch = %s,
                    backlogs = %s,
                    backlog_history = %s,
                    batch = %s,
                    tenth_score = COALESCE(%s, tenth_score),
                    inter_score = COALESCE(%s, inter_score),
                    phone_number = COALESCE(%s, phone_number),
                    tier_1 = %s,
                    tier_2 = %s,
                    tier_3 = %s,
                    selected_tier = %s,
                    gender = COALESCE(%s, gender),
                    category = COALESCE(%s, category),
                    physically_challenged = %s,
                    pc_percentage = COALESCE(%s, pc_percentage),
                    internships_count = %s,
                    home_address = COALESCE(%s, home_address),
                    jee_rank = COALESCE(%s, jee_rank),
                    academic_gap = COALESCE(%s, academic_gap),
                    alt_phone_number = COALESCE(%s, alt_phone_number),
                    aadhar = COALESCE(%s, aadhar),
                    pan = COALESCE(%s, pan),
                    dob = COALESCE(%s, dob),
                    career_option = COALESCE(%s, career_option),
                    personal_email = COALESCE(%s, personal_email)
                WHERE LOWER(email) = %s
            """, (
                roll_no, name, cgpa, branch.upper(),
                active_backlogs, backlog_history, batch,
                tenth_score_ms, inter_score_ms, phone_ms,
                tier_1_ms, tier_2_ms, tier_3_ms, selected_tier_ms,
                gender_ms, category_ms, physically_challenged_ms, pc_percentage_ms,
                internships_count_ms, home_address_ms, jee_rank_ms, academic_gap_ms,
                alt_phone_number_ms, aadhar_ms, pan_ms, dob_ms,
                career_option_ms, personal_email_ms,
                email
            ))
            updated += 1



        else:
            cursor.execute("SELECT COALESCE(MAX(student_id), 0) + 1 AS next_id FROM students")
            next_id = cursor.fetchone()["next_id"]

            cursor.execute("""
                INSERT INTO students
                (student_id, roll_number, name, email, password, branch, cgpa,
                 backlogs, backlog_history, batch, skills,
                 must_change_password, tenth_score, inter_score, phone_number,
                 tier_1, tier_2, tier_3, selected_tier,
                 gender, category, physically_challenged, pc_percentage, internships_count,
                 home_address, jee_rank, academic_gap, alt_phone_number, aadhar, pan, dob,
                 career_option, personal_email)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                next_id, roll_no, name, email, roll_no, branch.upper(), cgpa,
                active_backlogs, backlog_history, batch,
                "", 1, tenth_score_ms, inter_score_ms, phone_ms,
                tier_1_ms, tier_2_ms, tier_3_ms, selected_tier_ms,
                gender_ms, category_ms, physically_challenged_ms, pc_percentage_ms, internships_count_ms,
                home_address_ms, jee_rank_ms, academic_gap_ms, alt_phone_number_ms, aadhar_ms, pan_ms, dob_ms,
                career_option_ms, personal_email_ms
            ))

            inserted += 1

        sid = existing_student["student_id"] if existing_student else next_id
        
        # ── Ensure applications are created or updated for tier selections ──
        for t_label, t_company in [("Tier 1", tier_1_ms), ("Tier 2", tier_2_ms), ("Tier 3", tier_3_ms)]:
            if t_company and str(t_company).strip() and str(t_company).strip().lower() != 'nan':
                cursor.execute("SELECT job_id, role, company_name FROM jobs WHERE LOWER(company_name) = LOWER(%s) AND tier = %s LIMIT 1", (str(t_company).strip(), t_label))
                job_found = cursor.fetchone()
                if job_found:
                    cursor.execute("SELECT status FROM applications WHERE student_id=%s AND job_id=%s", (sid, job_found["job_id"]))
                    app_exists = cursor.fetchone()
                    if app_exists:
                        if app_exists["status"] != "Selected":
                            cursor.execute("UPDATE applications SET status='Selected', status_updated_at=NOW() WHERE student_id=%s AND job_id=%s", (sid, job_found["job_id"]))
                            sel_msg = f"🏆 Congratulations! You have been <strong>Selected</strong> by <strong>{job_found['company_name']}</strong> for the <strong>{job_found['role']}</strong> role!"
                            cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)", (sid, sel_msg, "/my_applications"))
                    else:
                        cursor.execute("INSERT INTO applications (student_id, job_id, status) VALUES (%s, %s, 'Selected')", (sid, job_found["job_id"]))
                        sel_msg = f"🏆 Congratulations! You have been <strong>Selected</strong> by <strong>{job_found['company_name']}</strong> for the <strong>{job_found['role']}</strong> role!"
                        cursor.execute("INSERT INTO notifications (student_id, message, link) VALUES (%s, %s, %s)", (sid, sel_msg, "/my_applications"))

    deleted = 0
    if uploaded_emails:
        format_strings = ','.join(['%s'] * len(uploaded_emails))
        cursor.execute(f"SELECT student_id FROM students WHERE LOWER(email) NOT IN ({format_strings})", tuple(uploaded_emails))
        to_delete = cursor.fetchall()
        
        if to_delete:
            delete_ids = [s["student_id"] for s in to_delete]
            del_format = ','.join(['%s'] * len(delete_ids))
            
            cursor.execute(f"DELETE FROM round_results WHERE student_id IN ({del_format})", tuple(delete_ids))
            cursor.execute(f"DELETE FROM applications WHERE student_id IN ({del_format})", tuple(delete_ids))
            cursor.execute(f"DELETE FROM notifications WHERE student_id IN ({del_format})", tuple(delete_ids))
            cursor.execute(f"DELETE FROM students WHERE student_id IN ({del_format})", tuple(delete_ids))
            
            deleted = len(delete_ids)

    db.commit()

    print("UPDATED:", updated)
    print("INSERTED:", inserted)
    print("DELETED:", deleted)
    flash(f"Master Sheet uploaded successfully! {inserted} added, {updated} updated, and {deleted} removed.", "success")
    return redirect("/faculty/master_sheet")


@app.route("/faculty/download_master_sheet")
def faculty_download_master_sheet():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    from flask import send_file
    active_year = session.get("active_year", "2025-2026")
    upload_dir = os.path.join(app.static_folder, "uploads", active_year)
    pdf_path = os.path.join(upload_dir, "master_sheet.pdf")
    xlsx_path = os.path.join(upload_dir, "master_sheet.xlsx")
    csv_path = os.path.join(upload_dir, "master_sheet.csv")

    if os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True, download_name="master_sheet.pdf")
    elif os.path.exists(xlsx_path):
        return send_file(xlsx_path, as_attachment=True, download_name="master_sheet.xlsx")
    elif os.path.exists(csv_path):
        return send_file(csv_path, as_attachment=True, download_name="master_sheet.csv")
    else:
        from flask import flash
        flash("No Master Sheet file found.", "error")
        return redirect("/faculty/master_sheet")


@app.route("/faculty/delete_master_sheet", methods=["POST"])
def faculty_delete_master_sheet():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    from flask import flash
    active_year = session.get("active_year", "2025-2026")
    upload_dir = os.path.join(app.static_folder, "uploads", active_year)
    deleted = False
    for fname in ["master_sheet.pdf", "master_sheet.xlsx", "master_sheet.csv"]:
        fpath = os.path.join(upload_dir, fname)
        if os.path.exists(fpath):
            os.remove(fpath)
            deleted = True
    if deleted:
        flash("Master Sheet deleted successfully. Student portals are now disabled.", "success")
    else:
        flash("No Master Sheet found to delete.", "error")
    return redirect("/faculty/master_sheet")


@app.route("/faculty/upload_students", methods=["POST"])
def faculty_upload_students():
    ensure_connection()
    redir = faculty_required()
    if redir: return redir

    if "file" not in request.files:
        from flask import flash
        flash("No file uploaded.", "error")
        return redirect("/faculty/students")

    file = request.files["file"]
    filename = file.filename.lower()

    if filename.endswith(".xlsx"):
        try:
            import pandas as pd
            df = pd.read_excel(file)
            for index, row in df.iterrows():
                try:
                    row_dict = {str(k).lower().replace(' ', '_'): v for k, v in row.items()}
                    student_id = int(row_dict.get('student_id', 0))
                    name       = str(row_dict.get('name', ''))
                    email      = str(row_dict.get('email', ''))
                    password   = str(row_dict.get('password', email.split('@')[0] if email else 'default123'))
                    branch     = normalize_branch(str(row_dict.get('branch', '')))
                    cgpa       = float(row_dict.get('cgpa', 0.0))
                    backlogs   = int(row_dict.get('backlogs', 0))
                    skills     = str(row_dict.get('skills', ''))
                    batch      = int(row_dict.get('batch', 2028))
                    if student_id == 0 or not email:
                        continue
                    cursor.execute("SELECT * FROM students WHERE student_id=%s", (student_id,))
                    if cursor.fetchone():
                        cursor.execute("""UPDATE students SET name=%s,email=%s,branch=%s,cgpa=%s,backlogs=%s,skills=%s,batch=%s WHERE student_id=%s""",
                                       (name,email,branch,cgpa,backlogs,skills,batch,student_id))
                    else:
                        cursor.execute("""INSERT INTO students (student_id,name,email,password,branch,cgpa,backlogs,skills,batch) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                                       (student_id,name,email,password,branch,cgpa,backlogs,skills,batch))
                except Exception as row_err:
                    print(f"Row {index} error: {row_err}")
            db.commit()
            from flask import flash
            flash("Students updated successfully from Excel!", "success")
        except Exception as e:
            from flask import flash
            flash(f"Failed to process Excel: {str(e)}", "error")
    else:
        from flask import flash
        flash("Please upload a .xlsx Excel file.", "error")

    return redirect("/faculty/students")


# ─── FORGOT PASSWORD ─────────────────────────────────────────────────────────

def send_reset_otp(to_email, otp):
    """
    Send a password-reset OTP email using the configured SMTP credentials from .env.
    Falls back to logging the OTP to console if SMTP is not configured.
    """
    html = f"""
    <html>
      <body>
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
          <h2 style="color: #d97706; text-align: center;">Placement Portal — Password Reset</h2>
          <p>Hello,</p>
          <p>You have requested to reset your password. Please use the following One-Time Password (OTP) to proceed:</p>
          <div style="text-align: center; margin: 20px 0;">
            <span style="font-size: 24px; font-weight: bold; padding: 10px 20px; background-color: #fef3c7; border-radius: 5px; letter-spacing: 2px;">{otp}</span>
          </div>
          <p>This OTP is valid for <strong>60 seconds</strong>. Do not share it with anyone.</p>
          <p>If you did not request a password reset, please ignore this email.</p>
          <br>
          <p>Best regards,<br>NIT AP Placement Portal Team</p>
        </div>
      </body>
    </html>
    """

    # Use the shared email_service with credentials from .env
    if _EMAIL_SVC_LOADED and _email_svc.is_configured():
        subject = "Password Reset OTP — NIT AP Placement Portal"
        result = _email_svc.send_email(to_email, subject, html)
        if result['success']:
            print(f"[OTP] Reset OTP sent successfully to {to_email}")
            return True
        else:
            print(f"[OTP] Failed to send reset email to {to_email}: {result.get('error')}")
            return False
    else:
        # SMTP not configured — log OTP to console only (never expose in UI)
        print(f"[OTP] SMTP not configured. OTP for {to_email}: {otp}")
        return False

@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        if not email:
            from flask import flash
            flash("Please enter an email address.", "error")
            return redirect("/forgot_password")
        
        # Check students table
        cursor.execute("SELECT * FROM students WHERE email=%s", (email,))
        student = cursor.fetchone()
        
        # Check faculty table
        cursor.execute("SELECT * FROM faculty WHERE email=%s", (email,))
        faculty = cursor.fetchone()
        
        if not student and not faculty:
            from flask import flash
            flash("No account found with that email address.", "error")
            return redirect("/forgot_password")
            
        role = "student" if student else "faculty"
        
        if role == "student":
            aadhar = request.form.get("aadhar", "").strip()
            pan = request.form.get("pan", "").strip()
            if not aadhar or not pan:
                from flask import flash
                flash("Students must provide their Aadhar and PAN numbers for verification.", "error")
                return redirect("/forgot_password")
            
            if str(student.get("aadhar", "")).strip() != aadhar or str(student.get("pan", "")).strip().upper() != pan.upper():
                from flask import flash
                flash("Aadhar or PAN number does not match our records.", "error")
                return redirect("/forgot_password")
        
        # Generate 6 digit OTP
        import random
        import time
        otp = str(random.randint(100000, 999999))
        session['reset_email'] = email
        session['reset_role'] = role
        session['reset_otp'] = otp
        session['reset_otp_expires'] = time.time() + 60  # Valid for 60 seconds
        
        # Send OTP via email using configured SMTP credentials
        email_sent = send_reset_otp(email, otp)

        from flask import flash
        if email_sent:
            flash("A password reset OTP has been sent to your email address. Please check your inbox.", "info")
        else:
            # SMTP not configured — surface OTP only on localhost for developer convenience
            import socket
            if request.host.startswith("127.") or request.host.startswith("localhost"):
                flash(f"[Dev Mode] SMTP not configured. Your OTP is: {otp}", "info")
            else:
                flash("Failed to send OTP email. Please contact the Placement Cell.", "error")
                return redirect("/forgot_password")
        return redirect("/verify_otp")
       
    return render_template("forgot_password.html")

@app.route("/verify_otp", methods=["GET", "POST"])
def verify_otp():
    if 'reset_email' not in session:
        return redirect("/forgot_password")
       
    if request.method == "POST":
        entered_otp = request.form.get("otp", "").strip()
        role = session.get('reset_role')
        
        import time
        if time.time() > session.get('reset_otp_expires', 0):
            from flask import flash
            flash("OTP has expired (Valid for 1 min). Please request a new one.", "error")
            return redirect("/forgot_password")

        if entered_otp == session.get('reset_otp'):
            session['reset_verified'] = True
            return redirect("/reset_password")
        else:
            from flask import flash
            flash("Invalid OTP. Please try again.", "error")
           
    return render_template("verify_otp.html", email=session.get('reset_email'), role=session.get('reset_role'))

@app.route("/reset_password", methods=["GET", "POST"])
def reset_password():
    if not session.get('reset_verified') or 'reset_email' not in session:
        return redirect("/forgot_password")
       
    if request.method == "POST":
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
       
        if len(new_password) < 6:
            from flask import flash
            flash("Password must be at least 6 characters long.", "error")
            return redirect("/reset_password")
           
        if new_password != confirm_password:
            from flask import flash
            flash("Passwords do not match.", "error")
            return redirect("/reset_password")
           
        email = session['reset_email']
        role = session['reset_role']
       
        try:
            if role == "student":
                cursor.execute("UPDATE students SET password=%s WHERE email=%s", (new_password, email))
            else:
                cursor.execute("UPDATE faculty SET password=%s WHERE email=%s", (new_password, email))
            db.commit()
           
            # Clear session
            session.pop('reset_email', None)
            session.pop('reset_role', None)
            session.pop('reset_otp', None)
            session.pop('reset_verified', None)
           
            from flask import flash
            active_year_name = get_active_batch_name()
            flash(f"Password has been reset successfully for {active_year_name}! You can now log in.", "success")
            return redirect("/")
        except Exception as e:
            db.rollback()
            from flask import flash
            flash(f"An error occurred: {str(e)}", "error")
           
    return render_template("reset_password.html")

import threading
import time
from datetime import datetime

def start_reminder_scheduler():
    def reminder_loop():
        while True:
            try:
                import mysql.connector
                from email_service import send_email, SMTP_EMAIL, is_configured
                
                # Use same credentials as the rest of the app
                DB_HOST = "localhost"
                DB_USER = "root"
                DB_PASS = "Pallavi@2007"
                
                conn = mysql.connector.connect(host=DB_HOST, user=DB_USER, password=DB_PASS)
                c = conn.cursor(dictionary=True)
                c.execute("SHOW DATABASES LIKE 'placement_portal%'")
                dbs = c.fetchall()
                
                # Fetch faculty emails from the global database
                faculty_emails = ["tap@nitandhra.ac.in"]
                try:
                    c.execute("SELECT email FROM placement_portal.faculty WHERE email IS NOT NULL AND email != ''")
                    emails_db = c.fetchall()
                    if emails_db:
                        faculty_emails = [e['email'] for e in emails_db]
                except Exception:
                    pass
                conn.close()
                
                for d in dbs:
                    db_name = list(d.values())[0]
                    try:
                        b_conn = mysql.connector.connect(
                            host=DB_HOST, user=DB_USER, password=DB_PASS, database=db_name
                        )
                        b_cur = b_conn.cursor(dictionary=True)
                        # Only check if the reminder columns exist
                        try:
                            b_cur.execute("SELECT id, company_name, role, reminder_note FROM jobs WHERE reminder_date <= NOW() AND (reminder_sent IS NULL OR reminder_sent = 0)")
                            due_jobs = b_cur.fetchall()
                        except Exception:
                            b_conn.close()
                            continue
                        
                        if due_jobs and is_configured():
                            for job in due_jobs:
                                subject = f"[Placement Portal] Follow-up Reminder: {job['company_name']} ({job['role']})"
                                body = f"""
<h3>&#128276; Follow-up Reminder</h3>
<p>This is a scheduled reminder from the NIT AP Placement Portal.</p>
<table style='border-collapse:collapse;width:100%;font-size:14px;'>
  <tr><td style='padding:8px;background:#fff8e1;font-weight:700;color:#92400e;'>Company</td><td style='padding:8px;'>{job['company_name']}</td></tr>
  <tr><td style='padding:8px;background:#fff8e1;font-weight:700;color:#92400e;'>Role</td><td style='padding:8px;'>{job['role']}</td></tr>
  <tr><td style='padding:8px;background:#fff8e1;font-weight:700;color:#92400e;'>Note</td><td style='padding:8px;'>{job['reminder_note'] or 'No note provided.'}</td></tr>
</table>
<p>Please take the necessary action and contact the HR as needed.</p>
"""
                                success = False
                                for email in faculty_emails:
                                    res = send_email(email, subject, body)
                                    if res['success']:
                                        success = True
                                        
                                print(f"[Reminder] {job['company_name']}: {'Sent' if success else 'Failed'}")
                                if success:
                                    b_cur.execute("UPDATE jobs SET reminder_sent = 1 WHERE id = %s", (job['id'],))
                                    b_conn.commit()
                        b_conn.close()
                    except Exception as e:
                        print(f"[Reminder] Error for {db_name}: {e}")
            except Exception as e:
                print(f"[Reminder] Scheduler error: {e}")
            
            # Sleep for 1 minute
            time.sleep(60)

    thread = threading.Thread(target=reminder_loop, daemon=True)
    thread.start()

# Start the scheduler when the app starts
start_reminder_scheduler()

if __name__ == "__main__":
    app.run(debug=True)
